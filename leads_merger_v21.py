
# # ==============================================================================
# #  Leads <-> Sales Activity Merger + Summary Report  (GitHub Actions)
# #  v21 -- GHA compat, Google Sheets output, Grand Total, whole numbers,
# #         reordered summary columns.
# # ==============================================================================
# # Changes from v20:
# #  (1) SUMMARY_METRIC_COLS reordered to:
# #        Segment | Attempted | Company Answered | Our Answered | AI Qualified |
# #        Paid | AI Qual & Paid | AI Qual & Paid Rev | Answered & Paid |
# #        Ans & Paid Rev | Conv% (Attempted) | Conv% (Answered) |
# #        AI Unqual & Paid | AI Unqual & Paid Rev
# #  (2) Grand Total row added to Today's Summary + Overall Summary
# #      (both in the XLSX tabs AND in the Google Sheets write).
# #  (3) Whole numbers throughout:
# #        Conv% -> round to nearest integer % (no ".1" decimal)
# #        Revenue columns -> int (no ".0" suffix)
# #  (4) Lead Summary Report written to Google Sheet via service account:
# #        https://docs.google.com/spreadsheets/d/1rP_72DZtCWxRXlTuCsy1J38gaMJIeWn4FpH6fP1LBDE
# #        First sheet (gid=0) is cleared and rewritten each run.
# #  (5) GitHub Actions compatible:
# #        - Auth via GOOGLE_SERVICE_ACCOUNT_JSON env var (GitHub Secret)
# #        - No google.colab dependencies
# #        - Output XLSX files saved to ./output/ (uploadable as GHA artifact)
# #        - No Google Drive copy steps
# # ==============================================================================
# # requirements.txt:
# #   gspread>=5.0
# #   openpyxl
# #   pandas
# #   google-auth
# # ==============================================================================
# # GitHub workflow snippet:
# #   - name: Run leads merger
# #     env:
# #       GOOGLE_SERVICE_ACCOUNT_JSON: ${{ secrets.GOOGLE_SERVICE_ACCOUNT_JSON }}
# #     run: python leads_merger_v21.py
# # ==============================================================================

# import re, json, os
# from collections import Counter
# import pandas as pd
# import gspread
# import openpyxl
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from pathlib import Path
# from google.oauth2.service_account import Credentials

# # ── CallerAI Sheet ─────────────────────────────────────────────────────────────
# CALLER_AI_SHEET_ID = "1pv8h-66WFdjJA91PR6p3y35nrUB4GiAQ2HEmicOlAOA"

# TTC_UNC_TAB_NAME  = "TTC+Uncalled"
# TTC_UNC_SOURCE    = "TTC+Uncalled"

# TTC_TAB_NAME      = "TTC"
# UNCALLED_TAB_NAME = "Uncalled"
# TTC_SOURCE        = "TTC"

# # WINBACK_TAB_NAME  = "Winback"
# WINBACK_SOURCE    = "Winback"
# WINBACK_SHEETS = [
#     (CALLER_AI_SHEET_ID, "Winback"),              # original sheet — existing data
#     ("1EmTqHH5yfcrdk2QL58pXdru3bumnmuGB9rGbinyPmuQ", "Winback"), # new overflow sheet — add tab name too
# ]

# RENEWAL_TTC_TAB_NAME = "Renewal_TTC"
# RENEWAL_TTC_SOURCE   = "Renewal_TTC"

# WINBACK_DISPLAY      = "Winback"
# TTC_UNC_DISPLAY      = "TTC+Uncalled (Set 1)"
# TTC_DISPLAY          = "TTC+Uncalled (SET2)"
# RENEWAL_TTC_DISPLAY  = "Renewal_TTC"

# # v20: tiebreak order when all disposition/confidence/date signals are equal
# CAMP_CONCAT_ORDER = {
#     TTC_UNC_SOURCE:    0,
#     TTC_SOURCE:        1,
#     WINBACK_SOURCE:    2,
#     RENEWAL_TTC_SOURCE: 3,
# }

# OUTPUT_TAB_GID           = 1114202623
# CALLER_AI_DISP_COL       = "Business Disposition"
# CALLER_AI_SENTIMENT_COL  = "Business Disposition Confidence"
# CALLER_AI_QUALIFIED_MAP  = {
#     "interested": ["high", "medium", "low"],
#     "callback":   ["high", "medium"],
#     "followup":   ["high", "medium"],
# }
# CALLER_AI_DISP_ORDER = {"interested": 1, "callback": 2, "followup": 3}
# CALLER_AI_CONF_ORDER = {"high": 1, "medium": 2, "low": 3}

# DISP_NORMALIZE = {
#     "callback requested":   "callback",
#     "call back later":      "callback",
#     "call back":            "callback",
#     "called back later":    "callback",
#     "follow-up required":   "followup",
#     "followup required":    "followup",
#     "follow up required":   "followup",
#     "follow up":            "followup",
#     "follow-up":            "followup",
#     "information provided": "interested",
# }

# CALLERAI_DISCONNECT_COL   = "Disconnect Reason"
# CALLERAI_ANSWERED_REASONS = {
#     "user_request", "CLIENT_INITIATED", "inactivity_timeout", "transferred",
# }
# COMPANY_ANSWERED_REASONS  = {
#     "user_request", "CLIENT_INITIATED", "inactivity_timeout", "transferred",
#     "voicemail", "screening", "dnc_request",
# }

# LEADS_PHONE1_COL   = "Phone"
# LEADS_DATE_COL     = "Start Time"
# LEADS_CAMPAIGN_COL = "Call Type"
# LEADS_STATUS_COL   = "Outcome"
# LEADS_OWNER_COL    = "Agent Name"
# LEADS_NAME_COL     = "Participant"

# # v21: relative paths for GHA (no /content/ or Drive paths)
# OUTPUT_DIR           = Path("output")
# OUTPUT_LOCAL         = OUTPUT_DIR / "Leads_Final_Output.csv"
# SUMMARY_LOCAL        = OUTPUT_DIR / "Leads_Summary_Report.xlsx"
# SUMMARY_MATRIX_LOCAL = OUTPUT_DIR / "Leads_Summary_Matrix.xlsx"

# # SD_SHEET_ID    = "1Zf3GLNtI5nLsOfa8EBISqvYezhqgbAEuB6dR2isB4_E"
# # SD_SHEET_GID   = 758486581
# # All cumulative disposition sheets — keep in sync with DISPOSITION_CUMULATIVE_SHEET_IDS in ozontel.py
# # Format: (sheet_id, gid)  — Sheet 1 uses specific gid; overflow sheets use gid=0 (first tab)
# SD_SHEETS = [
#     ("1Zf3GLNtI5nLsOfa8EBISqvYezhqgbAEuB6dR2isB4_E", 758486581),  # Sheet 1 — original
#     ("1Qt-n-QYnDykALFjYphDIIXdhOFbd7_ppqr6ZhmvXpNo",0)
#     # ("PASTE_CUMULATIVE_SHEET_2_ID_HERE", 0),                      # Sheet 2 — add when you create it
# ]
# SD_PHONE_COL   = "Phone Number"
# SD_DATE_COL    = "Activity Date"
# SD_DISP1_COL   = "Disposition 1"
# SD_DISP2_COL   = "Disposition 2"
# SD_ADDED_BY    = "Activity Added By"
# SD_FILTER_FROM = "2026-04-01"

# SHEET_ID             = "1Kh2b4q5kqdy6oCvGlwArNphcLy6es8ZmHcrW_udvM3w"
# SHEET_GID            = 473339099
# NUMBER_COL_IDX       = 36
# PLAN_COL_IDX         = 16
# REVENUE_COL_IDX      = 28
# ORDER_VALUE_COL_IDX  = 14
# SALE_DATE_COL_IDX    = 15
# TEST_PMT_CHK_COL_IDX = 33

# TEST_PAYMENT_PLAN_NAMES = {
#     "Tax Expert Onboarding",
#     "Salaried Individual - IncomeTax Return Filing ( ITR )",
# }

# OUTCOME_ORDER = {
#     "Purchased":                 1,
#     "DND":                       2,
#     "Invalid":                   3,
#     "Not Interested":            4,
#     "Awaiting payment":          5,
#     "Will buy later":            6,
#     "Qualified":                 7,
#     "Trying to Contact":         8,
#     "Inbound Call":              9,
#     "AI Qualified":              10,
#     "Answered - No Disposition": 11,
#     "Not Contacted":             12,
# }

# DETAIL_TAB_PURCHASED  = "Purchased"
# DETAIL_TAB_AI_QUAL    = "AI Qualified"
# DETAIL_TAB_UNANSWERED = "Unanswered"

# # v21: Primary cols first (as specified), remaining cols after Conv%
# SUMMARY_METRIC_COLS = [
#     "Attempted",
#     "Company Answered",
#     "Our Answered",
#     "AI Qualified",
#     "Paid",
#     "AI Qual & Paid",
#     "AI Qual & Paid Rev",
#     "Answered & Paid",
#     "Ans & Paid Rev",
#     "Conv% (Attempted)",
#     "Conv% (Answered)",
#     "AI Unqual & Paid",
#     "AI Unqual & Paid Rev",
# ]

# PLAN_COLS = [
#     "Plan Name", "Total Revenue (Rs)", "Total Orders",
#     "Winback Revenue",     "Winback Orders",
#     "TTC+Unc Revenue",     "TTC+Unc Orders",
#     "TTC Revenue",         "TTC Orders",
#     "Renewal TTC Revenue", "Renewal TTC Orders",
# ]

# # v21: Target Google Sheet for Lead Summary Report
# OUTPUT_SUMMARY_SHEET_ID = "1rP_72DZtCWxRXlTuCsy1J38gaMJIeWn4FpH6fP1LBDE"


# # ══════════════════════════════════════════════════════════════════════════════
# # HELPERS
# # ══════════════════════════════════════════════════════════════════════════════
# def clean_number(raw):
#     if pd.isna(raw): return ""
#     digits = re.sub(r"\D", "", str(raw))
#     return digits[-10:] if len(digits) >= 10 else digits

# def _is_qualified(row):
#     disp = str(row[CALLER_AI_DISP_COL]).strip().lower()      if pd.notna(row.get(CALLER_AI_DISP_COL))      else ""
#     sent = str(row[CALLER_AI_SENTIMENT_COL]).strip().lower() if pd.notna(row.get(CALLER_AI_SENTIMENT_COL)) else ""
#     if sent == "":
#         return "Yes" if disp in CALLER_AI_QUALIFIED_MAP else "No"
#     return "Yes" if sent in CALLER_AI_QUALIFIED_MAP.get(disp, []) else "No"

# def assign_tab(row):
#     if str(row.get("Paid Check", "")) == "Paid":
#         return DETAIL_TAB_PURCHASED
#     if str(row.get("CallerAI Qualified", "No")) == "Yes":
#         return DETAIL_TAB_AI_QUAL
#     co = row.get("Calling Outcome", "")
#     if pd.isna(co) or str(co).strip() in ("", "(blank)"):
#         return DETAIL_TAB_UNANSWERED
#     return str(co).strip()

# def is_test_payment(plan_name):
#     pn = str(plan_name).strip().lower()
#     if pn.startswith("cleartax"):
#         return True
#     return pn in {n.strip().lower() for n in TEST_PAYMENT_PLAN_NAMES}

# def src_count_str(df):
#     w  = int((df["Source"] == WINBACK_SOURCE).sum())     if "Source" in df.columns else 0
#     tu = int((df["Source"] == TTC_UNC_SOURCE).sum())     if "Source" in df.columns else 0
#     t  = int((df["Source"] == TTC_SOURCE).sum())         if "Source" in df.columns else 0
#     r  = int((df["Source"] == RENEWAL_TTC_SOURCE).sum()) if "Source" in df.columns else 0
#     return f"Winback: {w}  |  TTC+Unc: {tu}  |  TTC: {t}  |  Renewal_TTC: {r}"


# # ══════════════════════════════════════════════════════════════════════════════
# # STEP 1 — Auth  (GitHub Actions: service account JSON from env var)
# # ══════════════════════════════════════════════════════════════════════════════
# _sa_json_str = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "")
# if not _sa_json_str:
#     raise EnvironmentError(
#         "GOOGLE_SERVICE_ACCOUNT_JSON env var is not set.\n"
#         "Add it as a GitHub Secret and expose it in your workflow:\n"
#         "  env:\n"
#         "    GOOGLE_SERVICE_ACCOUNT_JSON: ${{ secrets.GOOGLE_SERVICE_ACCOUNT_JSON }}"
#     )
# _sa_info = json.loads(_sa_json_str)
# # Normalise private key — some secret stores escape newlines as \\n
# if "private_key" in _sa_info:
#     _sa_info["private_key"] = _sa_info["private_key"].replace("\\n", "\n").strip("\r")
# _creds = Credentials.from_service_account_info(
#     _sa_info,
#     scopes=["https://www.googleapis.com/auth/spreadsheets"],
# )
# gc = gspread.authorize(_creds)
# print("OK Auth done (service account)")


# # ══════════════════════════════════════════════════════════════════════════════
# # STEP 2 — Load CallerAI Leads (4 campaigns)
# # ══════════════════════════════════════════════════════════════════════════════
# sh_caller = gc.open_by_key(CALLER_AI_SHEET_ID)
# tabs_map  = {ws.title: ws for ws in sh_caller.worksheets()}
# print(f"OK CallerAI sheet -- available tabs: {list(tabs_map.keys())}")

# def gsheet_tab_to_df(tabs_map, tab_name):
#     if tab_name not in tabs_map:
#         print(f"WARNING Tab '{tab_name}' not found. Available: {list(tabs_map.keys())}")
#         return pd.DataFrame()
#     rows = tabs_map[tab_name].get_all_values()
#     if not rows:
#         print(f"WARNING Tab '{tab_name}' is empty.")
#         return pd.DataFrame()
#     df = pd.DataFrame(rows[1:], columns=[c.strip() for c in rows[0]])
#     print(f"   Loaded '{tab_name}': {len(df):,} rows  |  cols: {list(df.columns)}")
#     return df

# print("\n-- Campaign 1: TTC+Uncalled (single tab) --------------------------------")
# df_ttc_unc_raw = gsheet_tab_to_df(tabs_map, TTC_UNC_TAB_NAME)

# print(f"\n-- Campaign 2: TTC (merge '{TTC_TAB_NAME}' + '{UNCALLED_TAB_NAME}') ------")
# df_ttc_tab = gsheet_tab_to_df(tabs_map, TTC_TAB_NAME)
# df_unc_tab = gsheet_tab_to_df(tabs_map, UNCALLED_TAB_NAME)
# df_ttc_raw = pd.concat([df_ttc_tab, df_unc_tab], ignore_index=True)
# print(f"   TTC merged raw: {len(df_ttc_raw):,}  "
#       f"({TTC_TAB_NAME}: {len(df_ttc_tab):,} + {UNCALLED_TAB_NAME}: {len(df_unc_tab):,})")

# print("\n-- Campaign 3: Winback (single tab) -------------------------------------")
# # df_win_raw = gsheet_tab_to_df(tabs_map, WINBACK_TAB_NAME)
# # AFTER:
# print(f"\n-- Campaign 3: Winback (merging {len(WINBACK_SHEETS)} sheets) -------------")
# _win_frames = []
# for _wb_sid, _wb_tab in WINBACK_SHEETS:
#     try:
#         _wb_sh   = gc.open_by_key(_wb_sid)
#         _wb_tabs = {ws.title: ws for ws in _wb_sh.worksheets()}
#         _wb_df   = gsheet_tab_to_df(_wb_tabs, _wb_tab)
#         _win_frames.append(_wb_df)
#     except Exception as _wb_err:
#         print(f"   WARNING: Could not load Winback sheet {_wb_sid}: {_wb_err}")
# df_win_raw = pd.concat([f for f in _win_frames if not f.empty], ignore_index=True)
# print(f"   Winback merged raw: {len(df_win_raw):,}  "
#       + "  |  ".join(f"{t}: {len(f):,}" for (_,t), f in zip(WINBACK_SHEETS, _win_frames)))

# print("\n-- Campaign 4: Renewal_TTC (single tab) ---------------------------------")
# df_renewal_ttc_raw = gsheet_tab_to_df(tabs_map, RENEWAL_TTC_TAB_NAME)


# # ══════════════════════════════════════════════════════════════════════════════
# # STEP 2b — Normalize, qualify, dedup each campaign independently
# # ══════════════════════════════════════════════════════════════════════════════
# def process_tab(df, source_name):
#     if df.empty:
#         return df
#     df = df.copy()
#     df["_source"] = source_name

#     if CALLER_AI_DISP_COL in df.columns:
#         df[CALLER_AI_DISP_COL] = (
#             df[CALLER_AI_DISP_COL].astype(str).str.strip().str.lower()
#             .map(lambda v: DISP_NORMALIZE.get(v, v))
#         )
#         _bd_str = df[CALLER_AI_DISP_COL].astype(str).str.strip().str.lower()
#         _empty  = _bd_str.isin(["", "nan", "none"])
#         if _empty.any() and LEADS_STATUS_COL in df.columns:
#             _fallback = (df.loc[_empty, LEADS_STATUS_COL]
#                          .astype(str).str.strip().str.lower()
#                          .map(lambda v: DISP_NORMALIZE.get(v, v)))
#             df.loc[_empty, CALLER_AI_DISP_COL] = _fallback
#             print(f"   {source_name} -- {_empty.sum()} empty Business Dispositions filled from Lead Stage")

#         df["CallerAI Qualified"] = df.apply(_is_qualified, axis=1)
#         q = (df["CallerAI Qualified"] == "Yes").sum()
#         print(f"   {source_name} -- CallerAI Qualified (pre-dedup): {q:,}")

#         unique_disps = df[CALLER_AI_DISP_COL].dropna().unique()
#         print(f"   {source_name} -- Unique dispositions: {sorted(str(d) for d in unique_disps)}")
#     else:
#         df["CallerAI Qualified"] = "No"
#         print(f"   {source_name} -- Disposition col missing; Qualified=No for all")

#     df["_clean_phone"] = (
#         df[LEADS_PHONE1_COL].apply(clean_number) if LEADS_PHONE1_COL in df.columns else ""
#     )

#     if LEADS_DATE_COL in df.columns:
#         _cleaned       = df[LEADS_DATE_COL].astype(str).str.replace(r'\s*\([^)]*\)\s*$', '', regex=True).str.strip()
#         _parsed        = pd.to_datetime(_cleaned, utc=True, errors="coerce")
#         df["_date_ts"] = _parsed.dt.tz_convert(None)
#     else:
#         df["_date_ts"] = pd.NaT

#     pre_dedup = len(df)
#     df["_disp_sort"] = (df[CALLER_AI_DISP_COL].astype(str).str.strip().str.lower()
#                         .map(CALLER_AI_DISP_ORDER).fillna(999)) \
#                         if CALLER_AI_DISP_COL in df.columns else 999
#     df["_conf_sort"] = (df[CALLER_AI_SENTIMENT_COL].astype(str).str.strip().str.lower()
#                         .map(CALLER_AI_CONF_ORDER).fillna(999)) \
#                         if CALLER_AI_SENTIMENT_COL in df.columns else 999
#     df["_date_sort"] = df["_date_ts"].fillna(pd.Timestamp.min)

#     has_phone = df["_clean_phone"] != ""
#     with_ph = (
#         df[has_phone]
#         .sort_values(["_clean_phone", "_disp_sort", "_conf_sort", "_date_sort"],
#                      ascending=[True, True, True, False])
#         .drop_duplicates(subset=["_clean_phone"], keep="first")
#     )
#     df = pd.concat([with_ph, df[~has_phone]], ignore_index=True)
#     df.drop(columns=["_disp_sort", "_conf_sort", "_date_sort"], inplace=True)
#     print(f"   {source_name} -- dedup: {pre_dedup:,} -> {len(df):,}  (removed {pre_dedup - len(df):,})")
#     return df

# print("\n-- Processing campaigns -------------------------------------------------")
# df_ttc_unc     = process_tab(df_ttc_unc_raw,     TTC_UNC_SOURCE)
# df_ttc         = process_tab(df_ttc_raw,         TTC_SOURCE)
# df_win         = process_tab(df_win_raw,         WINBACK_SOURCE)
# df_renewal_ttc = process_tab(df_renewal_ttc_raw, RENEWAL_TTC_SOURCE)

# leads_raw = pd.concat([df_ttc_unc, df_ttc, df_win, df_renewal_ttc], ignore_index=True)
# print(f"\nOK Combined -- total: {len(leads_raw):,}  |  "
#       f"{TTC_UNC_SOURCE}: {len(df_ttc_unc):,}  |  "
#       f"{TTC_SOURCE}: {len(df_ttc):,}  |  "
#       f"{WINBACK_SOURCE}: {len(df_win):,}  |  "
#       f"{RENEWAL_TTC_SOURCE}: {len(df_renewal_ttc):,}")


# # ══════════════════════════════════════════════════════════════════════════════
# # STEP 3 — Paid Sheet
# # v20: skip zero-revenue rows entirely (treat as test payments)
# # ══════════════════════════════════════════════════════════════════════════════
# sh_paid  = gc.open_by_key(SHEET_ID)
# ws_paid  = next((w for w in sh_paid.worksheets() if w.id == SHEET_GID), sh_paid.worksheets()[0])
# all_rows = ws_paid.get_all_values()

# _today_ts = pd.Timestamp.today().normalize()

# paid_set             = set()
# paid_set_today       = set()
# paid_plan_map        = {}
# skipped_cleartax     = 0
# skipped_not_genuine  = 0
# skipped_zero_revenue = 0

# for row in all_rows[1:]:
#     if len(row) <= NUMBER_COL_IDX or not str(row[NUMBER_COL_IDX]).strip():
#         continue
#     phone = clean_number(row[NUMBER_COL_IDX])
#     if not phone:
#         continue

#     plan = row[PLAN_COL_IDX].strip() if len(row) > PLAN_COL_IDX else ""
#     if is_test_payment(plan):
#         skipped_cleartax += 1
#         continue

#     test_chk = row[TEST_PMT_CHK_COL_IDX].strip() if len(row) > TEST_PMT_CHK_COL_IDX else ""
#     if test_chk.lower() != "genuine":
#         skipped_not_genuine += 1
#         continue

#     rev_raw = row[REVENUE_COL_IDX].strip() if len(row) > REVENUE_COL_IDX else ""
#     try:    rev = float(re.sub(r"[^\d.]", "", rev_raw)) if rev_raw else 0.0
#     except: rev = 0.0

#     if rev == 0.0:
#         ov_raw = row[ORDER_VALUE_COL_IDX].strip() if len(row) > ORDER_VALUE_COL_IDX else ""
#         try:    rev = float(re.sub(r"[^\d.]", "", ov_raw)) if ov_raw else 0.0
#         except: rev = 0.0

#     if rev == 0.0:
#         skipped_zero_revenue += 1
#         continue

#     paid_set.add(phone)

#     if phone not in paid_plan_map or (not paid_plan_map[phone]["plan_name"] and plan):
#         paid_plan_map[phone] = {"plan_name": plan, "revenue": rev}

#     raw_sale = row[SALE_DATE_COL_IDX].strip() if len(row) > SALE_DATE_COL_IDX else ""
#     if raw_sale:
#         _sc = re.sub(r'\s*\([^)]*\)\s*$', '', raw_sale).strip()
#         _st = pd.to_datetime(_sc, utc=True, errors="coerce")
#         if pd.notna(_st) and _st.tz_convert(None).normalize() == _today_ts:
#             paid_set_today.add(phone)

# paid_set.discard("")
# paid_set_today.discard("")
# print(f"OK Paid sheet -- unique paid: {len(paid_set):,}  |  today: {len(paid_set_today):,}  "
#       f"|  ClearTax/test skipped: {skipped_cleartax:,}  |  Not Genuine: {skipped_not_genuine:,}  "
#       f"|  Zero revenue: {skipped_zero_revenue:,}")


# # ══════════════════════════════════════════════════════════════════════════════
# # STEP 4 — LSQ Sales Disposition
# # ══════════════════════════════════════════════════════════════════════════════
# # _disp_url = (f"https://docs.google.com/spreadsheets/d/{SD_SHEET_ID}"
# #              f"/export?format=csv&gid={SD_SHEET_GID}")
# # df_sd = pd.read_csv(_disp_url, low_memory=False)
# _sd_frames = []
# for _sd_id, _sd_gid in SD_SHEETS:
#     _disp_url = (f"https://docs.google.com/spreadsheets/d/{_sd_id}"
#                  f"/export?format=csv&gid={_sd_gid}")
#     try:
#         _df_chunk = pd.read_csv(_disp_url, low_memory=False)
#         _df_chunk.columns = _df_chunk.columns.str.strip()
#         _sd_frames.append(_df_chunk)
#         print(f"   Loaded SD sheet {_sd_id}: {len(_df_chunk):,} rows")
#     except Exception as _sd_err:
#         print(f"   WARNING: Could not load SD sheet {_sd_id}: {_sd_err}")

# if not _sd_frames:
#     raise RuntimeError("No cumulative disposition sheets could be loaded.")
# df_sd = pd.concat(_sd_frames, ignore_index=True)
# df_sd.columns = df_sd.columns.str.strip()
# if SD_ADDED_BY not in df_sd.columns:
#     raise KeyError(f"'{SD_ADDED_BY}' column missing.")

# df_sd["_act_date"]    = pd.to_datetime(df_sd[SD_DATE_COL], dayfirst=True, errors="coerce")
# df_sd = df_sd[df_sd["_act_date"] >= pd.Timestamp(SD_FILTER_FROM)].copy()
# df_sd["_clean_phone"] = df_sd[SD_PHONE_COL].apply(clean_number)
# df_sd = df_sd[df_sd["_clean_phone"] != ""].copy()

# d1 = df_sd[["_clean_phone", SD_DISP1_COL, SD_ADDED_BY]].rename(
#     columns={SD_DISP1_COL: "_disp", SD_ADDED_BY: "_added_by"})
# d2 = df_sd[["_clean_phone", SD_DISP2_COL, SD_ADDED_BY]].rename(
#     columns={SD_DISP2_COL: "_disp", SD_ADDED_BY: "_added_by"})
# all_disps = pd.concat([d1, d2], ignore_index=True)
# all_disps = all_disps[all_disps["_disp"].notna() & (all_disps["_disp"].str.strip() != "")].copy()
# all_disps["_order"] = all_disps["_disp"].map(OUTCOME_ORDER).fillna(999)
# best_disp = (
#     all_disps.sort_values("_order")
#     .groupby("_clean_phone").first().reset_index()
#     [["_clean_phone", "_disp", "_added_by"]]
#     .rename(columns={"_clean_phone": "_merge_phone", "_disp": "Calling Outcome",
#                      "_added_by": "_lsq_agent"})
# )
# print(f"OK LSQ loaded -- {len(df_sd):,} rows  |  {best_disp['_merge_phone'].nunique():,} unique phones")


# # ══════════════════════════════════════════════════════════════════════════════
# # STEP 5 — Build merged DataFrame
# # ══════════════════════════════════════════════════════════════════════════════
# def safe_col(df, col, fallback=""):
#     return df[col].values if col in df.columns else fallback

# paid_check = ["Paid" if (n and n in paid_set) else "Unpaid"
#               for n in leads_raw["_clean_phone"].values]

# out = pd.DataFrame({
#     "Date":                            pd.Series(leads_raw["_date_ts"].values).dt.strftime("%d/%m/%Y").fillna("").values,
#     "Customer Number":                 leads_raw["_clean_phone"].values,
#     "Lead Name":                       safe_col(leads_raw, LEADS_NAME_COL),
#     "Type Of Lead":                    safe_col(leads_raw, LEADS_CAMPAIGN_COL),
#     "Lead Stage":                      safe_col(leads_raw, LEADS_STATUS_COL),
#     "Disconnect Reason":               safe_col(leads_raw, CALLERAI_DISCONNECT_COL),
#     "Lead Owner":                      safe_col(leads_raw, LEADS_OWNER_COL),
#     "Paid Check":                      paid_check,
#     "CallerAI Qualified":              leads_raw["CallerAI Qualified"].values,
#     "Business Disposition":            safe_col(leads_raw, CALLER_AI_DISP_COL),
#     "Business Disposition Confidence": safe_col(leads_raw, CALLER_AI_SENTIMENT_COL),
#     "Source":                          leads_raw["_source"].values,
#     "_date_ts":                        pd.to_datetime(leads_raw["_date_ts"].values).normalize(),
# })

# out = out.merge(best_disp, left_on="Customer Number", right_on="_merge_phone", how="left")
# out.drop(columns=["_merge_phone"], inplace=True, errors="ignore")

# out.loc[out["Paid Check"] == "Paid", "Calling Outcome"] = "Purchased"
# out.loc[(out["Calling Outcome"] == "Purchased") & (out["Paid Check"] == "Unpaid"), "Calling Outcome"] = pd.NA

# out["Activity Added By"] = out.apply(
#     lambda row: row["_lsq_agent"]
#                 if pd.notna(row["_lsq_agent"]) and str(row["_lsq_agent"]).strip() != ""
#                 else row.get("Lead Owner", ""), axis=1)
# out.drop(columns=["_lsq_agent"], inplace=True)

# out["Plan Name"] = out["Customer Number"].map(lambda p: paid_plan_map.get(p, {}).get("plan_name", ""))
# out["Revenue"]   = out["Customer Number"].map(lambda p: paid_plan_map.get(p, {}).get("revenue", 0.0))
# out["Type Of Lead"] = out["Type Of Lead"].fillna("(blank)").astype(str).str.strip().replace("", "(blank)")
# out["Lead Stage"]   = out["Lead Stage"].fillna("(blank)").astype(str).str.strip().replace("", "(blank)")
# out["Business Disposition Confidence"] = out["Business Disposition Confidence"].fillna("").astype(str).str.strip()
# out["Business Disposition"]            = out["Business Disposition"].fillna("").astype(str).str.strip()

# out = out[[
#     "Date", "Customer Number", "Lead Name", "Type Of Lead", "Source",
#     "Lead Stage", "Disconnect Reason", "Business Disposition",
#     "CallerAI Qualified", "Business Disposition Confidence",
#     "Calling Outcome", "Paid Check", "Lead Owner", "Activity Added By",
#     "Plan Name", "Revenue", "_date_ts"
# ]]

# paid_c    = (out["Paid Check"] == "Paid").sum()
# ai_qual_c = (out["CallerAI Qualified"] == "Yes").sum()
# purch_c   = (out["Calling Outcome"] == "Purchased").sum()
# print(f"OK Merge done -- {len(out):,} leads  |  Paid: {paid_c}  |  Purchased: {purch_c}")
# print(f"   CallerAI Qualified (all):        {ai_qual_c:,}")
# print(f"   CallerAI Qualified (excl. Paid): "
#       f"{int(((out['CallerAI Qualified']=='Yes') & (out['Paid Check']!='Paid')).sum()):,}")
# print(f"   {TTC_UNC_SOURCE}: {(out['Source']==TTC_UNC_SOURCE).sum():,}  |  "
#       f"{TTC_SOURCE}: {(out['Source']==TTC_SOURCE).sum():,}  |  "
#       f"{WINBACK_SOURCE}: {(out['Source']==WINBACK_SOURCE).sum():,}  |  "
#       f"{RENEWAL_TTC_SOURCE}: {(out['Source']==RENEWAL_TTC_SOURCE).sum():,}")


# # ══════════════════════════════════════════════════════════════════════════════
# # STEP 6 — Save CSV (kept as GHA artifact; no Drive copy)
# # ══════════════════════════════════════════════════════════════════════════════
# OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
# out.drop(columns=["_date_ts"]).to_csv(OUTPUT_LOCAL, index=False)
# print(f"OK CSV -> {OUTPUT_LOCAL}")


# # ══════════════════════════════════════════════════════════════════════════════
# # STEP 7 — Build summary data + XLSX report
# # ══════════════════════════════════════════════════════════════════════════════
# today_ts  = _today_ts
# today_str = today_ts.strftime("%d-%m-%Y")

# out_all = out.copy()

# # ── v20: Revenue Attribution ──────────────────────────────────────────────────
# _paid_mask_all = out_all["Paid Check"] == "Paid"
# if _paid_mask_all.any():
#     _paid_for_attr = out_all[_paid_mask_all].copy()
#     _paid_for_attr["_disp_s"] = (_paid_for_attr["Business Disposition"]
#                                   .str.strip().str.lower()
#                                   .map(CALLER_AI_DISP_ORDER).fillna(999))
#     _paid_for_attr["_conf_s"] = (_paid_for_attr["Business Disposition Confidence"]
#                                   .str.strip().str.lower()
#                                   .map(CALLER_AI_CONF_ORDER).fillna(999))
#     _paid_for_attr["_date_s"] = _paid_for_attr["_date_ts"].fillna(pd.Timestamp.min)
#     _paid_for_attr["_camp_s"] = _paid_for_attr["Source"].map(CAMP_CONCAT_ORDER).fillna(99)

#     _rev_owner_df = (
#         _paid_for_attr
#         .sort_values(
#             ["Customer Number", "_disp_s", "_conf_s", "_date_s", "_camp_s"],
#             ascending=[True, True, True, False, True]
#         )
#         .drop_duplicates(subset=["Customer Number"], keep="first")
#     )
#     _rev_owner = dict(zip(_rev_owner_df["Customer Number"], _rev_owner_df["Source"]))
# else:
#     _rev_owner = {}

# out_all["Revenue_Attributed"] = out_all.apply(
#     lambda row: row["Revenue"]
#     if (row["Paid Check"] == "Paid"
#         and _rev_owner.get(row["Customer Number"]) == row["Source"])
#     else 0.0,
#     axis=1
# )

# if _rev_owner:
#     _attr_counts    = Counter(_rev_owner.values())
#     _attr_total_rev = out_all["Revenue_Attributed"].sum()
#     print(f"OK Revenue attribution -- {len(_rev_owner):,} paid phones attributed  |  "
#           + "  |  ".join(
#               f"{src}: {_attr_counts.get(src, 0)}"
#               for src in [TTC_UNC_SOURCE, TTC_SOURCE, WINBACK_SOURCE, RENEWAL_TTC_SOURCE])
#           + f"  |  Total attributed Rs: {_attr_total_rev:,.0f}")

# # Per-campaign DFs (built AFTER Revenue_Attributed so they inherit the column)
# out_ttc_unc     = out_all[out_all["Source"] == TTC_UNC_SOURCE].copy()
# out_ttc         = out_all[out_all["Source"] == TTC_SOURCE].copy()
# out_win         = out_all[out_all["Source"] == WINBACK_SOURCE].copy()
# out_renewal_ttc = out_all[out_all["Source"] == RENEWAL_TTC_SOURCE].copy()

# # Today's per-campaign DFs
# out_ttc_unc_today     = out_ttc_unc    [out_ttc_unc    ["_date_ts"] == today_ts].copy()
# out_ttc_today         = out_ttc        [out_ttc        ["_date_ts"] == today_ts].copy()
# out_win_today         = out_win        [out_win        ["_date_ts"] == today_ts].copy()
# out_renewal_ttc_today = out_renewal_ttc[out_renewal_ttc["_date_ts"] == today_ts].copy()


# # ── build_summary_row ─────────────────────────────────────────────────────────
# def build_summary_row(df, label, today_paid_set=None):
#     total = len(df)

#     if CALLERAI_DISCONNECT_COL in df.columns:
#         _dr = df[CALLERAI_DISCONNECT_COL].fillna("").str.strip().str.lower()
#         company_answered = int(_dr.isin({r.lower() for r in COMPANY_ANSWERED_REASONS}).sum())
#         our_answered     = int(_dr.isin({r.lower() for r in CALLERAI_ANSWERED_REASONS}).sum())
#         _is_our_answered = _dr.isin({r.lower() for r in CALLERAI_ANSWERED_REASONS})
#     else:
#         our_answered = company_answered = int(df["Calling Outcome"].notna().sum())
#         _is_our_answered = df["Calling Outcome"].notna()

#     if today_paid_set is not None:
#         _is_paid = df["Customer Number"].isin(today_paid_set)
#     else:
#         _is_paid = df["Paid Check"] == "Paid"

#     paid           = int(_is_paid.sum())
#     ai_qual_paid   = int((_is_paid & (df["CallerAI Qualified"] == "Yes")).sum())
#     ai_unqual_paid = int((_is_paid & (df["CallerAI Qualified"] != "Yes")).sum())
#     answered_paid  = int((_is_paid & _is_our_answered).sum())
#     ai_q           = int((df["CallerAI Qualified"] == "Yes").sum())

#     # v20: Revenue_Attributed already deduplicated — sum is safe
#     if "Revenue_Attributed" in df.columns:
#         _rev_a = pd.to_numeric(df["Revenue_Attributed"], errors="coerce").fillna(0)
#         ai_qual_paid_rev   = int(_rev_a[_is_paid & (df["CallerAI Qualified"] == "Yes")].sum())
#         ai_unqual_paid_rev = int(_rev_a[_is_paid & (df["CallerAI Qualified"] != "Yes")].sum())
#         ans_paid_rev       = int(_rev_a[_is_paid & _is_our_answered].sum())
#     else:
#         ai_qual_paid_rev = ai_unqual_paid_rev = ans_paid_rev = 0

#     # v21: whole-number percentages (no ".1" decimal)
#     conv_att = f"{round(paid / total        * 100)}%" if total        else "0%"
#     conv_ans = f"{round(paid / our_answered * 100)}%" if our_answered else "0%"

#     return {
#         "Segment":              label,
#         "Attempted":            total,
#         "Company Answered":     company_answered,
#         "Our Answered":         our_answered,
#         "AI Qualified":         ai_q,
#         "Paid":                 paid,
#         "AI Qual & Paid":       ai_qual_paid,
#         "AI Unqual & Paid":     ai_unqual_paid,
#         "Answered & Paid":      answered_paid,
#         "AI Qual & Paid Rev":   ai_qual_paid_rev,   # v21: int
#         "AI Unqual & Paid Rev": ai_unqual_paid_rev, # v21: int
#         "Ans & Paid Rev":       ans_paid_rev,       # v21: int
#         "Conv% (Attempted)":    conv_att,
#         "Conv% (Answered)":     conv_ans,
#     }


# # v21: Grand Total row — sums numeric cols, recomputes Conv% from totals
# def build_grand_total_row(rows):
#     def _sum(key):
#         return sum(r.get(key, 0) for r in rows
#                    if isinstance(r.get(key, 0), (int, float)))

#     total_att  = _sum("Attempted")
#     total_our  = _sum("Our Answered")
#     total_paid = _sum("Paid")

#     conv_att = f"{round(total_paid / total_att  * 100)}%" if total_att  else "0%"
#     conv_ans = f"{round(total_paid / total_our  * 100)}%" if total_our  else "0%"

#     return {
#         "Segment":              "Grand Total",
#         "Attempted":            total_att,
#         "Company Answered":     _sum("Company Answered"),
#         "Our Answered":         total_our,
#         "AI Qualified":         _sum("AI Qualified"),
#         "Paid":                 total_paid,
#         "AI Qual & Paid":       _sum("AI Qual & Paid"),
#         "AI Unqual & Paid":     _sum("AI Unqual & Paid"),
#         "Answered & Paid":      _sum("Answered & Paid"),
#         "AI Qual & Paid Rev":   _sum("AI Qual & Paid Rev"),
#         "AI Unqual & Paid Rev": _sum("AI Unqual & Paid Rev"),
#         "Ans & Paid Rev":       _sum("Ans & Paid Rev"),
#         "Conv% (Attempted)":    conv_att,
#         "Conv% (Answered)":     conv_ans,
#     }


# summary_today = [
#     build_summary_row(out_win_today,         WINBACK_DISPLAY,     today_paid_set=paid_set_today),
#     build_summary_row(out_ttc_unc_today,     TTC_UNC_DISPLAY,     today_paid_set=paid_set_today),
#     build_summary_row(out_ttc_today,         TTC_DISPLAY,         today_paid_set=paid_set_today),
#     build_summary_row(out_renewal_ttc_today, RENEWAL_TTC_DISPLAY, today_paid_set=paid_set_today),
# ]
# summary_overall = [
#     build_summary_row(out_win,         WINBACK_DISPLAY),
#     build_summary_row(out_ttc_unc,     TTC_UNC_DISPLAY),
#     build_summary_row(out_ttc,         TTC_DISPLAY),
#     build_summary_row(out_renewal_ttc, RENEWAL_TTC_DISPLAY),
# ]


# # ── Pivot helpers ──────────────────────────────────────────────────────────────
# def campaign_pivot(df):
#     if df.empty:
#         return pd.DataFrame(columns=["Paid", "Unpaid", "Grand Total", "%", "AI Qualified"])
#     df = df.copy()
#     df["Type Of Lead"] = df["Type Of Lead"].fillna("(blank)").astype(str).str.strip().replace("", "(blank)")
#     piv = (df.groupby("Type Of Lead", dropna=False)["Paid Check"]
#            .value_counts().unstack(fill_value=0)
#            .reindex(columns=["Paid", "Unpaid"], fill_value=0))
#     piv["Grand Total"] = piv["Paid"] + piv["Unpaid"]
#     piv.sort_values("Grand Total", ascending=False, inplace=True)
#     piv["%"] = [f"{round(p/t*100)}%" if t else "0%"
#                 for p, t in zip(piv["Paid"], piv["Grand Total"])]
#     qual_counts = df.groupby("Type Of Lead")["CallerAI Qualified"].apply(
#         lambda x: (x == "Yes").sum())
#     piv["AI Qualified"] = qual_counts.reindex(piv.index, fill_value=0)
#     tp, tu, tt, tq = (piv["Paid"].sum(), piv["Unpaid"].sum(),
#                       piv["Grand Total"].sum(), piv["AI Qualified"].sum())
#     return pd.concat([piv, pd.DataFrame(
#         [{"Paid": tp, "Unpaid": tu, "Grand Total": tt,
#           "%": f"{round(tp/tt*100)}%" if tt else "0%", "AI Qualified": tq}],
#         index=["Grand Total"])])

# def ai_qualified_funnel_pivot(df):
#     df_aiq = df[df["CallerAI Qualified"] == "Yes"].copy()
#     if df_aiq.empty:
#         return pd.DataFrame([{"LSQ Disposition": "(No AI Qualified leads)", "Count": 0, "%": "-"}])
#     co = df_aiq["Calling Outcome"].fillna("No LSQ Disposition").astype(str).str.strip()
#     co = co.replace("", "No LSQ Disposition")
#     counts = co.value_counts().reset_index()
#     counts.columns = ["LSQ Disposition", "Count"]
#     total = counts["Count"].sum()
#     counts["%"] = [f"{round(c/total*100)}%" for c in counts["Count"]]
#     counts["_sort"] = counts["LSQ Disposition"].map(OUTCOME_ORDER).fillna(999)
#     counts.sort_values(["_sort", "Count"], ascending=[True, False], inplace=True)
#     counts.drop(columns=["_sort"], inplace=True)
#     counts.reset_index(drop=True, inplace=True)
#     return pd.concat([counts, pd.DataFrame([{
#         "LSQ Disposition": "Grand Total", "Count": total, "%": ""}])], ignore_index=True)

# def date_pivot(df):
#     if df.empty: return pd.DataFrame()
#     def _our_ans(grp_df):
#         if CALLERAI_DISCONNECT_COL not in grp_df.columns: return 0
#         return int(grp_df[CALLERAI_DISCONNECT_COL].fillna("").str.strip().str.lower()
#                    .isin({r.lower() for r in CALLERAI_ANSWERED_REASONS}).sum())
#     grp = df.groupby("Date")
#     result = pd.DataFrame({
#         "Total":        grp.size(),
#         "Paid":         grp["Paid Check"].apply(lambda x: (x == "Paid").sum()),
#         "Unpaid":       grp["Paid Check"].apply(lambda x: (x == "Unpaid").sum()),
#         "Our Answered": grp.apply(_our_ans, include_groups=False),
#         "AI Qualified": grp["CallerAI Qualified"].apply(lambda x: (x == "Yes").sum()),
#     })
#     result["Not Answered"]       = result["Total"] - result["Our Answered"]
#     result["% Paid (Attempted)"] = [f"{round(p/t*100)}%" if t else "0%"
#                                     for p, t in zip(result["Paid"], result["Total"])]
#     result["% Paid (Answered)"]  = [f"{round(p/a*100)}%" if a else "0%"
#                                     for p, a in zip(result["Paid"], result["Our Answered"])]
#     result["_ts"] = pd.to_datetime(result.index, dayfirst=True, errors="coerce", format="mixed")
#     result.sort_values("_ts", inplace=True); result.drop(columns=["_ts"], inplace=True)
#     tt, tp, ta = result["Total"].sum(), result["Paid"].sum(), result["Our Answered"].sum()
#     return pd.concat([result, pd.DataFrame([{
#         "Total": tt, "Paid": tp, "Unpaid": result["Unpaid"].sum(),
#         "Our Answered": ta, "Not Answered": result["Not Answered"].sum(),
#         "AI Qualified": result["AI Qualified"].sum(),
#         "% Paid (Attempted)": f"{round(tp/tt*100)}%" if tt else "0%",
#         "% Paid (Answered)":  f"{round(tp/ta*100)}%" if ta else "0%",
#     }], index=["Grand Total"])])

# def plan_revenue_pivot(df):
#     if "Revenue_Attributed" in df.columns:
#         df_p = df[(df["Revenue_Attributed"] > 0) &
#                   df["Plan Name"].notna() &
#                   (df["Plan Name"].astype(str).str.strip() != "")].copy()
#         rev_col = "Revenue_Attributed"
#     else:
#         df_p = df[(df["Calling Outcome"] == "Purchased") &
#                   df["Plan Name"].notna() &
#                   (df["Plan Name"].astype(str).str.strip() != "")].copy()
#         rev_col = "Revenue"
#     if df_p.empty:
#         return pd.DataFrame(columns=["Plan Name", "Revenue (Rs)", "Order Count"])
#     df_p["_rev"] = pd.to_numeric(df_p[rev_col], errors="coerce").fillna(0)
#     grp = df_p.groupby("Plan Name").agg(
#         Revenue=("_rev", "sum"), Order_Count=("Plan Name", "count")).reset_index()
#     grp.columns = ["Plan Name", "Revenue (Rs)", "Order Count"]
#     grp.sort_values("Revenue (Rs)", ascending=False, inplace=True)
#     return pd.concat([grp, pd.DataFrame([{
#         "Plan Name": "Grand Total",
#         "Revenue (Rs)": grp["Revenue (Rs)"].sum(),
#         "Order Count":  grp["Order Count"].sum()
#     }])], ignore_index=True)

# def plan_revenue_pivot_today(df, today_paid_set):
#     if not today_paid_set:
#         return pd.DataFrame(columns=["Plan Name", "Revenue (Rs)", "Order Count"])
#     if "Revenue_Attributed" in df.columns:
#         df_p = df[df["Customer Number"].isin(today_paid_set) &
#                   (df["Revenue_Attributed"] > 0) &
#                   df["Plan Name"].notna() &
#                   (df["Plan Name"].astype(str).str.strip() != "")].copy()
#         rev_col = "Revenue_Attributed"
#     else:
#         df_p = df[df["Customer Number"].isin(today_paid_set) &
#                   df["Plan Name"].notna() &
#                   (df["Plan Name"].astype(str).str.strip() != "")].copy()
#         rev_col = "Revenue"
#     if df_p.empty:
#         return pd.DataFrame(columns=["Plan Name", "Revenue (Rs)", "Order Count"])
#     df_p["_rev"] = pd.to_numeric(df_p[rev_col], errors="coerce").fillna(0)
#     grp = df_p.groupby("Plan Name").agg(
#         Revenue=("_rev", "sum"), Order_Count=("Plan Name", "count")).reset_index()
#     grp.columns = ["Plan Name", "Revenue (Rs)", "Order Count"]
#     grp.sort_values("Revenue (Rs)", ascending=False, inplace=True)
#     return pd.concat([grp, pd.DataFrame([{
#         "Plan Name": "Grand Total",
#         "Revenue (Rs)": grp["Revenue (Rs)"].sum(),
#         "Order Count":  grp["Order Count"].sum()
#     }])], ignore_index=True)

# def plan_by_segment(df):
#     if "Revenue_Attributed" in df.columns:
#         df_p = df[(df["Revenue_Attributed"] > 0) &
#                   df["Plan Name"].notna() &
#                   (df["Plan Name"].astype(str).str.strip() != "")].copy()
#         rev_col = "Revenue_Attributed"
#     else:
#         df_p = df[(df["Calling Outcome"] == "Purchased") &
#                   df["Plan Name"].notna() &
#                   (df["Plan Name"].astype(str).str.strip() != "")].copy()
#         rev_col = "Revenue"
#     if df_p.empty:
#         return {}
#     df_p["_rev"] = pd.to_numeric(df_p[rev_col], errors="coerce").fillna(0)
#     grp = df_p.groupby("Plan Name").agg(revenue=("_rev", "sum"), count=("Plan Name", "count"))
#     return {plan: {"revenue": float(row["revenue"]), "count": int(row["count"])}
#             for plan, row in grp.iterrows()}

# def build_planwise_rows(df_win, df_ttc_unc, df_ttc, df_renewal_ttc):
#     p_win = plan_by_segment(df_win)
#     p_unc = plan_by_segment(df_ttc_unc)
#     p_ttc = plan_by_segment(df_ttc)
#     p_ren = plan_by_segment(df_renewal_ttc)
#     all_plans = sorted(set(list(p_win) + list(p_unc) + list(p_ttc) + list(p_ren)))
#     rows = []
#     for plan in all_plans:
#         w = p_win.get(plan, {"revenue": 0, "count": 0})
#         u = p_unc.get(plan, {"revenue": 0, "count": 0})
#         t = p_ttc.get(plan, {"revenue": 0, "count": 0})
#         r = p_ren.get(plan, {"revenue": 0, "count": 0})
#         rows.append([plan,
#                      w["revenue"] + u["revenue"] + t["revenue"] + r["revenue"],
#                      w["count"]   + u["count"]   + t["count"]   + r["count"],
#                      w["revenue"], w["count"], u["revenue"], u["count"],
#                      t["revenue"], t["count"], r["revenue"], r["count"]])
#     tr_w = sum(v["revenue"] for v in p_win.values()); tc_w = sum(v["count"] for v in p_win.values())
#     tr_u = sum(v["revenue"] for v in p_unc.values()); tc_u = sum(v["count"] for v in p_unc.values())
#     tr_t = sum(v["revenue"] for v in p_ttc.values()); tc_t = sum(v["count"] for v in p_ttc.values())
#     tr_r = sum(v["revenue"] for v in p_ren.values()); tc_r = sum(v["count"] for v in p_ren.values())
#     rows.append(["Grand Total",
#                  tr_w + tr_u + tr_t + tr_r, tc_w + tc_u + tc_t + tc_r,
#                  tr_w, tc_w, tr_u, tc_u, tr_t, tc_t, tr_r, tc_r])
#     return rows


# # ── Styles ─────────────────────────────────────────────────────────────────────
# TITLE_FILL          = PatternFill("solid", fgColor="2E75B6")
# HEADER_FILL         = PatternFill("solid", fgColor="BDD7EE")
# TOTAL_FILL          = PatternFill("solid", fgColor="9DC3E6")
# PAID_FILL           = PatternFill("solid", fgColor="E2EFDA")
# AI_QUAL_FILL        = PatternFill("solid", fgColor="FFF2CC")
# AI_COL_FILL         = PatternFill("solid", fgColor="FCE4D6")
# WIN_FILL            = PatternFill("solid", fgColor="E8F4FD")
# TTC_UNC_FILL        = PatternFill("solid", fgColor="FEF9E7")
# TTC_FILL            = PatternFill("solid", fgColor="EDE7F6")
# RENEWAL_TTC_FILL    = PatternFill("solid", fgColor="E8F8E8")
# ANS_COL_FILL        = PatternFill("solid", fgColor="E8F5E9")
# OUR_ANS_FILL        = PatternFill("solid", fgColor="C8E6C9")
# EMPTY_ROW_FILL      = PatternFill("solid", fgColor="F5F5F5")
# AI_QUAL_PAID_FILL   = PatternFill("solid", fgColor="C6EFCE")
# AI_UNQUAL_PAID_FILL = PatternFill("solid", fgColor="FFEB9C")
# REVENUE_FILL        = PatternFill("solid", fgColor="D6E4F0")

# THIN = Side(border_style="thin", color="B0B0B0")
# BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# seg_fills = {
#     WINBACK_SOURCE:      WIN_FILL,
#     TTC_UNC_SOURCE:      TTC_UNC_FILL,
#     TTC_SOURCE:          TTC_FILL,
#     RENEWAL_TTC_SOURCE:  RENEWAL_TTC_FILL,
#     WINBACK_DISPLAY:     WIN_FILL,
#     TTC_UNC_DISPLAY:     TTC_UNC_FILL,
#     TTC_DISPLAY:         TTC_FILL,
#     RENEWAL_TTC_DISPLAY: RENEWAL_TTC_FILL,
# }

# def _cell(ws, r, c, value, bold=False, fill=None, align="center", color="000000"):
#     cell = ws.cell(r, c)
#     cell.value = value
#     cell.font  = Font(bold=bold, color=color)
#     cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
#     cell.border = BOX
#     if fill: cell.fill = fill
#     return cell


# # ── Table writers ──────────────────────────────────────────────────────────────
# def write_summary_matrix(ws, title, rows_data, start_row, start_col):
#     """Write a summary matrix block with a Grand Total row appended."""
#     sr, sc = start_row, start_col
#     ws.merge_cells(start_row=sr, start_column=sc,
#                    end_row=sr, end_column=sc + len(SUMMARY_METRIC_COLS))
#     _cell(ws, sr, sc, title, bold=True, fill=TITLE_FILL, color="FFFFFF")
#     ws.row_dimensions[sr].height = 22
#     sr += 1

#     _header_fills = {
#         "AI Qualified":         AI_COL_FILL,
#         "Company Answered":     ANS_COL_FILL,
#         "Our Answered":         OUR_ANS_FILL,
#         "Answered & Paid":      PAID_FILL,
#         "AI Qual & Paid":       AI_QUAL_PAID_FILL,
#         "AI Unqual & Paid":     AI_UNQUAL_PAID_FILL,
#         "AI Qual & Paid Rev":   REVENUE_FILL,
#         "AI Unqual & Paid Rev": REVENUE_FILL,
#         "Ans & Paid Rev":       REVENUE_FILL,
#     }
#     for ci, h in enumerate(["Segment"] + SUMMARY_METRIC_COLS):
#         _cell(ws, sr, sc + ci, h, bold=True, fill=_header_fills.get(h, HEADER_FILL))
#     sr += 1

#     # Segment rows
#     for row in rows_data:
#         seg   = row["Segment"]
#         rfill = seg_fills.get(seg)
#         _cell(ws, sr, sc, seg, bold=True, fill=rfill, align="left")
#         for ci, col in enumerate(SUMMARY_METRIC_COLS, start=1):
#             cfill = (AI_QUAL_FILL        if col == "AI Qualified"
#                      else OUR_ANS_FILL        if col == "Our Answered"
#                      else ANS_COL_FILL        if col == "Company Answered"
#                      else AI_QUAL_PAID_FILL   if col == "AI Qual & Paid"
#                      else AI_UNQUAL_PAID_FILL if col == "AI Unqual & Paid"
#                      else REVENUE_FILL        if col in ("AI Qual & Paid Rev",
#                                                          "AI Unqual & Paid Rev",
#                                                          "Ans & Paid Rev")
#                      else rfill)
#             _cell(ws, sr, sc + ci, row.get(col, ""), fill=cfill)
#         sr += 1

#     # v21: Grand Total row
#     gt = build_grand_total_row(rows_data)
#     _cell(ws, sr, sc, gt["Segment"], bold=True, fill=TOTAL_FILL, align="left")
#     for ci, col in enumerate(SUMMARY_METRIC_COLS, start=1):
#         _cell(ws, sr, sc + ci, gt.get(col, ""), bold=True, fill=TOTAL_FILL)
#     sr += 1

#     return sr

# def write_ai_qual_funnel_table(ws, title, df, start_row, start_col):
#     sr, sc = start_row, start_col
#     ws.merge_cells(start_row=sr, start_column=sc, end_row=sr, end_column=sc + 2)
#     _cell(ws, sr, sc, title, bold=True, fill=TITLE_FILL, color="FFFFFF")
#     ws.row_dimensions[sr].height = 22; sr += 1
#     for ci, h in enumerate(["LSQ Disposition", "Count", "%"]):
#         _cell(ws, sr, sc + ci, h, bold=True,
#               fill=AI_QUAL_FILL if ci == 0 else HEADER_FILL)
#     sr += 1
#     for _, row in df.iterrows():
#         is_gt    = (row["LSQ Disposition"] == "Grand Total")
#         is_empty = (row["LSQ Disposition"] == "(No AI Qualified leads)")
#         fill      = TOTAL_FILL if is_gt else (EMPTY_ROW_FILL if is_empty else None)
#         txt_color = "888888" if is_empty else "000000"
#         _cell(ws, sr, sc,     row["LSQ Disposition"], bold=is_gt, fill=fill,
#               align="left", color=txt_color)
#         _cell(ws, sr, sc + 1, row["Count"], bold=is_gt, fill=fill)
#         _cell(ws, sr, sc + 2, row["%"],     bold=is_gt, fill=fill)
#         sr += 1
#     return sr

# def write_campaign_table(ws, title, df, start_row, start_col):
#     sr, sc = start_row, start_col
#     ws.merge_cells(start_row=sr, start_column=sc, end_row=sr, end_column=sc + 5)
#     _cell(ws, sr, sc, title, bold=True, fill=TITLE_FILL, color="FFFFFF")
#     ws.row_dimensions[sr].height = 22; sr += 1
#     ws.merge_cells(start_row=sr, start_column=sc, end_row=sr, end_column=sc + 5)
#     _cell(ws, sr, sc, "Count of Leads", bold=True, fill=HEADER_FILL, align="left"); sr += 1
#     for ci, h in enumerate(["Type Of Lead", "Paid", "Unpaid", "Grand Total", "%", "AI Qualified"]):
#         _cell(ws, sr, sc + ci, h, bold=True,
#               fill=AI_COL_FILL if h == "AI Qualified" else HEADER_FILL)
#     sr += 1
#     for label, row in df.iterrows():
#         is_gt  = (label == "Grand Total")
#         fill   = TOTAL_FILL if is_gt else None
#         ai_val = row.get("AI Qualified", 0)
#         _cell(ws, sr, sc,     label,                     bold=is_gt, fill=fill, align="left")
#         _cell(ws, sr, sc + 1, row.get("Paid", 0),        bold=is_gt, fill=fill)
#         _cell(ws, sr, sc + 2, row.get("Unpaid", 0),      bold=is_gt, fill=fill)
#         _cell(ws, sr, sc + 3, row.get("Grand Total", 0), bold=is_gt, fill=fill)
#         _cell(ws, sr, sc + 4, row.get("%", ""),           bold=is_gt, fill=fill)
#         _cell(ws, sr, sc + 5, ai_val, bold=is_gt,
#               fill=AI_QUAL_FILL if (not is_gt and ai_val) else fill)
#         sr += 1
#     return sr

# def write_date_table(ws, title, df, start_row, start_col):
#     sr, sc = start_row, start_col
#     DATA_COL = ["Total", "Paid", "Unpaid", "Our Answered", "Not Answered",
#                 "% Paid (Attempted)", "% Paid (Answered)", "AI Qualified"]
#     ws.merge_cells(start_row=sr, start_column=sc,
#                    end_row=sr, end_column=sc + len(DATA_COL))
#     _cell(ws, sr, sc, title, bold=True, fill=TITLE_FILL, color="FFFFFF")
#     ws.row_dimensions[sr].height = 22; sr += 1
#     for ci, h in enumerate(["Date"] + DATA_COL):
#         fill = (AI_COL_FILL   if h == "AI Qualified"
#                 else OUR_ANS_FILL if h == "Our Answered"
#                 else HEADER_FILL)
#         _cell(ws, sr, sc + ci, h, bold=True, fill=fill)
#     sr += 1
#     for label, row in df.iterrows():
#         is_gt = (label == "Grand Total")
#         fill  = TOTAL_FILL if is_gt else None
#         _cell(ws, sr, sc, label, bold=is_gt, fill=fill, align="left")
#         for ci, col in enumerate(DATA_COL, start=1):
#             cfill = (fill         if is_gt
#                      else AI_QUAL_FILL if col == "AI Qualified" and row.get(col, 0)
#                      else OUR_ANS_FILL if col == "Our Answered"
#                      else None)
#             _cell(ws, sr, sc + ci, row.get(col, 0), bold=is_gt, fill=cfill)
#         sr += 1
#     return sr

# def write_plan_revenue_table(ws, title, df, start_row, start_col):
#     sr, sc = start_row, start_col
#     ws.merge_cells(start_row=sr, start_column=sc, end_row=sr, end_column=sc + 2)
#     _cell(ws, sr, sc, title, bold=True, fill=TITLE_FILL, color="FFFFFF")
#     ws.row_dimensions[sr].height = 22; sr += 1
#     for ci, h in enumerate(["Plan Name", "Revenue (Rs)", "Order Count"]):
#         _cell(ws, sr, sc + ci, h, bold=True, fill=HEADER_FILL)
#     sr += 1
#     for _, row in df.iterrows():
#         is_gt = (row["Plan Name"] == "Grand Total")
#         fill  = TOTAL_FILL if is_gt else None
#         _cell(ws, sr, sc,     row["Plan Name"],    bold=is_gt, fill=fill, align="left")
#         _cell(ws, sr, sc + 1, row["Revenue (Rs)"], bold=is_gt, fill=fill)
#         _cell(ws, sr, sc + 2, row["Order Count"],  bold=is_gt, fill=fill)
#         sr += 1
#     return sr

# def write_planwise_block(ws, title, data_rows, start_row, start_col):
#     sr, sc = start_row, start_col
#     nc = len(PLAN_COLS)
#     ws.merge_cells(start_row=sr, start_column=sc, end_row=sr, end_column=sc + nc - 1)
#     _cell(ws, sr, sc, title, bold=True, fill=TITLE_FILL, color="FFFFFF")
#     ws.row_dimensions[sr].height = 22; sr += 1
#     for ci, h in enumerate(PLAN_COLS):
#         _cell(ws, sr, sc + ci, h, bold=True, fill=HEADER_FILL)
#     sr += 1
#     for row_vals in data_rows:
#         is_gt = (str(row_vals[0]) == "Grand Total")
#         fill  = TOTAL_FILL if is_gt else None
#         _cell(ws, sr, sc, row_vals[0], bold=is_gt, fill=fill, align="left")
#         for ci in range(1, nc):
#             _cell(ws, sr, sc + ci,
#                   row_vals[ci] if ci < len(row_vals) else "",
#                   bold=is_gt, fill=fill)
#         sr += 1
#     return sr


# # ── v21 enhanced: Rich formatted output to Google Sheet ──────────────────────
# def write_summary_to_gsheet(
#     gc, summary_today, summary_overall, today_str,
#     t_aiq_all, t_aiq_win, t_aiq_ttc_unc, t_aiq_ttc, t_aiq_renewal_ttc,
#     t_plan_rev
# ):
#     """
#     Writes a richly formatted Lead Summary Report to OUTPUT_SUMMARY_SHEET_ID (gid=0):
#       Block 1: Today's Summary  — navy title, blue headers, segment colors, amber GT
#       Block 2: Overall Summary  — same formatting
#       Block 3: 5 AI Qual Funnels side-by-side (display-name labels, alternating rows)
#       Block 4: Plan-wise Revenue (non-zero only, sorted by revenue desc)
#     Font: Lexend throughout.
#     """
#     def _rgb(h):
#         h = h.lstrip("#")
#         r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
#         return {"red": r / 255.0, "green": g / 255.0, "blue": b / 255.0}

#     NAVY, WHITE, BLUE_H = "1B2A4A", "FFFFFF", "2E75B6"
#     AMBER = "F4B942"
#     SEG_C = {
#         WINBACK_DISPLAY:     "E8F4FD",
#         TTC_UNC_DISPLAY:     "FEF9E7",
#         TTC_DISPLAY:         "EDE7F6",
#         RENEWAL_TTC_DISPLAY: "E8F8E8",
#     }
#     F_TITLE = "1565C0"; F_COL = "1976D2"; F_ALT = "E3F2FD"; F_GT = "BBDEFB"
#     P_TITLE = "1B2A4A"; P_COL  = "37474F"; P_ALT = "F1F8E9"; P_GT = "C8E6C9"
#     FONT = "Lexend"

#     DISP_COLS = ["Segment"] + SUMMARY_METRIC_COLS
#     N_SUM     = len(DISP_COLS)   # 14 cols A-N
#     TOTAL_W   = 19               # A-S for funnels

#     gt_today   = build_grand_total_row(summary_today)
#     gt_overall = build_grand_total_row(summary_overall)

#     def rv(d): return [d.get(c, "") for c in DISP_COLS]
#     def pad(row, w=TOTAL_W): return (list(row) + [""] * w)[:w]

#     reqs = []

#     def fmt(r1, r2, c1, c2, bg=None, fg="000000", bold=False,
#             size=10, align="CENTER", valign="MIDDLE"):
#         cell_fmt = {
#             "textFormat": {
#                 "fontFamily": FONT, "fontSize": size, "bold": bold,
#                 "foregroundColor": _rgb(fg),
#             },
#             "horizontalAlignment": align,
#             "verticalAlignment": valign,
#         }
#         if bg: cell_fmt["backgroundColor"] = _rgb(bg)
#         reqs.append({
#             "repeatCell": {
#                 "range": {"sheetId": 0,
#                           "startRowIndex": r1, "endRowIndex": r2,
#                           "startColumnIndex": c1, "endColumnIndex": c2},
#                 "cell": {"userEnteredFormat": cell_fmt},
#                 "fields": ("userEnteredFormat(backgroundColor,textFormat,"
#                            "horizontalAlignment,verticalAlignment)"),
#             }
#         })

#     def merge(r1, r2, c1, c2):
#         reqs.append({
#             "mergeCells": {
#                 "range": {"sheetId": 0,
#                           "startRowIndex": r1, "endRowIndex": r2,
#                           "startColumnIndex": c1, "endColumnIndex": c2},
#                 "mergeType": "MERGE_ALL",
#             }
#         })

#     def row_h(ri, h):
#         reqs.append({
#             "updateDimensionProperties": {
#                 "range": {"sheetId": 0, "dimension": "ROWS",
#                           "startIndex": ri, "endIndex": ri + 1},
#                 "properties": {"pixelSize": h}, "fields": "pixelSize",
#             }
#         })

#     def col_w(ci, w):
#         reqs.append({
#             "updateDimensionProperties": {
#                 "range": {"sheetId": 0, "dimension": "COLUMNS",
#                           "startIndex": ci, "endIndex": ci + 1},
#                 "properties": {"pixelSize": w}, "fields": "pixelSize",
#             }
#         })

#     grid = []
#     _r   = [0]

#     def add_row(values):
#         grid.append(pad(values))
#         r = _r[0]; _r[0] += 1
#         return r

#     # ═══ BLOCK 1: Today's Summary ═══════════════════════════════════════════
#     r = add_row([f"Today's Summary ({today_str})"])
#     fmt(r, r+1, 0, N_SUM, bg=NAVY, fg=WHITE, bold=True, size=11, align="LEFT")
#     merge(r, r+1, 0, N_SUM); row_h(r, 28)

#     r = add_row(DISP_COLS)
#     fmt(r, r+1, 0, N_SUM, bg=BLUE_H, fg=WHITE, bold=True)
#     fmt(r, r+1, 0, 1, bg=BLUE_H, fg=WHITE, bold=True, align="LEFT")
#     row_h(r, 24)

#     for row_d in summary_today:
#         seg = row_d["Segment"]; bg = SEG_C.get(seg, "FFFFFF")
#         r = add_row(rv(row_d))
#         fmt(r, r+1, 0, 1, bg=bg, align="LEFT")
#         fmt(r, r+1, 1, N_SUM, bg=bg)
#         row_h(r, 22)

#     r = add_row(rv(gt_today))
#     fmt(r, r+1, 0, 1, bg=AMBER, bold=True, align="LEFT")
#     fmt(r, r+1, 1, N_SUM, bg=AMBER, bold=True)
#     row_h(r, 24)

#     r = add_row([]); row_h(r, 10)

#     # ═══ BLOCK 2: Overall Summary ════════════════════════════════════════════
#     r = add_row(["Overall Summary (All Dates)"])
#     fmt(r, r+1, 0, N_SUM, bg=NAVY, fg=WHITE, bold=True, size=11, align="LEFT")
#     merge(r, r+1, 0, N_SUM); row_h(r, 28)

#     r = add_row(DISP_COLS)
#     fmt(r, r+1, 0, N_SUM, bg=BLUE_H, fg=WHITE, bold=True)
#     fmt(r, r+1, 0, 1, bg=BLUE_H, fg=WHITE, bold=True, align="LEFT")
#     row_h(r, 24)

#     for row_d in summary_overall:
#         seg = row_d["Segment"]; bg = SEG_C.get(seg, "FFFFFF")
#         r = add_row(rv(row_d))
#         fmt(r, r+1, 0, 1, bg=bg, align="LEFT")
#         fmt(r, r+1, 1, N_SUM, bg=bg)
#         row_h(r, 22)

#     r = add_row(rv(gt_overall))
#     fmt(r, r+1, 0, 1, bg=AMBER, bold=True, align="LEFT")
#     fmt(r, r+1, 1, N_SUM, bg=AMBER, bold=True)
#     row_h(r, 24)

#     r = add_row([]); row_h(r, 10)

#     # ═══ BLOCK 3: AI Qual Funnels (5 side-by-side) ═══════════════════════════
#     funnel_tables = [
#         ("Overall AI Qual Funnel",             t_aiq_all),
#         ("Winback AI Qual Funnel",              t_aiq_win),
#         (f"{TTC_UNC_DISPLAY} AI Qual Funnel",  t_aiq_ttc_unc),
#         (f"{TTC_DISPLAY} AI Qual Funnel",       t_aiq_ttc),
#         ("Renewal_TTC AI Qual Funnel",          t_aiq_renewal_ttc),
#     ]
#     F_OFF = [0, 4, 8, 12, 16]

#     max_df_rows  = max(len(df) for _, df in funnel_tables)
#     total_f_rows = 2 + max_df_rows   # title + col-hdr + data

#     f_start = _r[0]
#     for _ in range(total_f_rows):
#         add_row([])

#     for tbl_i, (tbl_title, tbl_df) in enumerate(funnel_tables):
#         off = F_OFF[tbl_i]

#         r = f_start
#         grid[r][off] = tbl_title
#         fmt(r, r+1, off, off+3, bg=F_TITLE, fg=WHITE, bold=True, size=10, align="LEFT")
#         merge(r, r+1, off, off+3)
#         if tbl_i == 0: row_h(r, 26)

#         r = f_start + 1
#         grid[r][off] = "LSQ Disposition"
#         grid[r][off+1] = "Count"
#         grid[r][off+2] = "%"
#         fmt(r, r+1, off, off+3, bg=F_COL, fg=WHITE, bold=True)
#         fmt(r, r+1, off, off+1, bg=F_COL, fg=WHITE, bold=True, align="LEFT")
#         if tbl_i == 0: row_h(r, 22)

#         for di, (_, drow) in enumerate(tbl_df.iterrows()):
#             r = f_start + 2 + di
#             disp  = str(drow.get("LSQ Disposition", ""))
#             is_gt = (disp == "Grand Total")
#             grid[r][off]   = disp
#             grid[r][off+1] = int(drow["Count"]) if not is_gt and pd.notna(drow.get("Count")) else drow.get("Count", "")
#             grid[r][off+2] = drow.get("%", "")
#             bg   = F_GT if is_gt else (F_ALT if di % 2 else "FFFFFF")
#             bold = is_gt
#             fmt(r, r+1, off, off+3, bg=bg, bold=bold)
#             fmt(r, r+1, off, off+1, bg=bg, bold=bold, align="LEFT")
#             if tbl_i == 0: row_h(r, 20)

#     r = add_row([]); row_h(r, 10)

#     # ═══ BLOCK 4: Plan-wise Revenue ══════════════════════════════════════════
#     r = add_row(["Plan-wise Revenue", "", ""])
#     fmt(r, r+1, 0, 3, bg=P_TITLE, fg=WHITE, bold=True, size=11, align="LEFT")
#     merge(r, r+1, 0, 3); row_h(r, 28)

#     r = add_row(["Plan Name", "Revenue (Rs)", "Order Count"])
#     fmt(r, r+1, 0, 3, bg=P_COL, fg=WHITE, bold=True)
#     fmt(r, r+1, 0, 1, bg=P_COL, fg=WHITE, bold=True, align="LEFT")
#     row_h(r, 22)

#     for di, (_, pr) in enumerate(t_plan_rev.iterrows()):
#         plan_n = str(pr.get("Plan Name", ""))
#         is_gt  = (plan_n == "Grand Total")
#         try:    rev_v = int(pr.get("Revenue (Rs)", 0))
#         except: rev_v = pr.get("Revenue (Rs)", 0)
#         try:    ord_v = int(pr.get("Order Count", 0))
#         except: ord_v = pr.get("Order Count", 0)
#         r = add_row([plan_n, rev_v, ord_v])
#         bg = P_GT if is_gt else (P_ALT if di % 2 else "FFFFFF")
#         fmt(r, r+1, 0, 3, bg=bg, bold=is_gt)
#         fmt(r, r+1, 0, 1, bg=bg, bold=is_gt, align="LEFT")
#         row_h(r, 20)

#     # ── Column widths ─────────────────────────────────────────────────────────
#     col_w(0, 155)
#     for c in range(1, N_SUM): col_w(c, 88)
#     for off in F_OFF:
#         col_w(off + 1, 70); col_w(off + 2, 55)
#     for off in [3, 7, 11, 15]: col_w(off, 12)

#     # ── Write to Google Sheet ─────────────────────────────────────────────────
#     sh = gc.open_by_key(OUTPUT_SUMMARY_SHEET_ID)
#     ws = sh.get_worksheet(0)
#     ws.clear()

#     try:
#         ws.update(range_name="A1", values=grid)
#     except TypeError:
#         ws.update("A1", grid)

#     all_reqs = [{
#         "unmergeCells": {
#             "range": {"sheetId": 0,
#                       "startRowIndex": 0, "endRowIndex": len(grid) + 5,
#                       "startColumnIndex": 0, "endColumnIndex": TOTAL_W}
#         }
#     }] + reqs

#     for i in range(0, len(all_reqs), 100):
#         sh.batch_update({"requests": all_reqs[i:i + 100]})

#     print(f"OK Lead Summary -> Google Sheet {OUTPUT_SUMMARY_SHEET_ID} "
#           f"({len(grid)} rows, {len(all_reqs)} format ops applied)")


# # ── Build pivot tables ─────────────────────────────────────────────────────────
# t1              = campaign_pivot(out_all)
# t_win_c         = campaign_pivot(out_win)
# t_ttc_unc_c     = campaign_pivot(out_ttc_unc)
# t_ttc_c         = campaign_pivot(out_ttc)
# t_renewal_ttc_c = campaign_pivot(out_renewal_ttc)
# t_today_c       = campaign_pivot(out_all[out_all["_date_ts"] == today_ts])

# t_aiq_all         = ai_qualified_funnel_pivot(out_all)
# t_aiq_win         = ai_qualified_funnel_pivot(out_win)
# t_aiq_ttc_unc     = ai_qualified_funnel_pivot(out_ttc_unc)
# t_aiq_ttc         = ai_qualified_funnel_pivot(out_ttc)
# t_aiq_renewal_ttc = ai_qualified_funnel_pivot(out_renewal_ttc)
# t_aiq_today       = ai_qualified_funnel_pivot(out_all[out_all["_date_ts"] == today_ts])

# t_plan_rev       = plan_revenue_pivot(out_all)
# t_plan_rev_today = plan_revenue_pivot_today(out_all, paid_set_today)
# t_wtd            = date_pivot(out_all)


# # ── Build Summary XLSX workbook ────────────────────────────────────────────────
# wb   = openpyxl.Workbook()
# wsum = wb.active
# wsum.title = "Summary"
# wsum.sheet_view.showGridLines = False

# nr = write_summary_matrix(wsum, f"Today's Summary ({today_str})", summary_today,   1, 1)
# nr = write_summary_matrix(wsum, "Overall Summary (All Dates)",    summary_overall, nr + 2, 1)
# gap0 = nr + 2

# nr_1 = write_ai_qual_funnel_table(wsum, "Overall AI Qual Funnel",     t_aiq_all,         gap0, 1)
# nr_2 = write_ai_qual_funnel_table(wsum, "Winback AI Qual Funnel",      t_aiq_win,         gap0, 5)
# nr_3 = write_ai_qual_funnel_table(wsum, "TTC+Uncalled AI Qual Funnel", t_aiq_ttc_unc,     gap0, 9)
# nr_4 = write_ai_qual_funnel_table(wsum, "TTC AI Qual Funnel",          t_aiq_ttc,         gap0, 13)
# nr_5 = write_ai_qual_funnel_table(wsum, "Renewal_TTC AI Qual Funnel",  t_aiq_renewal_ttc, gap0, 17)
# gap1 = max(nr_1, nr_2, nr_3, nr_4, nr_5) + 2

# nr_l = write_campaign_table(wsum,     "Overall Leads",     t1,         gap1, 1)
# nr_r = write_plan_revenue_table(wsum, "Plan-wise Revenue", t_plan_rev, gap1, 9)
# gap2 = max(nr_l, nr_r) + 2

# nr_l = write_campaign_table(wsum, "Winback Leads",      t_win_c,         gap2, 1)
# nr_m = write_campaign_table(wsum, "TTC+Uncalled Leads", t_ttc_unc_c,     gap2, 9)
# nr_r = write_campaign_table(wsum, "TTC Leads",          t_ttc_c,         gap2, 17)
# nr_x = write_campaign_table(wsum, "Renewal_TTC Leads",  t_renewal_ttc_c, gap2, 25)
# gap3 = max(nr_l, nr_m, nr_r, nr_x) + 2

# nr_l = write_date_table(wsum,           "Week to Date",                 t_wtd,       gap3, 1)
# nr_r = write_ai_qual_funnel_table(wsum, f"Today AI Qual ({today_str})", t_aiq_today, gap3, 13)
# gap4 = max(nr_l, nr_r) + 2

# write_campaign_table(wsum,     f"Today's Leads ({today_str})",        t_today_c,        gap4, 1)
# write_plan_revenue_table(wsum, f"Today's Plan Revenue ({today_str})", t_plan_rev_today, gap4, 9)

# for col, w in {"A": 22, "B": 12, "C": 14, "D": 14, "E": 12, "F": 14, "G": 14, "H": 12,
#                "I": 14, "J": 14, "K": 12, "L": 12, "M": 14, "N": 14, "O": 12, "P": 14,
#                "Q": 14, "R": 14, "S": 14, "T": 14, "U": 14, "V": 14, "W": 14,
#                "X": 14, "Y": 14, "Z": 14, "AA": 14, "AB": 14, "AC": 14, "AD": 14}.items():
#     wsum.column_dimensions[col].width = w


# # ══════════════════════════════════════════════════════════════════════════════
# # STEP 7b — Detail tabs (cascading funnel)
# # ══════════════════════════════════════════════════════════════════════════════
# DETAIL_COLS = [
#     "Date", "Customer Number", "Lead Name", "Source", "Type Of Lead",
#     "Lead Stage", "Disconnect Reason", "Business Disposition",
#     "CallerAI Qualified", "Business Disposition Confidence",
#     "Calling Outcome", "Paid Check", "Lead Owner", "Activity Added By",
#     "Plan Name", "Revenue"
# ]
# DETAIL_WIDTHS = [16, 16, 26, 16, 18, 18, 22, 20, 18, 24, 22, 12, 28, 28, 22, 12]

# detail_df = out[[c for c in DETAIL_COLS if c in out.columns]].copy()
# for col in ["Calling Outcome", "CallerAI Qualified", "Business Disposition",
#             "Business Disposition Confidence",
#             "Lead Owner", "Activity Added By", "Plan Name", "Disconnect Reason"]:
#     if col in detail_df.columns:
#         detail_df[col] = detail_df[col].fillna("")
# detail_df["Revenue"] = detail_df["Revenue"].fillna(0)

# detail_df["_tab"] = detail_df.apply(assign_tab, axis=1)
# cols_in_detail = [c for c in DETAIL_COLS if c in detail_df.columns]
# widths_map     = {c: w for c, w in zip(DETAIL_COLS, DETAIL_WIDTHS)}

# _disp_slots  = sorted([k for k in OUTCOME_ORDER if k != "Purchased"],
#                       key=lambda k: OUTCOME_ORDER[k])
# _extra_slots = sorted(
#     [t for t in detail_df["_tab"].unique()
#      if t not in (DETAIL_TAB_PURCHASED, DETAIL_TAB_AI_QUAL, DETAIL_TAB_UNANSWERED)
#      and t not in OUTCOME_ORDER],
#     key=lambda x: OUTCOME_ORDER.get(x, 998))
# PRIORITY_TAB_ORDER = ([DETAIL_TAB_PURCHASED, DETAIL_TAB_AI_QUAL]
#                       + _disp_slots + _extra_slots + [DETAIL_TAB_UNANSWERED])

# for tab_name in PRIORITY_TAB_ORDER:
#     subset = detail_df[detail_df["_tab"] == tab_name].copy()
#     if subset.empty: continue

#     if tab_name == DETAIL_TAB_AI_QUAL:
#         subset["_d"] = subset["Calling Outcome"].map(OUTCOME_ORDER).fillna(999)
#         subset.sort_values(["_d", "Business Disposition Confidence", "Date"], inplace=True)
#         subset.drop(columns=["_d"], inplace=True)
#     else:
#         subset["_ps"] = subset["Paid Check"].map({"Paid": 0, "Unpaid": 1}).fillna(1)
#         subset.sort_values(["_ps", "Date"], inplace=True)
#         subset.drop(columns=["_ps"], inplace=True)
#     subset.reset_index(drop=True, inplace=True)

#     paid_in   = int((subset["Paid Check"] == "Paid").sum())
#     unpaid_in = int((subset["Paid Check"] == "Unpaid").sum())
#     ai_q_in   = int((subset["CallerAI Qualified"] == "Yes").sum())
#     sheet_name = re.sub(r'[:\\/?*\[\]]', '', str(tab_name))[:31]

#     ws_d = wb.create_sheet(title=sheet_name)
#     ws_d.sheet_view.showGridLines = False

#     ws_d.merge_cells(start_row=1, start_column=1, end_row=1,
#                      end_column=len(cols_in_detail))
#     tc = ws_d.cell(1, 1)
#     tc.value = (f"{tab_name}  --  Total: {len(subset)}  |  Paid: {paid_in}  |  "
#                 f"Unpaid: {unpaid_in}  |  AI Qualified: {ai_q_in}  |  {src_count_str(subset)}")
#     tc.font = Font(bold=True, color="FFFFFF"); tc.fill = TITLE_FILL
#     tc.alignment = Alignment(horizontal="left", vertical="center")
#     ws_d.row_dimensions[1].height = 22

#     header_row = 2
#     if tab_name == DETAIL_TAB_AI_QUAL and not subset.empty:
#         _bk = (subset.groupby(["Calling Outcome", "Business Disposition Confidence"],
#                                dropna=False)
#                .size().reset_index(name="Count"))
#         bk_str = "  |  ".join(
#             f"{r['Calling Outcome']} / {r['Business Disposition Confidence']}: {r['Count']}"
#             for _, r in _bk.iterrows())
#         if bk_str:
#             ws_d.merge_cells(start_row=2, start_column=1, end_row=2,
#                              end_column=len(cols_in_detail))
#             bc = ws_d.cell(2, 1)
#             bc.value = bk_str; bc.font = Font(italic=True, color="444444")
#             bc.fill = AI_QUAL_FILL
#             bc.alignment = Alignment(horizontal="left", vertical="center")
#             bc.border = BOX; ws_d.row_dimensions[2].height = 18
#             header_row = 3

#     for ci, h in enumerate(cols_in_detail):
#         hfill = (AI_COL_FILL
#                  if h in ("CallerAI Qualified",
#                            "Business Disposition Confidence",
#                            "Business Disposition")
#                  else HEADER_FILL)
#         _cell(ws_d, header_row, ci + 1, h, bold=True, fill=hfill)

#     for ri, (_, row) in enumerate(subset[cols_in_detail].iterrows(),
#                                   start=header_row + 1):
#         is_paid = (row.get("Paid Check", "") == "Paid")
#         is_ai_q = (row.get("CallerAI Qualified", "No") == "Yes")
#         row_fill = (PAID_FILL   if is_paid
#                     else AI_QUAL_FILL if is_ai_q
#                     else seg_fills.get(row.get("Source", ""), None))
#         for ci, col in enumerate(cols_in_detail):
#             cell = ws_d.cell(ri, ci + 1)
#             cell.value = row[col]; cell.font = Font(bold=is_paid)
#             cell.alignment = Alignment(horizontal="left", vertical="center")
#             cell.border = BOX
#             if row_fill: cell.fill = row_fill

#     for ci, col in enumerate(cols_in_detail, start=1):
#         ws_d.column_dimensions[
#             openpyxl.utils.get_column_letter(ci)].width = widths_map.get(col, 16)


# # ══════════════════════════════════════════════════════════════════════════════
# # STEP 7c — AI Qualified tabs (per campaign + Overall) + AI Answered
# # ══════════════════════════════════════════════════════════════════════════════
# AI_QUAL_TAB_FILL = PatternFill("solid", fgColor="5B9BD5")

# def write_ai_qual_segment_tab(wb, sheet_name, df_segment, title):
#     df_aiq = df_segment[df_segment["CallerAI Qualified"] == "Yes"].copy()
#     for col in ["Calling Outcome", "Business Disposition",
#                 "Business Disposition Confidence",
#                 "Lead Owner", "Activity Added By", "Plan Name",
#                 "Disconnect Reason"]:
#         if col in df_aiq.columns: df_aiq[col] = df_aiq[col].fillna("")
#     if "Revenue" in df_aiq.columns:
#         df_aiq["Revenue"] = pd.to_numeric(df_aiq["Revenue"], errors="coerce").fillna(0)

#     df_aiq["_sort"] = df_aiq["Calling Outcome"].map(OUTCOME_ORDER).fillna(999)
#     df_aiq.sort_values(["_sort", "Business Disposition Confidence", "Date"],
#                        inplace=True)
#     df_aiq.drop(columns=["_sort"], inplace=True)
#     df_aiq.reset_index(drop=True, inplace=True)

#     disp_counts = {}
#     if "Calling Outcome" in df_aiq.columns:
#         disp_counts = (df_aiq["Calling Outcome"]
#                        .replace("", "No LSQ Dispo")
#                        .value_counts().to_dict())

#     cols_use = [c for c in DETAIL_COLS if c in df_aiq.columns]
#     w_map    = {c: w for c, w in zip(DETAIL_COLS, DETAIL_WIDTHS)}

#     ws = wb.create_sheet(title=re.sub(r'[:\\/?*\[\]]', '', sheet_name)[:31])
#     ws.sheet_view.showGridLines = False

#     ws.merge_cells(start_row=1, start_column=1, end_row=1,
#                    end_column=max(len(cols_use), 1))
#     tc = ws.cell(1, 1)
#     tc.value = (f"{title}  --  Total AI Qual: {len(df_aiq)}  |  "
#                 f"{src_count_str(df_aiq)}")
#     tc.font = Font(bold=True, color="FFFFFF"); tc.fill = AI_QUAL_TAB_FILL
#     tc.alignment = Alignment(horizontal="left", vertical="center")
#     ws.row_dimensions[1].height = 22

#     header_row = 2
#     if disp_counts:
#         bk_str = "  |  ".join(
#             f"{k}: {v}" for k, v in
#             sorted(disp_counts.items(),
#                    key=lambda x: OUTCOME_ORDER.get(x[0], 999)))
#         ws.merge_cells(start_row=2, start_column=1, end_row=2,
#                        end_column=max(len(cols_use), 1))
#         bc = ws.cell(2, 1)
#         bc.value = bk_str; bc.font = Font(italic=True, color="1A5276")
#         bc.fill = AI_QUAL_FILL
#         bc.alignment = Alignment(horizontal="left", vertical="center")
#         bc.border = BOX; ws.row_dimensions[2].height = 18
#         header_row = 3

#     for ci, h in enumerate(cols_use):
#         hfill = (AI_COL_FILL
#                  if h in ("CallerAI Qualified",
#                            "Business Disposition Confidence",
#                            "Business Disposition")
#                  else HEADER_FILL)
#         _cell(ws, header_row, ci + 1, h, bold=True, fill=hfill)

#     for ri, (_, row) in enumerate(df_aiq[cols_use].iterrows(),
#                                   start=header_row + 1):
#         row_fill = seg_fills.get(row.get("Source", ""), AI_QUAL_FILL)
#         for ci, col in enumerate(cols_use):
#             cell = ws.cell(ri, ci + 1)
#             cell.value = row[col]; cell.font = Font()
#             cell.alignment = Alignment(horizontal="left", vertical="center")
#             cell.border = BOX; cell.fill = row_fill

#     for ci, col in enumerate(cols_use, start=1):
#         ws.column_dimensions[
#             openpyxl.utils.get_column_letter(ci)].width = w_map.get(col, 16)
#     return ws

# def write_ai_answered_tab(wb, df, sheet_name="AI Answered"):
#     _mask = df["Disconnect Reason"].fillna("").str.strip().str.lower().isin(
#         {r.lower() for r in CALLERAI_ANSWERED_REASONS})
#     df_ans = df[_mask].copy()
#     for col in ["Calling Outcome", "CallerAI Qualified", "Business Disposition",
#                 "Business Disposition Confidence",
#                 "Lead Owner", "Activity Added By", "Plan Name",
#                 "Disconnect Reason"]:
#         if col in df_ans.columns: df_ans[col] = df_ans[col].fillna("")
#     if "Revenue" in df_ans.columns:
#         df_ans["Revenue"] = pd.to_numeric(df_ans["Revenue"], errors="coerce").fillna(0)

#     df_ans["_ps"] = df_ans["Paid Check"].map({"Paid": 0, "Unpaid": 1}).fillna(1)
#     df_ans.sort_values(["_ps", "Date"], inplace=True)
#     df_ans.drop(columns=["_ps"], inplace=True)
#     df_ans.reset_index(drop=True, inplace=True)

#     paid_c   = int((df_ans["Paid Check"] == "Paid").sum())
#     ai_q_c   = int((df_ans["CallerAI Qualified"] == "Yes").sum())
#     cols_use = [c for c in DETAIL_COLS if c in df_ans.columns]
#     w_map    = {c: w for c, w in zip(DETAIL_COLS, DETAIL_WIDTHS)}

#     ws = wb.create_sheet(title=sheet_name[:31])
#     ws.sheet_view.showGridLines = False

#     ws.merge_cells(start_row=1, start_column=1, end_row=1,
#                    end_column=max(len(cols_use), 1))
#     tc = ws.cell(1, 1)
#     tc.value = (f"AI Answered (user_request | CLIENT_INITIATED | "
#                 f"inactivity_timeout | transferred)"
#                 f"  --  Total: {len(df_ans)}  |  Paid: {paid_c}  |  "
#                 f"AI Qual: {ai_q_c}  |  {src_count_str(df_ans)}")
#     tc.font = Font(bold=True, color="FFFFFF")
#     tc.fill = PatternFill("solid", fgColor="1A7846")
#     tc.alignment = Alignment(horizontal="left", vertical="center")
#     ws.row_dimensions[1].height = 22

#     for ci, h in enumerate(cols_use):
#         hfill = (AI_COL_FILL
#                  if h in ("CallerAI Qualified",
#                            "Business Disposition Confidence",
#                            "Business Disposition")
#                  else HEADER_FILL)
#         _cell(ws, 2, ci + 1, h, bold=True, fill=hfill)

#     for ri, (_, row) in enumerate(df_ans[cols_use].iterrows(), start=3):
#         is_paid  = (str(row.get("Paid Check", "")) == "Paid")
#         is_ai_q  = (str(row.get("CallerAI Qualified", "")) == "Yes")
#         row_fill = (PAID_FILL   if is_paid
#                     else AI_QUAL_FILL if is_ai_q
#                     else seg_fills.get(row.get("Source", ""), None))
#         for ci, col in enumerate(cols_use):
#             cell = ws.cell(ri, ci + 1)
#             cell.value = row[col]; cell.font = Font(bold=is_paid)
#             cell.alignment = Alignment(horizontal="left", vertical="center")
#             cell.border = BOX
#             if row_fill: cell.fill = row_fill

#     for ci, col in enumerate(cols_use, start=1):
#         ws.column_dimensions[
#             openpyxl.utils.get_column_letter(ci)].width = w_map.get(col, 16)
#     return ws


# # ══════════════════════════════════════════════════════════════════════════════
# # STEP 7d — Per-campaign Purchased tabs + No LSQ Dispo tab
# # ══════════════════════════════════════════════════════════════════════════════
# def write_campaign_purchased_tab(wb, df_campaign, sheet_name, title):
#     df_p = df_campaign[df_campaign["Paid Check"] == "Paid"].copy()
#     for col in ["Calling Outcome", "Business Disposition",
#                 "Business Disposition Confidence",
#                 "Lead Owner", "Activity Added By", "Plan Name",
#                 "Disconnect Reason"]:
#         if col in df_p.columns: df_p[col] = df_p[col].fillna("")
#     if "Revenue" in df_p.columns:
#         df_p["Revenue"] = pd.to_numeric(df_p["Revenue"], errors="coerce").fillna(0)
#     if "Revenue_Attributed" in df_p.columns:
#         df_p["Revenue_Attributed"] = (
#             pd.to_numeric(df_p["Revenue_Attributed"], errors="coerce").fillna(0))
#     df_p.sort_values("Date", inplace=True, kind="stable")
#     df_p.reset_index(drop=True, inplace=True)

#     total_rev = (float(df_p["Revenue_Attributed"].sum())
#                  if "Revenue_Attributed" in df_p.columns
#                  else float(df_p["Revenue"].sum())
#                  if "Revenue" in df_p.columns else 0.0)

#     cols_use  = [c for c in DETAIL_COLS if c in df_p.columns]
#     w_map     = {c: w for c, w in zip(DETAIL_COLS, DETAIL_WIDTHS)}
#     PURCH_HDR = PatternFill("solid", fgColor="1F7743")

#     ws = wb.create_sheet(title=re.sub(r'[:\\/?*\[\]]', '', sheet_name)[:31])
#     ws.sheet_view.showGridLines = False

#     ws.merge_cells(start_row=1, start_column=1, end_row=1,
#                    end_column=max(len(cols_use), 1))
#     tc = ws.cell(1, 1)
#     tc.value = (f"{title}  --  Total: {len(df_p)}  |  "
#                 f"Attributed Revenue: Rs {total_rev:,.0f}  |  {src_count_str(df_p)}")
#     tc.font = Font(bold=True, color="FFFFFF"); tc.fill = PURCH_HDR
#     tc.alignment = Alignment(horizontal="left", vertical="center")
#     ws.row_dimensions[1].height = 22

#     header_row = 2
#     if "Plan Name" in df_p.columns:
#         _plans  = df_p["Plan Name"].replace("", "(No Plan)").value_counts()
#         bk_str  = "  |  ".join(f"{k}: {v}" for k, v in _plans.head(8).items())
#         ws.merge_cells(start_row=2, start_column=1, end_row=2,
#                        end_column=max(len(cols_use), 1))
#         bc = ws.cell(2, 1)
#         bc.value = bk_str; bc.font = Font(italic=True, color="1A5226")
#         bc.fill = PAID_FILL
#         bc.alignment = Alignment(horizontal="left", vertical="center")
#         bc.border = BOX; ws.row_dimensions[2].height = 18
#         header_row = 3

#     for ci, h in enumerate(cols_use):
#         _cell(ws, header_row, ci + 1, h, bold=True, fill=HEADER_FILL)

#     for ri, (_, row) in enumerate(df_p[cols_use].iterrows(),
#                                   start=header_row + 1):
#         for ci, col in enumerate(cols_use):
#             cell = ws.cell(ri, ci + 1)
#             cell.value = row[col]; cell.font = Font(bold=True)
#             cell.alignment = Alignment(horizontal="left", vertical="center")
#             cell.border = BOX; cell.fill = PAID_FILL

#     for ci, col in enumerate(cols_use, start=1):
#         ws.column_dimensions[
#             openpyxl.utils.get_column_letter(ci)].width = w_map.get(col, 16)
#     return ws


# def write_no_lsq_dispo_tab(wb, df, sheet_name="No LSQ Dispo Dump"):
#     df_no = df[
#         (df["Paid Check"] != "Paid") &
#         (df["Calling Outcome"].isna() |
#          (df["Calling Outcome"].astype(str).str.strip() == ""))
#     ].copy()
#     for col in ["Calling Outcome", "CallerAI Qualified", "Business Disposition",
#                 "Business Disposition Confidence", "Lead Owner",
#                 "Activity Added By", "Plan Name", "Disconnect Reason"]:
#         if col in df_no.columns: df_no[col] = df_no[col].fillna("")
#     if "Revenue" in df_no.columns:
#         df_no["Revenue"] = pd.to_numeric(df_no["Revenue"], errors="coerce").fillna(0)

#     df_no["_qs"]   = df_no["CallerAI Qualified"].map({"Yes": 0, "No": 1}).fillna(1)
#     df_no["_srcs"] = df_no["Source"].map({
#         WINBACK_SOURCE: 0, TTC_UNC_SOURCE: 1,
#         TTC_SOURCE: 2,     RENEWAL_TTC_SOURCE: 3
#     }).fillna(4)
#     df_no.sort_values(["_qs", "_srcs", "Date"], inplace=True)
#     df_no.drop(columns=["_qs", "_srcs"], inplace=True)
#     df_no.reset_index(drop=True, inplace=True)

#     total  = len(df_no)
#     ai_q_c = (int((df_no["CallerAI Qualified"] == "Yes").sum())
#                if "CallerAI Qualified" in df_no.columns else 0)
#     cols_use    = [c for c in DETAIL_COLS if c in df_no.columns]
#     w_map       = {c: w for c, w in zip(DETAIL_COLS, DETAIL_WIDTHS)}
#     NO_DISP_HDR = PatternFill("solid", fgColor="7B3F00")

#     ws = wb.create_sheet(title=sheet_name[:31])
#     ws.sheet_view.showGridLines = False

#     ws.merge_cells(start_row=1, start_column=1, end_row=1,
#                    end_column=max(len(cols_use), 1))
#     tc = ws.cell(1, 1)
#     tc.value = (f"No LSQ Disposition Dump  --  Total: {total}  |  "
#                 f"AI Qualified (no follow-up): {ai_q_c}  |  {src_count_str(df_no)}")
#     tc.font = Font(bold=True, color="FFFFFF"); tc.fill = NO_DISP_HDR
#     tc.alignment = Alignment(horizontal="left", vertical="center")
#     ws.row_dimensions[1].height = 22

#     for ci, h in enumerate(cols_use):
#         hfill = (AI_COL_FILL
#                  if h in ("CallerAI Qualified",
#                            "Business Disposition Confidence",
#                            "Business Disposition")
#                  else HEADER_FILL)
#         _cell(ws, 2, ci + 1, h, bold=True, fill=hfill)

#     for ri, (_, row) in enumerate(df_no[cols_use].iterrows(), start=3):
#         is_ai_q  = (str(row.get("CallerAI Qualified", "")) == "Yes")
#         row_fill = (AI_QUAL_FILL if is_ai_q
#                     else seg_fills.get(row.get("Source", ""), None))
#         for ci, col in enumerate(cols_use):
#             cell = ws.cell(ri, ci + 1)
#             cell.value = row[col]; cell.font = Font()
#             cell.alignment = Alignment(horizontal="left", vertical="center")
#             cell.border = BOX
#             if row_fill: cell.fill = row_fill

#     for ci, col in enumerate(cols_use, start=1):
#         ws.column_dimensions[
#             openpyxl.utils.get_column_letter(ci)].width = w_map.get(col, 16)
#     return ws


# write_ai_qual_segment_tab(wb, "Winback AI Qual",     out_win,
#                           "Winback -- AI Qualified Leads")
# write_ai_qual_segment_tab(wb, "TTC+Unc AI Qual",     out_ttc_unc,
#                           "TTC+Uncalled -- AI Qualified Leads")
# write_ai_qual_segment_tab(wb, "TTC AI Qual",         out_ttc,
#                           "TTC -- AI Qualified Leads")
# write_ai_qual_segment_tab(wb, "Renewal TTC AI Qual", out_renewal_ttc,
#                           "Renewal_TTC -- AI Qualified Leads")
# write_ai_qual_segment_tab(wb, "Overall AI Qual",     out_all,
#                           "Overall -- AI Qualified Leads (All Campaigns)")
# write_ai_answered_tab(wb, out_all, "AI Answered")

# write_campaign_purchased_tab(wb, out_ttc_unc,     "TTC+Unc Purchased",
#                              "TTC+Uncalled -- Purchased Leads")
# write_campaign_purchased_tab(wb, out_ttc,         "TTC Purchased",
#                              "TTC -- Purchased Leads")
# write_campaign_purchased_tab(wb, out_win,         "Winback Purchased",
#                              "Winback -- Purchased Leads")
# write_campaign_purchased_tab(wb, out_renewal_ttc, "Renewal TTC Purchased",
#                              "Renewal_TTC -- Purchased Leads")

# write_no_lsq_dispo_tab(wb, out_all, "No LSQ Dispo Dump")

# tab_counts = detail_df["_tab"].value_counts()
# print("\n-- Detail tab assignment ------------------------------------------------")
# for t in PRIORITY_TAB_ORDER:
#     if t in tab_counts.index:
#         print(f"   {t:<30} {tab_counts[t]:>5}")
# print(f"   {'TOTAL':<30} {tab_counts.sum():>5}")

# print(f"\n-- AI Qualified tabs (incl. Paid) ---------------------------------------")
# for seg_nm, seg_df in [
#     (WINBACK_SOURCE,      out_win),
#     (TTC_UNC_SOURCE,      out_ttc_unc),
#     (TTC_SOURCE,          out_ttc),
#     (RENEWAL_TTC_SOURCE,  out_renewal_ttc),
#     ("Overall",           out_all),
# ]:
#     n_total = int((seg_df["CallerAI Qualified"] == "Yes").sum())
#     n_paid  = int(((seg_df["CallerAI Qualified"] == "Yes") &
#                    (seg_df["Paid Check"] == "Paid")).sum())
#     print(f"   {seg_nm:<30} {n_total:>5}  (incl. {n_paid} Paid)")
# _ans_c = int(out_all["Disconnect Reason"].fillna("").str.strip().str.lower()
#              .isin({r.lower() for r in CALLERAI_ANSWERED_REASONS}).sum())
# print(f"   {'AI Answered':<30} {_ans_c:>5}")


# # ══════════════════════════════════════════════════════════════════════════════
# # STEP 8 — Save Leads_Summary_Report.xlsx  (local artifact for GHA)
# # ══════════════════════════════════════════════════════════════════════════════
# OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
# wb.save(SUMMARY_LOCAL)
# print(f"\nOK Leads_Summary_Report.xlsx -> {SUMMARY_LOCAL}")


# # ══════════════════════════════════════════════════════════════════════════════
# # STEP 9 — Build Leads_Summary_Matrix.xlsx
# # ══════════════════════════════════════════════════════════════════════════════
# wb_m = openpyxl.Workbook()
# ws_s = wb_m.active
# ws_s.title = "Summary"
# ws_s.sheet_view.showGridLines = False

# nr = write_summary_matrix(ws_s, f"Today  {today_str}", summary_today,   1, 1)
# nr += 2
# nr = write_summary_matrix(ws_s, "Overall", summary_overall, nr, 1)
# nr += 2

# nr_1 = write_ai_qual_funnel_table(ws_s, "Overall AI Qualified Funnel",
#                                   t_aiq_all,         nr, 1)
# nr_2 = write_ai_qual_funnel_table(ws_s, "Winback AI Qualified Funnel",
#                                   t_aiq_win,         nr, 5)
# nr_3 = write_ai_qual_funnel_table(ws_s, "TTC+Uncalled AI Qualified Funnel",
#                                   t_aiq_ttc_unc,     nr, 9)
# nr_4 = write_ai_qual_funnel_table(ws_s, "TTC AI Qualified Funnel",
#                                   t_aiq_ttc,         nr, 13)
# nr_5 = write_ai_qual_funnel_table(ws_s, "Renewal_TTC AI Qualified Funnel",
#                                   t_aiq_renewal_ttc, nr, 17)
# nr   = max(nr_1, nr_2, nr_3, nr_4, nr_5) + 2

# nr = write_planwise_block(
#     ws_s, f"Planwise Today  {today_str}",
#     build_planwise_rows(out_win_today, out_ttc_unc_today,
#                         out_ttc_today, out_renewal_ttc_today), nr, 1)
# nr += 2
# nr = write_planwise_block(
#     ws_s, "Planwise Overall",
#     build_planwise_rows(out_win, out_ttc_unc, out_ttc, out_renewal_ttc), nr, 1)
# nr += 2

# nr = write_plan_revenue_table(
#     ws_s, "Overall Plan-wise Revenue (All Campaigns)", t_plan_rev, nr, 1)
# nr += 2
# write_plan_revenue_table(
#     ws_s, f"Today's Plan-wise Revenue ({today_str})", t_plan_rev_today, nr, 1)

# for col, w in {"A": 22, "B": 14, "C": 16, "D": 14, "E": 14, "F": 14, "G": 14,
#                "H": 14, "I": 14, "J": 14, "K": 14, "L": 14, "M": 14, "N": 14,
#                "O": 14, "P": 14, "Q": 14, "R": 14, "S": 14, "T": 14}.items():
#     ws_s.column_dimensions[col].width = w

# # Matrix detail tabs
# base_df = out.drop(columns=["_date_ts"], errors="ignore").copy()
# _is_purch   = base_df["Paid Check"] == "Paid"
# _is_aiq     = base_df["CallerAI Qualified"] == "Yes"
# _has_nodisp = base_df["Calling Outcome"].isna()
# _is_ttcsrc  = base_df["Source"] == TTC_SOURCE

# df_purchased_m  = base_df[_is_purch].copy()
# df_qualified_m  = base_df[~_is_purch & _is_aiq].copy()
# df_ttc_m        = base_df[~_is_purch & ~_is_aiq &
#                            (base_df["Calling Outcome"] == "Trying to Contact")].copy()
# df_unanswered_m = base_df[~_is_purch & ~_is_aiq & _has_nodisp].copy()
# df_uncalled_m   = base_df[~_is_purch & ~_is_aiq & _has_nodisp & _is_ttcsrc].copy()

# def write_matrix_detail_sheet(wb, sheet_title, df_sub, header_title):
#     ws = wb.create_sheet(title=sheet_title[:31])
#     ws.sheet_view.showGridLines = False
#     df_s = df_sub.copy()
#     for col in ["Calling Outcome", "CallerAI Qualified", "Business Disposition",
#                 "Business Disposition Confidence", "Lead Owner",
#                 "Activity Added By", "Plan Name", "Disconnect Reason"]:
#         if col in df_s.columns: df_s[col] = df_s[col].fillna("")
#     if "Revenue" in df_s.columns:
#         df_s["Revenue"] = pd.to_numeric(df_s["Revenue"], errors="coerce").fillna(0)

#     sort_keys, sort_asc = [], []
#     if "Paid Check" in df_s.columns:
#         df_s["_ps"] = df_s["Paid Check"].map({"Paid": 0, "Unpaid": 1}).fillna(1)
#         sort_keys.append("_ps"); sort_asc.append(True)
#     if "CallerAI Qualified" in df_s.columns:
#         df_s["_qs"] = df_s["CallerAI Qualified"].map({"Yes": 0, "No": 1}).fillna(1)
#         sort_keys.append("_qs"); sort_asc.append(True)
#     if "Date" in df_s.columns:
#         sort_keys.append("Date"); sort_asc.append(False)
#     if sort_keys:
#         df_s.sort_values(sort_keys, ascending=sort_asc, inplace=True)
#     df_s.drop(columns=[c for c in ["_ps", "_qs"] if c in df_s.columns],
#               inplace=True)
#     df_s.reset_index(drop=True, inplace=True)

#     paid_c = (int((df_s["Paid Check"] == "Paid").sum())
#                if "Paid Check" in df_s.columns else 0)
#     ai_q_c = (int((df_s["CallerAI Qualified"] == "Yes").sum())
#                if "CallerAI Qualified" in df_s.columns else 0)
#     cols_use = [c for c in DETAIL_COLS if c in df_s.columns]
#     w_map    = {c: w for c, w in zip(DETAIL_COLS, DETAIL_WIDTHS)}

#     ws.merge_cells(start_row=1, start_column=1, end_row=1,
#                    end_column=max(len(cols_use), 1))
#     tc = ws.cell(1, 1)
#     tc.value = (f"{header_title}  --  Total: {len(df_s)}  |  Paid: {paid_c}  |  "
#                 f"AI Qualified: {ai_q_c}  |  {src_count_str(df_s)}")
#     tc.font = Font(bold=True, color="FFFFFF"); tc.fill = TITLE_FILL
#     tc.alignment = Alignment(horizontal="left", vertical="center")
#     ws.row_dimensions[1].height = 22

#     for ci, h in enumerate(cols_use):
#         hfill = (AI_COL_FILL
#                  if h in ("CallerAI Qualified",
#                            "Business Disposition Confidence",
#                            "Business Disposition")
#                  else HEADER_FILL)
#         _cell(ws, 2, ci + 1, h, bold=True, fill=hfill)

#     for ri, (_, row) in enumerate(df_s[cols_use].iterrows(), start=3):
#         is_paid  = (str(row.get("Paid Check", "")) == "Paid")
#         is_ai_q  = (str(row.get("CallerAI Qualified", "")) == "Yes")
#         row_fill = (PAID_FILL   if is_paid
#                     else AI_QUAL_FILL if is_ai_q
#                     else seg_fills.get(row.get("Source", ""), None))
#         for ci, col in enumerate(cols_use):
#             cell = ws.cell(ri, ci + 1)
#             cell.value = row[col]; cell.font = Font(bold=is_paid)
#             cell.alignment = Alignment(horizontal="left", vertical="center")
#             cell.border = BOX
#             if row_fill: cell.fill = row_fill

#     for ci, col in enumerate(cols_use, start=1):
#         ws.column_dimensions[
#             openpyxl.utils.get_column_letter(ci)].width = w_map.get(col, 16)
#     return ws

# write_matrix_detail_sheet(wb_m, "Purchased",
#                           df_purchased_m,  "Purchased Leads")
# write_matrix_detail_sheet(wb_m, "Qualified",
#                           df_qualified_m,  "AI Qualified (excl. Purchased)")
# write_matrix_detail_sheet(wb_m, "Trying to Contact",
#                           df_ttc_m,        "Trying to Contact (excl. Purchased & AI Qualified)")
# write_matrix_detail_sheet(wb_m, "Unanswered",
#                           df_unanswered_m, "Unanswered -- No Disposition (excl. Purchased & AI Qualified)")
# write_matrix_detail_sheet(wb_m, "Uncalled",
#                           df_uncalled_m,   "Uncalled -- TTC Campaign, No LSQ Disposition")

# wb_m.save(SUMMARY_MATRIX_LOCAL)
# print(f"OK Leads_Summary_Matrix.xlsx -> {SUMMARY_MATRIX_LOCAL}")
# print(f"   Purchased ({len(df_purchased_m)})  |  AI Qualified ({len(df_qualified_m)})  |  "
#       f"TTC ({len(df_ttc_m)})  |  Unanswered ({len(df_unanswered_m)})  |  "
#       f"Uncalled ({len(df_uncalled_m)})")


# # ══════════════════════════════════════════════════════════════════════════════
# # STEP 10 — Write Lead Summary Report to Google Sheet  (v21 new step)
# # ══════════════════════════════════════════════════════════════════════════════
# write_summary_to_gsheet(
#     gc, summary_today, summary_overall, today_str,
#     t_aiq_all, t_aiq_win, t_aiq_ttc_unc, t_aiq_ttc, t_aiq_renewal_ttc,
#     t_plan_rev
# )


# # ══════════════════════════════════════════════════════════════════════════════
# # Final console summary
# # ══════════════════════════════════════════════════════════════════════════════
# print(f"\n-- Final Summary --------------------------------------------------------")
# print(f"   Today  {today_str}:")
# for row in summary_today:
#     rev = row.get("AI Qual & Paid Rev", 0) + row.get("AI Unqual & Paid Rev", 0)
#     print(f"     {row['Segment']:<22}: {row['Attempted']:>6} attempted | "
#           f"{row['AI Qualified']:>5} AI Qual | {row['Paid']:>5} paid | "
#           f"Rs {rev:,} attributed")
# # Grand total today
# gt_t = build_grand_total_row(summary_today)
# rev_t = gt_t.get("AI Qual & Paid Rev", 0) + gt_t.get("AI Unqual & Paid Rev", 0)
# print(f"     {'GRAND TOTAL':<22}: {gt_t['Attempted']:>6} attempted | "
#       f"{gt_t['AI Qualified']:>5} AI Qual | {gt_t['Paid']:>5} paid | "
#       f"Rs {rev_t:,} attributed")

# print(f"   Overall:")
# for row in summary_overall:
#     rev = row.get("AI Qual & Paid Rev", 0) + row.get("AI Unqual & Paid Rev", 0)
#     print(f"     {row['Segment']:<22}: {row['Attempted']:>6} attempted | "
#           f"{row['AI Qualified']:>5} AI Qual | {row['Paid']:>5} paid | "
#           f"Rs {rev:,} attributed")
# # Grand total overall
# gt_o = build_grand_total_row(summary_overall)
# rev_o = gt_o.get("AI Qual & Paid Rev", 0) + gt_o.get("AI Unqual & Paid Rev", 0)
# print(f"     {'GRAND TOTAL':<22}: {gt_o['Attempted']:>6} attempted | "
#       f"{gt_o['AI Qualified']:>5} AI Qual | {gt_o['Paid']:>5} paid | "
#       f"Rs {rev_o:,} attributed")

# print("\nDone!")
# ==============================================================================
#  Leads <-> Sales Activity Merger + Summary Report  (GitHub Actions)
#  v22 -- Added Broker segment (Campaign 5).
# ==============================================================================
# Changes from v20:
#  (1) SUMMARY_METRIC_COLS reordered to:
#        Segment | Attempted | Company Answered | Our Answered | AI Qualified |
#        Paid | AI Qual & Paid | AI Qual & Paid Rev | Answered & Paid |
#        Ans & Paid Rev | Conv% (Attempted) | Conv% (Answered) |
#        AI Unqual & Paid | AI Unqual & Paid Rev
#  (2) Grand Total row added to Today's Summary + Overall Summary
#      (both in the XLSX tabs AND in the Google Sheets write).
#  (3) Whole numbers throughout:
#        Conv% -> round to nearest integer % (no ".1" decimal)
#        Revenue columns -> int (no ".0" suffix)
#  (4) Lead Summary Report written to Google Sheet via service account:
#        https://docs.google.com/spreadsheets/d/1rP_72DZtCWxRXlTuCsy1J38gaMJIeWn4FpH6fP1LBDE
#        First sheet (gid=0) is cleared and rewritten each run.
#  (5) GitHub Actions compatible:
#        - Auth via GOOGLE_SERVICE_ACCOUNT_JSON env var (GitHub Secret)
#        - No google.colab dependencies
#        - Output XLSX files saved to ./output/ (uploadable as GHA artifact)
#        - No Google Drive copy steps
# ==============================================================================
# requirements.txt:
#   gspread>=5.0
#   openpyxl
#   pandas
#   google-auth
# ==============================================================================
# GitHub workflow snippet:
#   - name: Run leads merger
#     env:
#       GOOGLE_SERVICE_ACCOUNT_JSON: ${{ secrets.GOOGLE_SERVICE_ACCOUNT_JSON }}
#     run: python leads_merger_v22.py
# ==============================================================================

import re, json, os
from collections import Counter
import pandas as pd
import gspread
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from google.oauth2.service_account import Credentials

# ── CallerAI Sheet ─────────────────────────────────────────────────────────────
CALLER_AI_SHEET_ID = "1pv8h-66WFdjJA91PR6p3y35nrUB4GiAQ2HEmicOlAOA"

TTC_UNC_TAB_NAME  = "TTC+Uncalled"
TTC_UNC_SOURCE    = "TTC+Uncalled"

TTC_TAB_NAME      = "TTC"
UNCALLED_TAB_NAME = "Uncalled"
TTC_SOURCE        = "TTC"

WINBACK_SOURCE = "Winback"
WINBACK_SHEETS = [
    (CALLER_AI_SHEET_ID,                             "Winback"),
    ("1EmTqHH5yfcrdk2QL58pXdru3bumnmuGB9rGbinyPmuQ", "Winback"),
]

RENEWAL_TTC_TAB_NAME = "Renewal_TTC"
RENEWAL_TTC_SOURCE   = "Renewal_TTC"

# v22: Broker (Campaign 5)
BROKER_SHEET_ID = "1EmTqHH5yfcrdk2QL58pXdru3bumnmuGB9rGbinyPmuQ"
BROKER_TAB_NAME = "Broker"
BROKER_SOURCE   = "Broker"
BROKER_DISPLAY  = "Broker"

WINBACK_DISPLAY      = "Winback"
TTC_UNC_DISPLAY      = "TTC+Uncalled (Set 1)"
TTC_DISPLAY          = "TTC+Uncalled (SET2)"
RENEWAL_TTC_DISPLAY  = "Renewal_TTC"

# v22: BROKER_SOURCE added at position 4
CAMP_CONCAT_ORDER = {
    TTC_UNC_SOURCE:     0,
    TTC_SOURCE:         1,
    WINBACK_SOURCE:     2,
    RENEWAL_TTC_SOURCE: 3,
    BROKER_SOURCE:      4,
}

OUTPUT_TAB_GID           = 1114202623
CALLER_AI_DISP_COL       = "Business Disposition"
CALLER_AI_SENTIMENT_COL  = "Business Disposition Confidence"
CALLER_AI_QUALIFIED_MAP  = {
    "interested": ["high", "medium", "low"],
    "callback":   ["high", "medium"],
    "followup":   ["high", "medium"],
}
CALLER_AI_DISP_ORDER = {"interested": 1, "callback": 2, "followup": 3}
CALLER_AI_CONF_ORDER = {"high": 1, "medium": 2, "low": 3}

DISP_NORMALIZE = {
    "callback requested":   "callback",
    "call back later":      "callback",
    "call back":            "callback",
    "called back later":    "callback",
    "follow-up required":   "followup",
    "followup required":    "followup",
    "follow up required":   "followup",
    "follow up":            "followup",
    "follow-up":            "followup",
    "information provided": "interested",
}

CALLERAI_DISCONNECT_COL   = "Disconnect Reason"
CALLERAI_ANSWERED_REASONS = {
    "user_request", "CLIENT_INITIATED", "inactivity_timeout", "transferred",
}
COMPANY_ANSWERED_REASONS  = {
    "user_request", "CLIENT_INITIATED", "inactivity_timeout", "transferred",
    "voicemail", "screening", "dnc_request",
}

LEADS_PHONE1_COL   = "Phone"
LEADS_DATE_COL     = "Start Time"
LEADS_CAMPAIGN_COL = "Call Type"
LEADS_STATUS_COL   = "Outcome"
LEADS_OWNER_COL    = "Agent Name"
LEADS_NAME_COL     = "Participant"

# v21: relative paths for GHA (no /content/ or Drive paths)
OUTPUT_DIR           = Path("output")
OUTPUT_LOCAL         = OUTPUT_DIR / "Leads_Final_Output.csv"
SUMMARY_LOCAL        = OUTPUT_DIR / "Leads_Summary_Report.xlsx"
SUMMARY_MATRIX_LOCAL = OUTPUT_DIR / "Leads_Summary_Matrix.xlsx"

# v22: SD_SHEETS replaces SD_SHEET_ID / SD_SHEET_GID
SD_SHEETS = [
    ("1Zf3GLNtI5nLsOfa8EBISqvYezhqgbAEuB6dR2isB4_E", 758486581),
    ("1Qt-n-QYnDykALFjYphDIIXdhOFbd7_ppqr6ZhmvXpNo",  0),
]
SD_PHONE_COL   = "Phone Number"
SD_DATE_COL    = "Activity Date"
SD_DISP1_COL   = "Disposition 1"
SD_DISP2_COL   = "Disposition 2"
SD_ADDED_BY    = "Activity Added By"
SD_FILTER_FROM = "2026-04-01"

SHEET_ID             = "1Kh2b4q5kqdy6oCvGlwArNphcLy6es8ZmHcrW_udvM3w"
SHEET_GID            = 473339099
NUMBER_COL_IDX       = 36
PLAN_COL_IDX         = 16
REVENUE_COL_IDX      = 28
ORDER_VALUE_COL_IDX  = 14
SALE_DATE_COL_IDX    = 15
TEST_PMT_CHK_COL_IDX = 33

TEST_PAYMENT_PLAN_NAMES = {
    "Tax Expert Onboarding",
    "Salaried Individual - IncomeTax Return Filing ( ITR )",
}

OUTCOME_ORDER = {
    "Purchased":                 1,
    "DND":                       2,
    "Invalid":                   3,
    "Not Interested":            4,
    "Awaiting payment":          5,
    "Will buy later":            6,
    "Qualified":                 7,
    "Trying to Contact":         8,
    "Inbound Call":              9,
    "AI Qualified":              10,
    "Answered - No Disposition": 11,
    "Not Contacted":             12,
}

DETAIL_TAB_PURCHASED  = "Purchased"
DETAIL_TAB_AI_QUAL    = "AI Qualified"
DETAIL_TAB_UNANSWERED = "Unanswered"

# v21: Primary cols first (as specified), remaining cols after Conv%
SUMMARY_METRIC_COLS = [
    "Attempted",
    "Company Answered",
    "Our Answered",
    "AI Qualified",
    "Paid",
    "AI Qual & Paid",
    "AI Qual & Paid Rev",
    "Answered & Paid",
    "Ans & Paid Rev",
    "Conv% (Attempted)",
    "Conv% (Answered)",
    "AI Unqual & Paid",
    "AI Unqual & Paid Rev",
]

# v22: 13 cols — added Broker Revenue / Broker Orders
PLAN_COLS = [
    "Plan Name", "Total Revenue (Rs)", "Total Orders",
    "Winback Revenue",     "Winback Orders",
    "TTC+Unc Revenue",     "TTC+Unc Orders",
    "TTC Revenue",         "TTC Orders",
    "Renewal TTC Revenue", "Renewal TTC Orders",
    "Broker Revenue",      "Broker Orders",
]

# v21: Target Google Sheet for Lead Summary Report
OUTPUT_SUMMARY_SHEET_ID = "1rP_72DZtCWxRXlTuCsy1J38gaMJIeWn4FpH6fP1LBDE"


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def clean_number(raw):
    if pd.isna(raw): return ""
    digits = re.sub(r"\D", "", str(raw))
    return digits[-10:] if len(digits) >= 10 else digits

def _is_qualified(row):
    disp = str(row[CALLER_AI_DISP_COL]).strip().lower()      if pd.notna(row.get(CALLER_AI_DISP_COL))      else ""
    sent = str(row[CALLER_AI_SENTIMENT_COL]).strip().lower() if pd.notna(row.get(CALLER_AI_SENTIMENT_COL)) else ""
    if sent == "":
        return "Yes" if disp in CALLER_AI_QUALIFIED_MAP else "No"
    return "Yes" if sent in CALLER_AI_QUALIFIED_MAP.get(disp, []) else "No"

def assign_tab(row):
    if str(row.get("Paid Check", "")) == "Paid":
        return DETAIL_TAB_PURCHASED
    if str(row.get("CallerAI Qualified", "No")) == "Yes":
        return DETAIL_TAB_AI_QUAL
    co = row.get("Calling Outcome", "")
    if pd.isna(co) or str(co).strip() in ("", "(blank)"):
        return DETAIL_TAB_UNANSWERED
    return str(co).strip()

def is_test_payment(plan_name):
    pn = str(plan_name).strip().lower()
    if pn.startswith("cleartax"):
        return True
    return pn in {n.strip().lower() for n in TEST_PAYMENT_PLAN_NAMES}

# v22: includes Broker
def src_count_str(df):
    w  = int((df["Source"] == WINBACK_SOURCE).sum())     if "Source" in df.columns else 0
    tu = int((df["Source"] == TTC_UNC_SOURCE).sum())     if "Source" in df.columns else 0
    t  = int((df["Source"] == TTC_SOURCE).sum())         if "Source" in df.columns else 0
    r  = int((df["Source"] == RENEWAL_TTC_SOURCE).sum()) if "Source" in df.columns else 0
    b  = int((df["Source"] == BROKER_SOURCE).sum())      if "Source" in df.columns else 0
    return (f"Winback: {w}  |  TTC+Unc: {tu}  |  TTC: {t}  |  "
            f"Renewal_TTC: {r}  |  Broker: {b}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Auth  (GitHub Actions: service account JSON from env var)
# ══════════════════════════════════════════════════════════════════════════════
_sa_json_str = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "")
if not _sa_json_str:
    raise EnvironmentError(
        "GOOGLE_SERVICE_ACCOUNT_JSON env var is not set.\n"
        "Add it as a GitHub Secret and expose it in your workflow:\n"
        "  env:\n"
        "    GOOGLE_SERVICE_ACCOUNT_JSON: ${{ secrets.GOOGLE_SERVICE_ACCOUNT_JSON }}"
    )
_sa_info = json.loads(_sa_json_str)
# Normalise private key — some secret stores escape newlines as \\n
if "private_key" in _sa_info:
    _sa_info["private_key"] = _sa_info["private_key"].replace("\\n", "\n").strip("\r")
_creds = Credentials.from_service_account_info(
    _sa_info,
    scopes=["https://www.googleapis.com/auth/spreadsheets"],
)
gc = gspread.authorize(_creds)
print("OK Auth done (service account)")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — Load CallerAI Leads (4 campaigns) + Broker (Campaign 5)
# ══════════════════════════════════════════════════════════════════════════════
sh_caller = gc.open_by_key(CALLER_AI_SHEET_ID)
tabs_map  = {ws.title: ws for ws in sh_caller.worksheets()}
print(f"OK CallerAI sheet -- available tabs: {list(tabs_map.keys())}")

def gsheet_tab_to_df(tabs_map, tab_name):
    if tab_name not in tabs_map:
        print(f"WARNING Tab '{tab_name}' not found. Available: {list(tabs_map.keys())}")
        return pd.DataFrame()
    rows = tabs_map[tab_name].get_all_values()
    if not rows:
        print(f"WARNING Tab '{tab_name}' is empty.")
        return pd.DataFrame()
    df = pd.DataFrame(rows[1:], columns=[c.strip() for c in rows[0]])
    print(f"   Loaded '{tab_name}': {len(df):,} rows  |  cols: {list(df.columns)}")
    return df

print("\n-- Campaign 1: TTC+Uncalled (single tab) --------------------------------")
df_ttc_unc_raw = gsheet_tab_to_df(tabs_map, TTC_UNC_TAB_NAME)

print(f"\n-- Campaign 2: TTC (merge '{TTC_TAB_NAME}' + '{UNCALLED_TAB_NAME}') ------")
df_ttc_tab = gsheet_tab_to_df(tabs_map, TTC_TAB_NAME)
df_unc_tab = gsheet_tab_to_df(tabs_map, UNCALLED_TAB_NAME)
df_ttc_raw = pd.concat([df_ttc_tab, df_unc_tab], ignore_index=True)
print(f"   TTC merged raw: {len(df_ttc_raw):,}  "
      f"({TTC_TAB_NAME}: {len(df_ttc_tab):,} + {UNCALLED_TAB_NAME}: {len(df_unc_tab):,})")

print("\n-- Campaign 3: Winback (single tab) -------------------------------------")
df_win_raw = gsheet_tab_to_df(tabs_map, WINBACK_SHEETS[0][1])

print("\n-- Campaign 4: Renewal_TTC (single tab) ---------------------------------")
df_renewal_ttc_raw = gsheet_tab_to_df(tabs_map, RENEWAL_TTC_TAB_NAME)

print("\n-- Campaign 5: Broker (separate sheet, 'Broker' tab) --------------------")
try:
    _brk_sh       = gc.open_by_key(BROKER_SHEET_ID)
    _brk_tabs     = {ws.title: ws for ws in _brk_sh.worksheets()}
    df_broker_raw = gsheet_tab_to_df(_brk_tabs, BROKER_TAB_NAME)
    print(f"   Broker sheet tabs: {list(_brk_tabs.keys())}")
except Exception as _brk_err:
    print(f"   WARNING: Could not load Broker sheet: {_brk_err}")
    df_broker_raw = pd.DataFrame()


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2b — Normalize, qualify, dedup each campaign independently
# ══════════════════════════════════════════════════════════════════════════════
def process_tab(df, source_name):
    if df.empty:
        return df
    df = df.copy()
    df["_source"] = source_name

    if CALLER_AI_DISP_COL in df.columns:
        df[CALLER_AI_DISP_COL] = (
            df[CALLER_AI_DISP_COL].astype(str).str.strip().str.lower()
            .map(lambda v: DISP_NORMALIZE.get(v, v))
        )
        _bd_str = df[CALLER_AI_DISP_COL].astype(str).str.strip().str.lower()
        _empty  = _bd_str.isin(["", "nan", "none"])
        if _empty.any() and LEADS_STATUS_COL in df.columns:
            _fallback = (df.loc[_empty, LEADS_STATUS_COL]
                         .astype(str).str.strip().str.lower()
                         .map(lambda v: DISP_NORMALIZE.get(v, v)))
            df.loc[_empty, CALLER_AI_DISP_COL] = _fallback
            print(f"   {source_name} -- {_empty.sum()} empty Business Dispositions filled from Lead Stage")

        df["CallerAI Qualified"] = df.apply(_is_qualified, axis=1)
        q = (df["CallerAI Qualified"] == "Yes").sum()
        print(f"   {source_name} -- CallerAI Qualified (pre-dedup): {q:,}")

        unique_disps = df[CALLER_AI_DISP_COL].dropna().unique()
        print(f"   {source_name} -- Unique dispositions: {sorted(str(d) for d in unique_disps)}")
    else:
        df["CallerAI Qualified"] = "No"
        print(f"   {source_name} -- Disposition col missing; Qualified=No for all")

    df["_clean_phone"] = (
        df[LEADS_PHONE1_COL].apply(clean_number) if LEADS_PHONE1_COL in df.columns else ""
    )

    if LEADS_DATE_COL in df.columns:
        _cleaned       = df[LEADS_DATE_COL].astype(str).str.replace(r'\s*\([^)]*\)\s*$', '', regex=True).str.strip()
        _parsed        = pd.to_datetime(_cleaned, utc=True, errors="coerce")
        df["_date_ts"] = _parsed.dt.tz_convert(None)
    else:
        df["_date_ts"] = pd.NaT

    pre_dedup = len(df)
    df["_disp_sort"] = (df[CALLER_AI_DISP_COL].astype(str).str.strip().str.lower()
                        .map(CALLER_AI_DISP_ORDER).fillna(999)) \
                        if CALLER_AI_DISP_COL in df.columns else 999
    df["_conf_sort"] = (df[CALLER_AI_SENTIMENT_COL].astype(str).str.strip().str.lower()
                        .map(CALLER_AI_CONF_ORDER).fillna(999)) \
                        if CALLER_AI_SENTIMENT_COL in df.columns else 999
    df["_date_sort"] = df["_date_ts"].fillna(pd.Timestamp.min)

    has_phone = df["_clean_phone"] != ""
    with_ph = (
        df[has_phone]
        .sort_values(["_clean_phone", "_disp_sort", "_conf_sort", "_date_sort"],
                     ascending=[True, True, True, False])
        .drop_duplicates(subset=["_clean_phone"], keep="first")
    )
    df = pd.concat([with_ph, df[~has_phone]], ignore_index=True)
    df.drop(columns=["_disp_sort", "_conf_sort", "_date_sort"], inplace=True)
    print(f"   {source_name} -- dedup: {pre_dedup:,} -> {len(df):,}  (removed {pre_dedup - len(df):,})")
    return df

print("\n-- Processing campaigns -------------------------------------------------")
df_ttc_unc     = process_tab(df_ttc_unc_raw,     TTC_UNC_SOURCE)
df_ttc         = process_tab(df_ttc_raw,         TTC_SOURCE)
df_win         = process_tab(df_win_raw,         WINBACK_SOURCE)
df_renewal_ttc = process_tab(df_renewal_ttc_raw, RENEWAL_TTC_SOURCE)
df_broker      = process_tab(df_broker_raw,      BROKER_SOURCE)

leads_raw = pd.concat([df_ttc_unc, df_ttc, df_win, df_renewal_ttc, df_broker], ignore_index=True)
print(f"\nOK Combined -- total: {len(leads_raw):,}  |  "
      f"{TTC_UNC_SOURCE}: {len(df_ttc_unc):,}  |  "
      f"{TTC_SOURCE}: {len(df_ttc):,}  |  "
      f"{WINBACK_SOURCE}: {len(df_win):,}  |  "
      f"{RENEWAL_TTC_SOURCE}: {len(df_renewal_ttc):,}  |  "
      f"{BROKER_SOURCE}: {len(df_broker):,}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — Paid Sheet
# v20: skip zero-revenue rows entirely (treat as test payments)
# ══════════════════════════════════════════════════════════════════════════════
sh_paid  = gc.open_by_key(SHEET_ID)
ws_paid  = next((w for w in sh_paid.worksheets() if w.id == SHEET_GID), sh_paid.worksheets()[0])
all_rows = ws_paid.get_all_values()

_today_ts = pd.Timestamp.today().normalize()

paid_set             = set()
paid_set_today       = set()
paid_plan_map        = {}
skipped_cleartax     = 0
skipped_not_genuine  = 0
skipped_zero_revenue = 0

for row in all_rows[1:]:
    if len(row) <= NUMBER_COL_IDX or not str(row[NUMBER_COL_IDX]).strip():
        continue
    phone = clean_number(row[NUMBER_COL_IDX])
    if not phone:
        continue

    plan = row[PLAN_COL_IDX].strip() if len(row) > PLAN_COL_IDX else ""
    if is_test_payment(plan):
        skipped_cleartax += 1
        continue

    test_chk = row[TEST_PMT_CHK_COL_IDX].strip() if len(row) > TEST_PMT_CHK_COL_IDX else ""
    if test_chk.lower() != "genuine":
        skipped_not_genuine += 1
        continue

    rev_raw = row[REVENUE_COL_IDX].strip() if len(row) > REVENUE_COL_IDX else ""
    try:    rev = float(re.sub(r"[^\d.]", "", rev_raw)) if rev_raw else 0.0
    except: rev = 0.0

    if rev == 0.0:
        ov_raw = row[ORDER_VALUE_COL_IDX].strip() if len(row) > ORDER_VALUE_COL_IDX else ""
        try:    rev = float(re.sub(r"[^\d.]", "", ov_raw)) if ov_raw else 0.0
        except: rev = 0.0

    if rev == 0.0:
        skipped_zero_revenue += 1
        continue

    paid_set.add(phone)

    if phone not in paid_plan_map or (not paid_plan_map[phone]["plan_name"] and plan):
        paid_plan_map[phone] = {"plan_name": plan, "revenue": rev}

    raw_sale = row[SALE_DATE_COL_IDX].strip() if len(row) > SALE_DATE_COL_IDX else ""
    if raw_sale:
        _sc = re.sub(r'\s*\([^)]*\)\s*$', '', raw_sale).strip()
        _st = pd.to_datetime(_sc, utc=True, errors="coerce")
        if pd.notna(_st) and _st.tz_convert(None).normalize() == _today_ts:
            paid_set_today.add(phone)

paid_set.discard("")
paid_set_today.discard("")
print(f"OK Paid sheet -- unique paid: {len(paid_set):,}  |  today: {len(paid_set_today):,}  "
      f"|  ClearTax/test skipped: {skipped_cleartax:,}  |  Not Genuine: {skipped_not_genuine:,}  "
      f"|  Zero revenue: {skipped_zero_revenue:,}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — LSQ Sales Disposition (v22: load from SD_SHEETS list)
# ══════════════════════════════════════════════════════════════════════════════
_sd_dfs = []
for _sd_id, _sd_gid in SD_SHEETS:
    try:
        _sd_url  = (f"https://docs.google.com/spreadsheets/d/{_sd_id}"
                    f"/export?format=csv&gid={_sd_gid}")
        _df_part = pd.read_csv(_sd_url, low_memory=False)
        _df_part.columns = _df_part.columns.str.strip()
        print(f"   SD sheet {_sd_id} gid={_sd_gid}: {len(_df_part):,} rows")
        _sd_dfs.append(_df_part)
    except Exception as _sd_e:
        print(f"   WARNING SD sheet {_sd_id} gid={_sd_gid}: {_sd_e}")
df_sd = pd.concat(_sd_dfs, ignore_index=True) if _sd_dfs else pd.DataFrame()
df_sd.columns = df_sd.columns.str.strip()
if SD_ADDED_BY not in df_sd.columns:
    raise KeyError(f"'{SD_ADDED_BY}' column missing in Sales Disposition sheets.")

df_sd["_act_date"]    = pd.to_datetime(df_sd[SD_DATE_COL], dayfirst=True, errors="coerce")
df_sd = df_sd[df_sd["_act_date"] >= pd.Timestamp(SD_FILTER_FROM)].copy()
df_sd["_clean_phone"] = df_sd[SD_PHONE_COL].apply(clean_number)
df_sd = df_sd[df_sd["_clean_phone"] != ""].copy()

d1 = df_sd[["_clean_phone", SD_DISP1_COL, SD_ADDED_BY]].rename(
    columns={SD_DISP1_COL: "_disp", SD_ADDED_BY: "_added_by"})
d2 = df_sd[["_clean_phone", SD_DISP2_COL, SD_ADDED_BY]].rename(
    columns={SD_DISP2_COL: "_disp", SD_ADDED_BY: "_added_by"})
all_disps = pd.concat([d1, d2], ignore_index=True)
all_disps = all_disps[all_disps["_disp"].notna() & (all_disps["_disp"].str.strip() != "")].copy()
all_disps["_order"] = all_disps["_disp"].map(OUTCOME_ORDER).fillna(999)
best_disp = (
    all_disps.sort_values("_order")
    .groupby("_clean_phone").first().reset_index()
    [["_clean_phone", "_disp", "_added_by"]]
    .rename(columns={"_clean_phone": "_merge_phone", "_disp": "Calling Outcome",
                     "_added_by": "_lsq_agent"})
)
print(f"OK LSQ loaded -- {len(df_sd):,} rows  |  {best_disp['_merge_phone'].nunique():,} unique phones")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 5 — Build merged DataFrame
# ══════════════════════════════════════════════════════════════════════════════
def safe_col(df, col, fallback=""):
    return df[col].values if col in df.columns else fallback

paid_check = ["Paid" if (n and n in paid_set) else "Unpaid"
              for n in leads_raw["_clean_phone"].values]

out = pd.DataFrame({
    "Date":                            pd.Series(leads_raw["_date_ts"].values).dt.strftime("%d/%m/%Y").fillna("").values,
    "Customer Number":                 leads_raw["_clean_phone"].values,
    "Lead Name":                       safe_col(leads_raw, LEADS_NAME_COL),
    "Type Of Lead":                    safe_col(leads_raw, LEADS_CAMPAIGN_COL),
    "Lead Stage":                      safe_col(leads_raw, LEADS_STATUS_COL),
    "Disconnect Reason":               safe_col(leads_raw, CALLERAI_DISCONNECT_COL),
    "Lead Owner":                      safe_col(leads_raw, LEADS_OWNER_COL),
    "Paid Check":                      paid_check,
    "CallerAI Qualified":              leads_raw["CallerAI Qualified"].values,
    "Business Disposition":            safe_col(leads_raw, CALLER_AI_DISP_COL),
    "Business Disposition Confidence": safe_col(leads_raw, CALLER_AI_SENTIMENT_COL),
    "Source":                          leads_raw["_source"].values,
    "_date_ts":                        pd.to_datetime(leads_raw["_date_ts"].values).normalize(),
})

out = out.merge(best_disp, left_on="Customer Number", right_on="_merge_phone", how="left")
out.drop(columns=["_merge_phone"], inplace=True, errors="ignore")

out.loc[out["Paid Check"] == "Paid", "Calling Outcome"] = "Purchased"
out.loc[(out["Calling Outcome"] == "Purchased") & (out["Paid Check"] == "Unpaid"), "Calling Outcome"] = pd.NA

out["Activity Added By"] = out.apply(
    lambda row: row["_lsq_agent"]
                if pd.notna(row["_lsq_agent"]) and str(row["_lsq_agent"]).strip() != ""
                else row.get("Lead Owner", ""), axis=1)
out.drop(columns=["_lsq_agent"], inplace=True)

out["Plan Name"] = out["Customer Number"].map(lambda p: paid_plan_map.get(p, {}).get("plan_name", ""))
out["Revenue"]   = out["Customer Number"].map(lambda p: paid_plan_map.get(p, {}).get("revenue", 0.0))
out["Type Of Lead"] = out["Type Of Lead"].fillna("(blank)").astype(str).str.strip().replace("", "(blank)")
out["Lead Stage"]   = out["Lead Stage"].fillna("(blank)").astype(str).str.strip().replace("", "(blank)")
out["Business Disposition Confidence"] = out["Business Disposition Confidence"].fillna("").astype(str).str.strip()
out["Business Disposition"]            = out["Business Disposition"].fillna("").astype(str).str.strip()

out = out[[
    "Date", "Customer Number", "Lead Name", "Type Of Lead", "Source",
    "Lead Stage", "Disconnect Reason", "Business Disposition",
    "CallerAI Qualified", "Business Disposition Confidence",
    "Calling Outcome", "Paid Check", "Lead Owner", "Activity Added By",
    "Plan Name", "Revenue", "_date_ts"
]]

paid_c    = (out["Paid Check"] == "Paid").sum()
ai_qual_c = (out["CallerAI Qualified"] == "Yes").sum()
purch_c   = (out["Calling Outcome"] == "Purchased").sum()
print(f"OK Merge done -- {len(out):,} leads  |  Paid: {paid_c}  |  Purchased: {purch_c}")
print(f"   CallerAI Qualified (all):        {ai_qual_c:,}")
print(f"   CallerAI Qualified (excl. Paid): "
      f"{int(((out['CallerAI Qualified']=='Yes') & (out['Paid Check']!='Paid')).sum()):,}")
print(f"   {TTC_UNC_SOURCE}: {(out['Source']==TTC_UNC_SOURCE).sum():,}  |  "
      f"{TTC_SOURCE}: {(out['Source']==TTC_SOURCE).sum():,}  |  "
      f"{WINBACK_SOURCE}: {(out['Source']==WINBACK_SOURCE).sum():,}  |  "
      f"{RENEWAL_TTC_SOURCE}: {(out['Source']==RENEWAL_TTC_SOURCE).sum():,}  |  "
      f"{BROKER_SOURCE}: {(out['Source']==BROKER_SOURCE).sum():,}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 6 — Save CSV (kept as GHA artifact; no Drive copy)
# ══════════════════════════════════════════════════════════════════════════════
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
out.drop(columns=["_date_ts"]).to_csv(OUTPUT_LOCAL, index=False)
print(f"OK CSV -> {OUTPUT_LOCAL}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 7 — Build summary data + XLSX report
# ══════════════════════════════════════════════════════════════════════════════
today_ts  = _today_ts
today_str = today_ts.strftime("%d-%m-%Y")

out_all = out.copy()

# ── v20: Revenue Attribution ──────────────────────────────────────────────────
_paid_mask_all = out_all["Paid Check"] == "Paid"
if _paid_mask_all.any():
    _paid_for_attr = out_all[_paid_mask_all].copy()
    _paid_for_attr["_disp_s"] = (_paid_for_attr["Business Disposition"]
                                  .str.strip().str.lower()
                                  .map(CALLER_AI_DISP_ORDER).fillna(999))
    _paid_for_attr["_conf_s"] = (_paid_for_attr["Business Disposition Confidence"]
                                  .str.strip().str.lower()
                                  .map(CALLER_AI_CONF_ORDER).fillna(999))
    _paid_for_attr["_date_s"] = _paid_for_attr["_date_ts"].fillna(pd.Timestamp.min)
    _paid_for_attr["_camp_s"] = _paid_for_attr["Source"].map(CAMP_CONCAT_ORDER).fillna(99)

    _rev_owner_df = (
        _paid_for_attr
        .sort_values(
            ["Customer Number", "_disp_s", "_conf_s", "_date_s", "_camp_s"],
            ascending=[True, True, True, False, True]
        )
        .drop_duplicates(subset=["Customer Number"], keep="first")
    )
    _rev_owner = dict(zip(_rev_owner_df["Customer Number"], _rev_owner_df["Source"]))
else:
    _rev_owner = {}

out_all["Revenue_Attributed"] = out_all.apply(
    lambda row: row["Revenue"]
    if (row["Paid Check"] == "Paid"
        and _rev_owner.get(row["Customer Number"]) == row["Source"])
    else 0.0,
    axis=1
)

if _rev_owner:
    _attr_counts    = Counter(_rev_owner.values())
    _attr_total_rev = out_all["Revenue_Attributed"].sum()
    print(f"OK Revenue attribution -- {len(_rev_owner):,} paid phones attributed  |  "
          + "  |  ".join(
              f"{src}: {_attr_counts.get(src, 0)}"
              for src in [TTC_UNC_SOURCE, TTC_SOURCE, WINBACK_SOURCE,
                          RENEWAL_TTC_SOURCE, BROKER_SOURCE])
          + f"  |  Total attributed Rs: {_attr_total_rev:,.0f}")

# Per-campaign DFs (built AFTER Revenue_Attributed so they inherit the column)
out_ttc_unc     = out_all[out_all["Source"] == TTC_UNC_SOURCE].copy()
out_ttc         = out_all[out_all["Source"] == TTC_SOURCE].copy()
out_win         = out_all[out_all["Source"] == WINBACK_SOURCE].copy()
out_renewal_ttc = out_all[out_all["Source"] == RENEWAL_TTC_SOURCE].copy()
out_broker      = out_all[out_all["Source"] == BROKER_SOURCE].copy()

# Today's per-campaign DFs
out_ttc_unc_today     = out_ttc_unc    [out_ttc_unc    ["_date_ts"] == today_ts].copy()
out_ttc_today         = out_ttc        [out_ttc        ["_date_ts"] == today_ts].copy()
out_win_today         = out_win        [out_win        ["_date_ts"] == today_ts].copy()
out_renewal_ttc_today = out_renewal_ttc[out_renewal_ttc["_date_ts"] == today_ts].copy()
out_broker_today      = out_broker     [out_broker     ["_date_ts"] == today_ts].copy()


# ── build_summary_row ─────────────────────────────────────────────────────────
def build_summary_row(df, label, today_paid_set=None):
    total = len(df)

    if CALLERAI_DISCONNECT_COL in df.columns:
        _dr = df[CALLERAI_DISCONNECT_COL].fillna("").str.strip().str.lower()
        company_answered = int(_dr.isin({r.lower() for r in COMPANY_ANSWERED_REASONS}).sum())
        our_answered     = int(_dr.isin({r.lower() for r in CALLERAI_ANSWERED_REASONS}).sum())
        _is_our_answered = _dr.isin({r.lower() for r in CALLERAI_ANSWERED_REASONS})
    else:
        our_answered = company_answered = int(df["Calling Outcome"].notna().sum())
        _is_our_answered = df["Calling Outcome"].notna()

    if today_paid_set is not None:
        _is_paid = df["Customer Number"].isin(today_paid_set)
    else:
        _is_paid = df["Paid Check"] == "Paid"

    paid           = int(_is_paid.sum())
    ai_qual_paid   = int((_is_paid & (df["CallerAI Qualified"] == "Yes")).sum())
    ai_unqual_paid = int((_is_paid & (df["CallerAI Qualified"] != "Yes")).sum())
    answered_paid  = int((_is_paid & _is_our_answered).sum())
    ai_q           = int((df["CallerAI Qualified"] == "Yes").sum())

    # v20: Revenue_Attributed already deduplicated — sum is safe
    if "Revenue_Attributed" in df.columns:
        _rev_a = pd.to_numeric(df["Revenue_Attributed"], errors="coerce").fillna(0)
        ai_qual_paid_rev   = int(_rev_a[_is_paid & (df["CallerAI Qualified"] == "Yes")].sum())
        ai_unqual_paid_rev = int(_rev_a[_is_paid & (df["CallerAI Qualified"] != "Yes")].sum())
        ans_paid_rev       = int(_rev_a[_is_paid & _is_our_answered].sum())
    else:
        ai_qual_paid_rev = ai_unqual_paid_rev = ans_paid_rev = 0

    # v21: whole-number percentages (no ".1" decimal)
    conv_att = f"{round(paid / total        * 100)}%" if total        else "0%"
    conv_ans = f"{round(paid / our_answered * 100)}%" if our_answered else "0%"

    return {
        "Segment":              label,
        "Attempted":            total,
        "Company Answered":     company_answered,
        "Our Answered":         our_answered,
        "AI Qualified":         ai_q,
        "Paid":                 paid,
        "AI Qual & Paid":       ai_qual_paid,
        "AI Unqual & Paid":     ai_unqual_paid,
        "Answered & Paid":      answered_paid,
        "AI Qual & Paid Rev":   ai_qual_paid_rev,   # v21: int
        "AI Unqual & Paid Rev": ai_unqual_paid_rev, # v21: int
        "Ans & Paid Rev":       ans_paid_rev,       # v21: int
        "Conv% (Attempted)":    conv_att,
        "Conv% (Answered)":     conv_ans,
    }


# v21: Grand Total row — sums numeric cols, recomputes Conv% from totals
def build_grand_total_row(rows):
    def _sum(key):
        return sum(r.get(key, 0) for r in rows
                   if isinstance(r.get(key, 0), (int, float)))

    total_att  = _sum("Attempted")
    total_our  = _sum("Our Answered")
    total_paid = _sum("Paid")

    conv_att = f"{round(total_paid / total_att  * 100)}%" if total_att  else "0%"
    conv_ans = f"{round(total_paid / total_our  * 100)}%" if total_our  else "0%"

    return {
        "Segment":              "Grand Total",
        "Attempted":            total_att,
        "Company Answered":     _sum("Company Answered"),
        "Our Answered":         total_our,
        "AI Qualified":         _sum("AI Qualified"),
        "Paid":                 total_paid,
        "AI Qual & Paid":       _sum("AI Qual & Paid"),
        "AI Unqual & Paid":     _sum("AI Unqual & Paid"),
        "Answered & Paid":      _sum("Answered & Paid"),
        "AI Qual & Paid Rev":   _sum("AI Qual & Paid Rev"),
        "AI Unqual & Paid Rev": _sum("AI Unqual & Paid Rev"),
        "Ans & Paid Rev":       _sum("Ans & Paid Rev"),
        "Conv% (Attempted)":    conv_att,
        "Conv% (Answered)":     conv_ans,
    }


# v22: 5 rows each (added Broker)
summary_today = [
    build_summary_row(out_win_today,         WINBACK_DISPLAY,     today_paid_set=paid_set_today),
    build_summary_row(out_ttc_unc_today,     TTC_UNC_DISPLAY,     today_paid_set=paid_set_today),
    build_summary_row(out_ttc_today,         TTC_DISPLAY,         today_paid_set=paid_set_today),
    build_summary_row(out_renewal_ttc_today, RENEWAL_TTC_DISPLAY, today_paid_set=paid_set_today),
    build_summary_row(out_broker_today,      BROKER_DISPLAY,      today_paid_set=paid_set_today),
]
summary_overall = [
    build_summary_row(out_win,         WINBACK_DISPLAY),
    build_summary_row(out_ttc_unc,     TTC_UNC_DISPLAY),
    build_summary_row(out_ttc,         TTC_DISPLAY),
    build_summary_row(out_renewal_ttc, RENEWAL_TTC_DISPLAY),
    build_summary_row(out_broker,      BROKER_DISPLAY),
]


# ── Pivot helpers ──────────────────────────────────────────────────────────────
def campaign_pivot(df):
    if df.empty:
        return pd.DataFrame(columns=["Paid", "Unpaid", "Grand Total", "%", "AI Qualified"])
    df = df.copy()
    df["Type Of Lead"] = df["Type Of Lead"].fillna("(blank)").astype(str).str.strip().replace("", "(blank)")
    piv = (df.groupby("Type Of Lead", dropna=False)["Paid Check"]
           .value_counts().unstack(fill_value=0)
           .reindex(columns=["Paid", "Unpaid"], fill_value=0))
    piv["Grand Total"] = piv["Paid"] + piv["Unpaid"]
    piv.sort_values("Grand Total", ascending=False, inplace=True)
    piv["%"] = [f"{round(p/t*100)}%" if t else "0%"
                for p, t in zip(piv["Paid"], piv["Grand Total"])]
    qual_counts = df.groupby("Type Of Lead")["CallerAI Qualified"].apply(
        lambda x: (x == "Yes").sum())
    piv["AI Qualified"] = qual_counts.reindex(piv.index, fill_value=0)
    tp, tu, tt, tq = (piv["Paid"].sum(), piv["Unpaid"].sum(),
                      piv["Grand Total"].sum(), piv["AI Qualified"].sum())
    return pd.concat([piv, pd.DataFrame(
        [{"Paid": tp, "Unpaid": tu, "Grand Total": tt,
          "%": f"{round(tp/tt*100)}%" if tt else "0%", "AI Qualified": tq}],
        index=["Grand Total"])])

def ai_qualified_funnel_pivot(df):
    df_aiq = df[df["CallerAI Qualified"] == "Yes"].copy()
    if df_aiq.empty:
        return pd.DataFrame([{"LSQ Disposition": "(No AI Qualified leads)", "Count": 0, "%": "-"}])
    co = df_aiq["Calling Outcome"].fillna("No LSQ Disposition").astype(str).str.strip()
    co = co.replace("", "No LSQ Disposition")
    counts = co.value_counts().reset_index()
    counts.columns = ["LSQ Disposition", "Count"]
    total = counts["Count"].sum()
    counts["%"] = [f"{round(c/total*100)}%" for c in counts["Count"]]
    counts["_sort"] = counts["LSQ Disposition"].map(OUTCOME_ORDER).fillna(999)
    counts.sort_values(["_sort", "Count"], ascending=[True, False], inplace=True)
    counts.drop(columns=["_sort"], inplace=True)
    counts.reset_index(drop=True, inplace=True)
    return pd.concat([counts, pd.DataFrame([{
        "LSQ Disposition": "Grand Total", "Count": total, "%": ""}])], ignore_index=True)

def date_pivot(df):
    if df.empty: return pd.DataFrame()
    def _our_ans(grp_df):
        if CALLERAI_DISCONNECT_COL not in grp_df.columns: return 0
        return int(grp_df[CALLERAI_DISCONNECT_COL].fillna("").str.strip().str.lower()
                   .isin({r.lower() for r in CALLERAI_ANSWERED_REASONS}).sum())
    grp = df.groupby("Date")
    result = pd.DataFrame({
        "Total":        grp.size(),
        "Paid":         grp["Paid Check"].apply(lambda x: (x == "Paid").sum()),
        "Unpaid":       grp["Paid Check"].apply(lambda x: (x == "Unpaid").sum()),
        "Our Answered": grp.apply(_our_ans, include_groups=False),
        "AI Qualified": grp["CallerAI Qualified"].apply(lambda x: (x == "Yes").sum()),
    })
    result["Not Answered"]       = result["Total"] - result["Our Answered"]
    result["% Paid (Attempted)"] = [f"{round(p/t*100)}%" if t else "0%"
                                    for p, t in zip(result["Paid"], result["Total"])]
    result["% Paid (Answered)"]  = [f"{round(p/a*100)}%" if a else "0%"
                                    for p, a in zip(result["Paid"], result["Our Answered"])]
    result["_ts"] = pd.to_datetime(result.index, dayfirst=True, errors="coerce", format="mixed")
    result.sort_values("_ts", inplace=True); result.drop(columns=["_ts"], inplace=True)
    tt, tp, ta = result["Total"].sum(), result["Paid"].sum(), result["Our Answered"].sum()
    return pd.concat([result, pd.DataFrame([{
        "Total": tt, "Paid": tp, "Unpaid": result["Unpaid"].sum(),
        "Our Answered": ta, "Not Answered": result["Not Answered"].sum(),
        "AI Qualified": result["AI Qualified"].sum(),
        "% Paid (Attempted)": f"{round(tp/tt*100)}%" if tt else "0%",
        "% Paid (Answered)":  f"{round(tp/ta*100)}%" if ta else "0%",
    }], index=["Grand Total"])])

def plan_revenue_pivot(df):
    if "Revenue_Attributed" in df.columns:
        df_p = df[(df["Revenue_Attributed"] > 0) &
                  df["Plan Name"].notna() &
                  (df["Plan Name"].astype(str).str.strip() != "")].copy()
        rev_col = "Revenue_Attributed"
    else:
        df_p = df[(df["Calling Outcome"] == "Purchased") &
                  df["Plan Name"].notna() &
                  (df["Plan Name"].astype(str).str.strip() != "")].copy()
        rev_col = "Revenue"
    if df_p.empty:
        return pd.DataFrame(columns=["Plan Name", "Revenue (Rs)", "Order Count"])
    df_p["_rev"] = pd.to_numeric(df_p[rev_col], errors="coerce").fillna(0)
    grp = df_p.groupby("Plan Name").agg(
        Revenue=("_rev", "sum"), Order_Count=("Plan Name", "count")).reset_index()
    grp.columns = ["Plan Name", "Revenue (Rs)", "Order Count"]
    grp.sort_values("Revenue (Rs)", ascending=False, inplace=True)
    return pd.concat([grp, pd.DataFrame([{
        "Plan Name": "Grand Total",
        "Revenue (Rs)": grp["Revenue (Rs)"].sum(),
        "Order Count":  grp["Order Count"].sum()
    }])], ignore_index=True)

def plan_revenue_pivot_today(df, today_paid_set):
    if not today_paid_set:
        return pd.DataFrame(columns=["Plan Name", "Revenue (Rs)", "Order Count"])
    if "Revenue_Attributed" in df.columns:
        df_p = df[df["Customer Number"].isin(today_paid_set) &
                  (df["Revenue_Attributed"] > 0) &
                  df["Plan Name"].notna() &
                  (df["Plan Name"].astype(str).str.strip() != "")].copy()
        rev_col = "Revenue_Attributed"
    else:
        df_p = df[df["Customer Number"].isin(today_paid_set) &
                  df["Plan Name"].notna() &
                  (df["Plan Name"].astype(str).str.strip() != "")].copy()
        rev_col = "Revenue"
    if df_p.empty:
        return pd.DataFrame(columns=["Plan Name", "Revenue (Rs)", "Order Count"])
    df_p["_rev"] = pd.to_numeric(df_p[rev_col], errors="coerce").fillna(0)
    grp = df_p.groupby("Plan Name").agg(
        Revenue=("_rev", "sum"), Order_Count=("Plan Name", "count")).reset_index()
    grp.columns = ["Plan Name", "Revenue (Rs)", "Order Count"]
    grp.sort_values("Revenue (Rs)", ascending=False, inplace=True)
    return pd.concat([grp, pd.DataFrame([{
        "Plan Name": "Grand Total",
        "Revenue (Rs)": grp["Revenue (Rs)"].sum(),
        "Order Count":  grp["Order Count"].sum()
    }])], ignore_index=True)

def plan_by_segment(df):
    if "Revenue_Attributed" in df.columns:
        df_p = df[(df["Revenue_Attributed"] > 0) &
                  df["Plan Name"].notna() &
                  (df["Plan Name"].astype(str).str.strip() != "")].copy()
        rev_col = "Revenue_Attributed"
    else:
        df_p = df[(df["Calling Outcome"] == "Purchased") &
                  df["Plan Name"].notna() &
                  (df["Plan Name"].astype(str).str.strip() != "")].copy()
        rev_col = "Revenue"
    if df_p.empty:
        return {}
    df_p["_rev"] = pd.to_numeric(df_p[rev_col], errors="coerce").fillna(0)
    grp = df_p.groupby("Plan Name").agg(revenue=("_rev", "sum"), count=("Plan Name", "count"))
    return {plan: {"revenue": float(row["revenue"]), "count": int(row["count"])}
            for plan, row in grp.iterrows()}

# v22: 5-arg version (adds df_broker)
def build_planwise_rows(df_win, df_ttc_unc, df_ttc, df_renewal_ttc, df_broker):
    p_win = plan_by_segment(df_win)
    p_unc = plan_by_segment(df_ttc_unc)
    p_ttc = plan_by_segment(df_ttc)
    p_ren = plan_by_segment(df_renewal_ttc)
    p_bro = plan_by_segment(df_broker)
    all_plans = sorted(set(list(p_win)+list(p_unc)+list(p_ttc)+list(p_ren)+list(p_bro)))
    rows = []
    for plan in all_plans:
        w = p_win.get(plan, {"revenue": 0, "count": 0})
        u = p_unc.get(plan, {"revenue": 0, "count": 0})
        t = p_ttc.get(plan, {"revenue": 0, "count": 0})
        r = p_ren.get(plan, {"revenue": 0, "count": 0})
        b = p_bro.get(plan, {"revenue": 0, "count": 0})
        rows.append([plan,
                     w["revenue"]+u["revenue"]+t["revenue"]+r["revenue"]+b["revenue"],
                     w["count"]  +u["count"]  +t["count"]  +r["count"]  +b["count"],
                     w["revenue"], w["count"], u["revenue"], u["count"],
                     t["revenue"], t["count"], r["revenue"], r["count"],
                     b["revenue"], b["count"]])
    tr_w=sum(v["revenue"] for v in p_win.values()); tc_w=sum(v["count"] for v in p_win.values())
    tr_u=sum(v["revenue"] for v in p_unc.values()); tc_u=sum(v["count"] for v in p_unc.values())
    tr_t=sum(v["revenue"] for v in p_ttc.values()); tc_t=sum(v["count"] for v in p_ttc.values())
    tr_r=sum(v["revenue"] for v in p_ren.values()); tc_r=sum(v["count"] for v in p_ren.values())
    tr_b=sum(v["revenue"] for v in p_bro.values()); tc_b=sum(v["count"] for v in p_bro.values())
    rows.append(["Grand Total",
                 tr_w+tr_u+tr_t+tr_r+tr_b, tc_w+tc_u+tc_t+tc_r+tc_b,
                 tr_w, tc_w, tr_u, tc_u, tr_t, tc_t, tr_r, tc_r,
                 tr_b, tc_b])
    return rows


# ── Styles ─────────────────────────────────────────────────────────────────────
TITLE_FILL          = PatternFill("solid", fgColor="2E75B6")
HEADER_FILL         = PatternFill("solid", fgColor="BDD7EE")
TOTAL_FILL          = PatternFill("solid", fgColor="9DC3E6")
PAID_FILL           = PatternFill("solid", fgColor="E2EFDA")
AI_QUAL_FILL        = PatternFill("solid", fgColor="FFF2CC")
AI_COL_FILL         = PatternFill("solid", fgColor="FCE4D6")
WIN_FILL            = PatternFill("solid", fgColor="E8F4FD")
TTC_UNC_FILL        = PatternFill("solid", fgColor="FEF9E7")
TTC_FILL            = PatternFill("solid", fgColor="EDE7F6")
RENEWAL_TTC_FILL    = PatternFill("solid", fgColor="E8F8E8")
BROKER_FILL         = PatternFill("solid", fgColor="FDEBD0")   # v22 peach
ANS_COL_FILL        = PatternFill("solid", fgColor="E8F5E9")
OUR_ANS_FILL        = PatternFill("solid", fgColor="C8E6C9")
EMPTY_ROW_FILL      = PatternFill("solid", fgColor="F5F5F5")
AI_QUAL_PAID_FILL   = PatternFill("solid", fgColor="C6EFCE")
AI_UNQUAL_PAID_FILL = PatternFill("solid", fgColor="FFEB9C")
REVENUE_FILL        = PatternFill("solid", fgColor="D6E4F0")

THIN = Side(border_style="thin", color="B0B0B0")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

seg_fills = {
    WINBACK_SOURCE:      WIN_FILL,
    TTC_UNC_SOURCE:      TTC_UNC_FILL,
    TTC_SOURCE:          TTC_FILL,
    RENEWAL_TTC_SOURCE:  RENEWAL_TTC_FILL,
    BROKER_SOURCE:       BROKER_FILL,       # v22
    WINBACK_DISPLAY:     WIN_FILL,
    TTC_UNC_DISPLAY:     TTC_UNC_FILL,
    TTC_DISPLAY:         TTC_FILL,
    RENEWAL_TTC_DISPLAY: RENEWAL_TTC_FILL,
    BROKER_DISPLAY:      BROKER_FILL,       # v22
}

def _cell(ws, r, c, value, bold=False, fill=None, align="center", color="000000"):
    cell = ws.cell(r, c)
    cell.value = value
    cell.font  = Font(bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    cell.border = BOX
    if fill: cell.fill = fill
    return cell


# ── Table writers ──────────────────────────────────────────────────────────────
def write_summary_matrix(ws, title, rows_data, start_row, start_col):
    """Write a summary matrix block with a Grand Total row appended."""
    sr, sc = start_row, start_col
    ws.merge_cells(start_row=sr, start_column=sc,
                   end_row=sr, end_column=sc + len(SUMMARY_METRIC_COLS))
    _cell(ws, sr, sc, title, bold=True, fill=TITLE_FILL, color="FFFFFF")
    ws.row_dimensions[sr].height = 22
    sr += 1

    _header_fills = {
        "AI Qualified":         AI_COL_FILL,
        "Company Answered":     ANS_COL_FILL,
        "Our Answered":         OUR_ANS_FILL,
        "Answered & Paid":      PAID_FILL,
        "AI Qual & Paid":       AI_QUAL_PAID_FILL,
        "AI Unqual & Paid":     AI_UNQUAL_PAID_FILL,
        "AI Qual & Paid Rev":   REVENUE_FILL,
        "AI Unqual & Paid Rev": REVENUE_FILL,
        "Ans & Paid Rev":       REVENUE_FILL,
    }
    for ci, h in enumerate(["Segment"] + SUMMARY_METRIC_COLS):
        _cell(ws, sr, sc + ci, h, bold=True, fill=_header_fills.get(h, HEADER_FILL))
    sr += 1

    # Segment rows
    for row in rows_data:
        seg   = row["Segment"]
        rfill = seg_fills.get(seg)
        _cell(ws, sr, sc, seg, bold=True, fill=rfill, align="left")
        for ci, col in enumerate(SUMMARY_METRIC_COLS, start=1):
            cfill = (AI_QUAL_FILL        if col == "AI Qualified"
                     else OUR_ANS_FILL        if col == "Our Answered"
                     else ANS_COL_FILL        if col == "Company Answered"
                     else AI_QUAL_PAID_FILL   if col == "AI Qual & Paid"
                     else AI_UNQUAL_PAID_FILL if col == "AI Unqual & Paid"
                     else REVENUE_FILL        if col in ("AI Qual & Paid Rev",
                                                         "AI Unqual & Paid Rev",
                                                         "Ans & Paid Rev")
                     else rfill)
            _cell(ws, sr, sc + ci, row.get(col, ""), fill=cfill)
        sr += 1

    # v21: Grand Total row
    gt = build_grand_total_row(rows_data)
    _cell(ws, sr, sc, gt["Segment"], bold=True, fill=TOTAL_FILL, align="left")
    for ci, col in enumerate(SUMMARY_METRIC_COLS, start=1):
        _cell(ws, sr, sc + ci, gt.get(col, ""), bold=True, fill=TOTAL_FILL)
    sr += 1

    return sr

def write_ai_qual_funnel_table(ws, title, df, start_row, start_col):
    sr, sc = start_row, start_col
    ws.merge_cells(start_row=sr, start_column=sc, end_row=sr, end_column=sc + 2)
    _cell(ws, sr, sc, title, bold=True, fill=TITLE_FILL, color="FFFFFF")
    ws.row_dimensions[sr].height = 22; sr += 1
    for ci, h in enumerate(["LSQ Disposition", "Count", "%"]):
        _cell(ws, sr, sc + ci, h, bold=True,
              fill=AI_QUAL_FILL if ci == 0 else HEADER_FILL)
    sr += 1
    for _, row in df.iterrows():
        is_gt    = (row["LSQ Disposition"] == "Grand Total")
        is_empty = (row["LSQ Disposition"] == "(No AI Qualified leads)")
        fill      = TOTAL_FILL if is_gt else (EMPTY_ROW_FILL if is_empty else None)
        txt_color = "888888" if is_empty else "000000"
        _cell(ws, sr, sc,     row["LSQ Disposition"], bold=is_gt, fill=fill,
              align="left", color=txt_color)
        _cell(ws, sr, sc + 1, row["Count"], bold=is_gt, fill=fill)
        _cell(ws, sr, sc + 2, row["%"],     bold=is_gt, fill=fill)
        sr += 1
    return sr

def write_campaign_table(ws, title, df, start_row, start_col):
    sr, sc = start_row, start_col
    ws.merge_cells(start_row=sr, start_column=sc, end_row=sr, end_column=sc + 5)
    _cell(ws, sr, sc, title, bold=True, fill=TITLE_FILL, color="FFFFFF")
    ws.row_dimensions[sr].height = 22; sr += 1
    ws.merge_cells(start_row=sr, start_column=sc, end_row=sr, end_column=sc + 5)
    _cell(ws, sr, sc, "Count of Leads", bold=True, fill=HEADER_FILL, align="left"); sr += 1
    for ci, h in enumerate(["Type Of Lead", "Paid", "Unpaid", "Grand Total", "%", "AI Qualified"]):
        _cell(ws, sr, sc + ci, h, bold=True,
              fill=AI_COL_FILL if h == "AI Qualified" else HEADER_FILL)
    sr += 1
    for label, row in df.iterrows():
        is_gt  = (label == "Grand Total")
        fill   = TOTAL_FILL if is_gt else None
        ai_val = row.get("AI Qualified", 0)
        _cell(ws, sr, sc,     label,                     bold=is_gt, fill=fill, align="left")
        _cell(ws, sr, sc + 1, row.get("Paid", 0),        bold=is_gt, fill=fill)
        _cell(ws, sr, sc + 2, row.get("Unpaid", 0),      bold=is_gt, fill=fill)
        _cell(ws, sr, sc + 3, row.get("Grand Total", 0), bold=is_gt, fill=fill)
        _cell(ws, sr, sc + 4, row.get("%", ""),           bold=is_gt, fill=fill)
        _cell(ws, sr, sc + 5, ai_val, bold=is_gt,
              fill=AI_QUAL_FILL if (not is_gt and ai_val) else fill)
        sr += 1
    return sr

def write_date_table(ws, title, df, start_row, start_col):
    sr, sc = start_row, start_col
    DATA_COL = ["Total", "Paid", "Unpaid", "Our Answered", "Not Answered",
                "% Paid (Attempted)", "% Paid (Answered)", "AI Qualified"]
    ws.merge_cells(start_row=sr, start_column=sc,
                   end_row=sr, end_column=sc + len(DATA_COL))
    _cell(ws, sr, sc, title, bold=True, fill=TITLE_FILL, color="FFFFFF")
    ws.row_dimensions[sr].height = 22; sr += 1
    for ci, h in enumerate(["Date"] + DATA_COL):
        fill = (AI_COL_FILL   if h == "AI Qualified"
                else OUR_ANS_FILL if h == "Our Answered"
                else HEADER_FILL)
        _cell(ws, sr, sc + ci, h, bold=True, fill=fill)
    sr += 1
    for label, row in df.iterrows():
        is_gt = (label == "Grand Total")
        fill  = TOTAL_FILL if is_gt else None
        _cell(ws, sr, sc, label, bold=is_gt, fill=fill, align="left")
        for ci, col in enumerate(DATA_COL, start=1):
            cfill = (fill         if is_gt
                     else AI_QUAL_FILL if col == "AI Qualified" and row.get(col, 0)
                     else OUR_ANS_FILL if col == "Our Answered"
                     else None)
            _cell(ws, sr, sc + ci, row.get(col, 0), bold=is_gt, fill=cfill)
        sr += 1
    return sr

def write_plan_revenue_table(ws, title, df, start_row, start_col):
    sr, sc = start_row, start_col
    ws.merge_cells(start_row=sr, start_column=sc, end_row=sr, end_column=sc + 2)
    _cell(ws, sr, sc, title, bold=True, fill=TITLE_FILL, color="FFFFFF")
    ws.row_dimensions[sr].height = 22; sr += 1
    for ci, h in enumerate(["Plan Name", "Revenue (Rs)", "Order Count"]):
        _cell(ws, sr, sc + ci, h, bold=True, fill=HEADER_FILL)
    sr += 1
    for _, row in df.iterrows():
        is_gt = (row["Plan Name"] == "Grand Total")
        fill  = TOTAL_FILL if is_gt else None
        _cell(ws, sr, sc,     row["Plan Name"],    bold=is_gt, fill=fill, align="left")
        _cell(ws, sr, sc + 1, row["Revenue (Rs)"], bold=is_gt, fill=fill)
        _cell(ws, sr, sc + 2, row["Order Count"],  bold=is_gt, fill=fill)
        sr += 1
    return sr

def write_planwise_block(ws, title, data_rows, start_row, start_col):
    sr, sc = start_row, start_col
    nc = len(PLAN_COLS)
    ws.merge_cells(start_row=sr, start_column=sc, end_row=sr, end_column=sc + nc - 1)
    _cell(ws, sr, sc, title, bold=True, fill=TITLE_FILL, color="FFFFFF")
    ws.row_dimensions[sr].height = 22; sr += 1
    for ci, h in enumerate(PLAN_COLS):
        _cell(ws, sr, sc + ci, h, bold=True, fill=HEADER_FILL)
    sr += 1
    for row_vals in data_rows:
        is_gt = (str(row_vals[0]) == "Grand Total")
        fill  = TOTAL_FILL if is_gt else None
        _cell(ws, sr, sc, row_vals[0], bold=is_gt, fill=fill, align="left")
        for ci in range(1, nc):
            _cell(ws, sr, sc + ci,
                  row_vals[ci] if ci < len(row_vals) else "",
                  bold=is_gt, fill=fill)
        sr += 1
    return sr


# ── v21 enhanced: Rich formatted output to Google Sheet ──────────────────────
def write_summary_to_gsheet(
    gc, summary_today, summary_overall, today_str,
    t_aiq_all, t_aiq_win, t_aiq_ttc_unc, t_aiq_ttc, t_aiq_renewal_ttc,
    t_aiq_broker,   # v22
    t_plan_rev
):
    """
    Writes a richly formatted Lead Summary Report to OUTPUT_SUMMARY_SHEET_ID (gid=0):
      Block 1: Today's Summary  — navy title, blue headers, segment colors, amber GT
      Block 2: Overall Summary  — same formatting
      Block 3: 5 AI Qual Funnels side-by-side (display-name labels, alternating rows)
      Block 4: Plan-wise Revenue (non-zero only, sorted by revenue desc)
    Font: Lexend throughout.
    """
    def _rgb(h):
        h = h.lstrip("#")
        r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        return {"red": r / 255.0, "green": g / 255.0, "blue": b / 255.0}

    NAVY, WHITE, BLUE_H = "1B2A4A", "FFFFFF", "2E75B6"
    AMBER = "F4B942"
    SEG_C = {
        WINBACK_DISPLAY:     "E8F4FD",
        TTC_UNC_DISPLAY:     "FEF9E7",
        TTC_DISPLAY:         "EDE7F6",
        RENEWAL_TTC_DISPLAY: "E8F8E8",
        BROKER_DISPLAY:      "FDEBD0",   # v22
    }
    F_TITLE = "1565C0"; F_COL = "1976D2"; F_ALT = "E3F2FD"; F_GT = "BBDEFB"
    P_TITLE = "1B2A4A"; P_COL  = "37474F"; P_ALT = "F1F8E9"; P_GT = "C8E6C9"
    FONT = "Lexend"

    DISP_COLS = ["Segment"] + SUMMARY_METRIC_COLS
    N_SUM     = len(DISP_COLS)   # 14 cols A-N
    TOTAL_W   = 24               # v22: A-X for 6 funnels

    gt_today   = build_grand_total_row(summary_today)
    gt_overall = build_grand_total_row(summary_overall)

    def rv(d): return [d.get(c, "") for c in DISP_COLS]
    def pad(row, w=TOTAL_W): return (list(row) + [""] * w)[:w]

    reqs = []

    def fmt(r1, r2, c1, c2, bg=None, fg="000000", bold=False,
            size=10, align="CENTER", valign="MIDDLE"):
        cell_fmt = {
            "textFormat": {
                "fontFamily": FONT, "fontSize": size, "bold": bold,
                "foregroundColor": _rgb(fg),
            },
            "horizontalAlignment": align,
            "verticalAlignment": valign,
        }
        if bg: cell_fmt["backgroundColor"] = _rgb(bg)
        reqs.append({
            "repeatCell": {
                "range": {"sheetId": 0,
                          "startRowIndex": r1, "endRowIndex": r2,
                          "startColumnIndex": c1, "endColumnIndex": c2},
                "cell": {"userEnteredFormat": cell_fmt},
                "fields": ("userEnteredFormat(backgroundColor,textFormat,"
                           "horizontalAlignment,verticalAlignment)"),
            }
        })

    def merge(r1, r2, c1, c2):
        reqs.append({
            "mergeCells": {
                "range": {"sheetId": 0,
                          "startRowIndex": r1, "endRowIndex": r2,
                          "startColumnIndex": c1, "endColumnIndex": c2},
                "mergeType": "MERGE_ALL",
            }
        })

    def row_h(ri, h):
        reqs.append({
            "updateDimensionProperties": {
                "range": {"sheetId": 0, "dimension": "ROWS",
                          "startIndex": ri, "endIndex": ri + 1},
                "properties": {"pixelSize": h}, "fields": "pixelSize",
            }
        })

    def col_w(ci, w):
        reqs.append({
            "updateDimensionProperties": {
                "range": {"sheetId": 0, "dimension": "COLUMNS",
                          "startIndex": ci, "endIndex": ci + 1},
                "properties": {"pixelSize": w}, "fields": "pixelSize",
            }
        })

    grid = []
    _r   = [0]

    def add_row(values):
        grid.append(pad(values))
        r = _r[0]; _r[0] += 1
        return r

    # ═══ BLOCK 1: Today's Summary ═══════════════════════════════════════════
    r = add_row([f"Today's Summary ({today_str})"])
    fmt(r, r+1, 0, N_SUM, bg=NAVY, fg=WHITE, bold=True, size=11, align="LEFT")
    merge(r, r+1, 0, N_SUM); row_h(r, 28)

    r = add_row(DISP_COLS)
    fmt(r, r+1, 0, N_SUM, bg=BLUE_H, fg=WHITE, bold=True)
    fmt(r, r+1, 0, 1, bg=BLUE_H, fg=WHITE, bold=True, align="LEFT")
    row_h(r, 24)

    for row_d in summary_today:
        seg = row_d["Segment"]; bg = SEG_C.get(seg, "FFFFFF")
        r = add_row(rv(row_d))
        fmt(r, r+1, 0, 1, bg=bg, align="LEFT")
        fmt(r, r+1, 1, N_SUM, bg=bg)
        row_h(r, 22)

    r = add_row(rv(gt_today))
    fmt(r, r+1, 0, 1, bg=AMBER, bold=True, align="LEFT")
    fmt(r, r+1, 1, N_SUM, bg=AMBER, bold=True)
    row_h(r, 24)

    r = add_row([]); row_h(r, 10)

    # ═══ BLOCK 2: Overall Summary ════════════════════════════════════════════
    r = add_row(["Overall Summary (All Dates)"])
    fmt(r, r+1, 0, N_SUM, bg=NAVY, fg=WHITE, bold=True, size=11, align="LEFT")
    merge(r, r+1, 0, N_SUM); row_h(r, 28)

    r = add_row(DISP_COLS)
    fmt(r, r+1, 0, N_SUM, bg=BLUE_H, fg=WHITE, bold=True)
    fmt(r, r+1, 0, 1, bg=BLUE_H, fg=WHITE, bold=True, align="LEFT")
    row_h(r, 24)

    for row_d in summary_overall:
        seg = row_d["Segment"]; bg = SEG_C.get(seg, "FFFFFF")
        r = add_row(rv(row_d))
        fmt(r, r+1, 0, 1, bg=bg, align="LEFT")
        fmt(r, r+1, 1, N_SUM, bg=bg)
        row_h(r, 22)

    r = add_row(rv(gt_overall))
    fmt(r, r+1, 0, 1, bg=AMBER, bold=True, align="LEFT")
    fmt(r, r+1, 1, N_SUM, bg=AMBER, bold=True)
    row_h(r, 24)

    r = add_row([]); row_h(r, 10)

    # ═══ BLOCK 3: AI Qual Funnels (5 side-by-side) ═══════════════════════════
    funnel_tables = [
        ("Overall AI Qual Funnel",             t_aiq_all),
        ("Winback AI Qual Funnel",              t_aiq_win),
        (f"{TTC_UNC_DISPLAY} AI Qual Funnel",  t_aiq_ttc_unc),
        (f"{TTC_DISPLAY} AI Qual Funnel",       t_aiq_ttc),
        ("Renewal_TTC AI Qual Funnel",          t_aiq_renewal_ttc),
        ("Broker AI Qual Funnel",               t_aiq_broker),   # v22
    ]
    F_OFF = [0, 4, 8, 12, 16, 20]   # v22: +20

    max_df_rows  = max(len(df) for _, df in funnel_tables)
    total_f_rows = 2 + max_df_rows   # title + col-hdr + data

    f_start = _r[0]
    for _ in range(total_f_rows):
        add_row([])

    for tbl_i, (tbl_title, tbl_df) in enumerate(funnel_tables):
        off = F_OFF[tbl_i]

        r = f_start
        grid[r][off] = tbl_title
        fmt(r, r+1, off, off+3, bg=F_TITLE, fg=WHITE, bold=True, size=10, align="LEFT")
        merge(r, r+1, off, off+3)
        if tbl_i == 0: row_h(r, 26)

        r = f_start + 1
        grid[r][off] = "LSQ Disposition"
        grid[r][off+1] = "Count"
        grid[r][off+2] = "%"
        fmt(r, r+1, off, off+3, bg=F_COL, fg=WHITE, bold=True)
        fmt(r, r+1, off, off+1, bg=F_COL, fg=WHITE, bold=True, align="LEFT")
        if tbl_i == 0: row_h(r, 22)

        for di, (_, drow) in enumerate(tbl_df.iterrows()):
            r = f_start + 2 + di
            disp  = str(drow.get("LSQ Disposition", ""))
            is_gt = (disp == "Grand Total")
            grid[r][off]   = disp
            grid[r][off+1] = int(drow["Count"]) if not is_gt and pd.notna(drow.get("Count")) else drow.get("Count", "")
            grid[r][off+2] = drow.get("%", "")
            bg   = F_GT if is_gt else (F_ALT if di % 2 else "FFFFFF")
            bold = is_gt
            fmt(r, r+1, off, off+3, bg=bg, bold=bold)
            fmt(r, r+1, off, off+1, bg=bg, bold=bold, align="LEFT")
            if tbl_i == 0: row_h(r, 20)

    r = add_row([]); row_h(r, 10)

    # ═══ BLOCK 4: Plan-wise Revenue ══════════════════════════════════════════
    r = add_row(["Plan-wise Revenue", "", ""])
    fmt(r, r+1, 0, 3, bg=P_TITLE, fg=WHITE, bold=True, size=11, align="LEFT")
    merge(r, r+1, 0, 3); row_h(r, 28)

    r = add_row(["Plan Name", "Revenue (Rs)", "Order Count"])
    fmt(r, r+1, 0, 3, bg=P_COL, fg=WHITE, bold=True)
    fmt(r, r+1, 0, 1, bg=P_COL, fg=WHITE, bold=True, align="LEFT")
    row_h(r, 22)

    for di, (_, pr) in enumerate(t_plan_rev.iterrows()):
        plan_n = str(pr.get("Plan Name", ""))
        is_gt  = (plan_n == "Grand Total")
        try:    rev_v = int(pr.get("Revenue (Rs)", 0))
        except: rev_v = pr.get("Revenue (Rs)", 0)
        try:    ord_v = int(pr.get("Order Count", 0))
        except: ord_v = pr.get("Order Count", 0)
        r = add_row([plan_n, rev_v, ord_v])
        bg = P_GT if is_gt else (P_ALT if di % 2 else "FFFFFF")
        fmt(r, r+1, 0, 3, bg=bg, bold=is_gt)
        fmt(r, r+1, 0, 1, bg=bg, bold=is_gt, align="LEFT")
        row_h(r, 20)

    # ── Column widths ─────────────────────────────────────────────────────────
    col_w(0, 155)
    for c in range(1, N_SUM): col_w(c, 88)
    for off in F_OFF:
        col_w(off + 1, 70); col_w(off + 2, 55)
    for off in [3, 7, 11, 15, 19]: col_w(off, 12)   # v22: +19

    # ── Write to Google Sheet ─────────────────────────────────────────────────
    sh = gc.open_by_key(OUTPUT_SUMMARY_SHEET_ID)
    ws = sh.get_worksheet(0)
    ws.clear()

    try:
        ws.update(range_name="A1", values=grid)
    except TypeError:
        ws.update("A1", grid)

    all_reqs = [{
        "unmergeCells": {
            "range": {"sheetId": 0,
                      "startRowIndex": 0, "endRowIndex": len(grid) + 5,
                      "startColumnIndex": 0, "endColumnIndex": TOTAL_W}
        }
    }] + reqs

    for i in range(0, len(all_reqs), 100):
        sh.batch_update({"requests": all_reqs[i:i + 100]})

    print(f"OK Lead Summary -> Google Sheet {OUTPUT_SUMMARY_SHEET_ID} "
          f"({len(grid)} rows, {len(all_reqs)} format ops applied)")


# ── Build pivot tables ─────────────────────────────────────────────────────────
t1              = campaign_pivot(out_all)
t_win_c         = campaign_pivot(out_win)
t_ttc_unc_c     = campaign_pivot(out_ttc_unc)
t_ttc_c         = campaign_pivot(out_ttc)
t_renewal_ttc_c = campaign_pivot(out_renewal_ttc)
t_today_c       = campaign_pivot(out_all[out_all["_date_ts"] == today_ts])

t_aiq_all         = ai_qualified_funnel_pivot(out_all)
t_aiq_win         = ai_qualified_funnel_pivot(out_win)
t_aiq_ttc_unc     = ai_qualified_funnel_pivot(out_ttc_unc)
t_aiq_ttc         = ai_qualified_funnel_pivot(out_ttc)
t_aiq_renewal_ttc = ai_qualified_funnel_pivot(out_renewal_ttc)
t_broker_c        = campaign_pivot(out_broker)            # v22
t_aiq_broker      = ai_qualified_funnel_pivot(out_broker) # v22
t_aiq_today       = ai_qualified_funnel_pivot(out_all[out_all["_date_ts"] == today_ts])

t_plan_rev       = plan_revenue_pivot(out_all)
t_plan_rev_today = plan_revenue_pivot_today(out_all, paid_set_today)
t_wtd            = date_pivot(out_all)


# ── Build Summary XLSX workbook ────────────────────────────────────────────────
wb   = openpyxl.Workbook()
wsum = wb.active
wsum.title = "Summary"
wsum.sheet_view.showGridLines = False

nr = write_summary_matrix(wsum, f"Today's Summary ({today_str})", summary_today,   1, 1)
nr = write_summary_matrix(wsum, "Overall Summary (All Dates)",    summary_overall, nr + 2, 1)
gap0 = nr + 2

nr_1 = write_ai_qual_funnel_table(wsum, "Overall AI Qual Funnel",     t_aiq_all,         gap0, 1)
nr_2 = write_ai_qual_funnel_table(wsum, "Winback AI Qual Funnel",      t_aiq_win,         gap0, 5)
nr_3 = write_ai_qual_funnel_table(wsum, "TTC+Uncalled AI Qual Funnel", t_aiq_ttc_unc,     gap0, 9)
nr_4 = write_ai_qual_funnel_table(wsum, "TTC AI Qual Funnel",          t_aiq_ttc,         gap0, 13)
nr_5 = write_ai_qual_funnel_table(wsum, "Renewal_TTC AI Qual Funnel",  t_aiq_renewal_ttc, gap0, 17)
nr_6 = write_ai_qual_funnel_table(wsum, "Broker AI Qual Funnel",       t_aiq_broker,      gap0, 21)  # v22
gap1 = max(nr_1, nr_2, nr_3, nr_4, nr_5, nr_6) + 2

nr_l = write_campaign_table(wsum,     "Overall Leads",     t1,         gap1, 1)
nr_r = write_plan_revenue_table(wsum, "Plan-wise Revenue", t_plan_rev, gap1, 9)
gap2 = max(nr_l, nr_r) + 2

nr_l = write_campaign_table(wsum, "Winback Leads",      t_win_c,         gap2, 1)
nr_m = write_campaign_table(wsum, "TTC+Uncalled Leads", t_ttc_unc_c,     gap2, 9)
nr_r = write_campaign_table(wsum, "TTC Leads",          t_ttc_c,         gap2, 17)
nr_x = write_campaign_table(wsum, "Renewal_TTC Leads",  t_renewal_ttc_c, gap2, 25)
nr_y = write_campaign_table(wsum, "Broker Leads",       t_broker_c,      gap2, 33)  # v22
gap3 = max(nr_l, nr_m, nr_r, nr_x, nr_y) + 2

nr_l = write_date_table(wsum,           "Week to Date",                 t_wtd,       gap3, 1)
nr_r = write_ai_qual_funnel_table(wsum, f"Today AI Qual ({today_str})", t_aiq_today, gap3, 13)
gap4 = max(nr_l, nr_r) + 2

write_campaign_table(wsum,     f"Today's Leads ({today_str})",        t_today_c,        gap4, 1)
write_plan_revenue_table(wsum, f"Today's Plan Revenue ({today_str})", t_plan_rev_today, gap4, 9)

for col, w in {"A": 22, "B": 12, "C": 14, "D": 14, "E": 12, "F": 14, "G": 14, "H": 12,
               "I": 14, "J": 14, "K": 12, "L": 12, "M": 14, "N": 14, "O": 12, "P": 14,
               "Q": 14, "R": 14, "S": 14, "T": 14, "U": 14, "V": 14, "W": 14,
               "X": 14, "Y": 14, "Z": 14, "AA": 14, "AB": 14, "AC": 14, "AD": 14,
               "AE": 14, "AF": 14, "AG": 14, "AH": 14}.items():  # v22
    wsum.column_dimensions[col].width = w


# ══════════════════════════════════════════════════════════════════════════════
# STEP 7b — Detail tabs (cascading funnel)
# ══════════════════════════════════════════════════════════════════════════════
DETAIL_COLS = [
    "Date", "Customer Number", "Lead Name", "Source", "Type Of Lead",
    "Lead Stage", "Disconnect Reason", "Business Disposition",
    "CallerAI Qualified", "Business Disposition Confidence",
    "Calling Outcome", "Paid Check", "Lead Owner", "Activity Added By",
    "Plan Name", "Revenue"
]
DETAIL_WIDTHS = [16, 16, 26, 16, 18, 18, 22, 20, 18, 24, 22, 12, 28, 28, 22, 12]

detail_df = out[[c for c in DETAIL_COLS if c in out.columns]].copy()
for col in ["Calling Outcome", "CallerAI Qualified", "Business Disposition",
            "Business Disposition Confidence",
            "Lead Owner", "Activity Added By", "Plan Name", "Disconnect Reason"]:
    if col in detail_df.columns:
        detail_df[col] = detail_df[col].fillna("")
detail_df["Revenue"] = detail_df["Revenue"].fillna(0)

detail_df["_tab"] = detail_df.apply(assign_tab, axis=1)
cols_in_detail = [c for c in DETAIL_COLS if c in detail_df.columns]
widths_map     = {c: w for c, w in zip(DETAIL_COLS, DETAIL_WIDTHS)}

_disp_slots  = sorted([k for k in OUTCOME_ORDER if k != "Purchased"],
                      key=lambda k: OUTCOME_ORDER[k])
_extra_slots = sorted(
    [t for t in detail_df["_tab"].unique()
     if t not in (DETAIL_TAB_PURCHASED, DETAIL_TAB_AI_QUAL, DETAIL_TAB_UNANSWERED)
     and t not in OUTCOME_ORDER],
    key=lambda x: OUTCOME_ORDER.get(x, 998))
PRIORITY_TAB_ORDER = ([DETAIL_TAB_PURCHASED, DETAIL_TAB_AI_QUAL]
                      + _disp_slots + _extra_slots + [DETAIL_TAB_UNANSWERED])

for tab_name in PRIORITY_TAB_ORDER:
    subset = detail_df[detail_df["_tab"] == tab_name].copy()
    if subset.empty: continue

    if tab_name == DETAIL_TAB_AI_QUAL:
        subset["_d"] = subset["Calling Outcome"].map(OUTCOME_ORDER).fillna(999)
        subset.sort_values(["_d", "Business Disposition Confidence", "Date"], inplace=True)
        subset.drop(columns=["_d"], inplace=True)
    else:
        subset["_ps"] = subset["Paid Check"].map({"Paid": 0, "Unpaid": 1}).fillna(1)
        subset.sort_values(["_ps", "Date"], inplace=True)
        subset.drop(columns=["_ps"], inplace=True)
    subset.reset_index(drop=True, inplace=True)

    paid_in   = int((subset["Paid Check"] == "Paid").sum())
    unpaid_in = int((subset["Paid Check"] == "Unpaid").sum())
    ai_q_in   = int((subset["CallerAI Qualified"] == "Yes").sum())
    sheet_name = re.sub(r'[:\\/?*\[\]]', '', str(tab_name))[:31]

    ws_d = wb.create_sheet(title=sheet_name)
    ws_d.sheet_view.showGridLines = False

    ws_d.merge_cells(start_row=1, start_column=1, end_row=1,
                     end_column=len(cols_in_detail))
    tc = ws_d.cell(1, 1)
    tc.value = (f"{tab_name}  --  Total: {len(subset)}  |  Paid: {paid_in}  |  "
                f"Unpaid: {unpaid_in}  |  AI Qualified: {ai_q_in}  |  {src_count_str(subset)}")
    tc.font = Font(bold=True, color="FFFFFF"); tc.fill = TITLE_FILL
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws_d.row_dimensions[1].height = 22

    header_row = 2
    if tab_name == DETAIL_TAB_AI_QUAL and not subset.empty:
        _bk = (subset.groupby(["Calling Outcome", "Business Disposition Confidence"],
                               dropna=False)
               .size().reset_index(name="Count"))
        bk_str = "  |  ".join(
            f"{r['Calling Outcome']} / {r['Business Disposition Confidence']}: {r['Count']}"
            for _, r in _bk.iterrows())
        if bk_str:
            ws_d.merge_cells(start_row=2, start_column=1, end_row=2,
                             end_column=len(cols_in_detail))
            bc = ws_d.cell(2, 1)
            bc.value = bk_str; bc.font = Font(italic=True, color="444444")
            bc.fill = AI_QUAL_FILL
            bc.alignment = Alignment(horizontal="left", vertical="center")
            bc.border = BOX; ws_d.row_dimensions[2].height = 18
            header_row = 3

    for ci, h in enumerate(cols_in_detail):
        hfill = (AI_COL_FILL
                 if h in ("CallerAI Qualified",
                           "Business Disposition Confidence",
                           "Business Disposition")
                 else HEADER_FILL)
        _cell(ws_d, header_row, ci + 1, h, bold=True, fill=hfill)

    for ri, (_, row) in enumerate(subset[cols_in_detail].iterrows(),
                                  start=header_row + 1):
        is_paid = (row.get("Paid Check", "") == "Paid")
        is_ai_q = (row.get("CallerAI Qualified", "No") == "Yes")
        row_fill = (PAID_FILL   if is_paid
                    else AI_QUAL_FILL if is_ai_q
                    else seg_fills.get(row.get("Source", ""), None))
        for ci, col in enumerate(cols_in_detail):
            cell = ws_d.cell(ri, ci + 1)
            cell.value = row[col]; cell.font = Font(bold=is_paid)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = BOX
            if row_fill: cell.fill = row_fill

    for ci, col in enumerate(cols_in_detail, start=1):
        ws_d.column_dimensions[
            openpyxl.utils.get_column_letter(ci)].width = widths_map.get(col, 16)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 7c — AI Qualified tabs (per campaign + Overall) + AI Answered
# ══════════════════════════════════════════════════════════════════════════════
AI_QUAL_TAB_FILL = PatternFill("solid", fgColor="5B9BD5")

def write_ai_qual_segment_tab(wb, sheet_name, df_segment, title):
    df_aiq = df_segment[df_segment["CallerAI Qualified"] == "Yes"].copy()
    for col in ["Calling Outcome", "Business Disposition",
                "Business Disposition Confidence",
                "Lead Owner", "Activity Added By", "Plan Name",
                "Disconnect Reason"]:
        if col in df_aiq.columns: df_aiq[col] = df_aiq[col].fillna("")
    if "Revenue" in df_aiq.columns:
        df_aiq["Revenue"] = pd.to_numeric(df_aiq["Revenue"], errors="coerce").fillna(0)

    df_aiq["_sort"] = df_aiq["Calling Outcome"].map(OUTCOME_ORDER).fillna(999)
    df_aiq.sort_values(["_sort", "Business Disposition Confidence", "Date"],
                       inplace=True)
    df_aiq.drop(columns=["_sort"], inplace=True)
    df_aiq.reset_index(drop=True, inplace=True)

    disp_counts = {}
    if "Calling Outcome" in df_aiq.columns:
        disp_counts = (df_aiq["Calling Outcome"]
                       .replace("", "No LSQ Dispo")
                       .value_counts().to_dict())

    cols_use = [c for c in DETAIL_COLS if c in df_aiq.columns]
    w_map    = {c: w for c, w in zip(DETAIL_COLS, DETAIL_WIDTHS)}

    ws = wb.create_sheet(title=re.sub(r'[:\\/?*\[\]]', '', sheet_name)[:31])
    ws.sheet_view.showGridLines = False

    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=max(len(cols_use), 1))
    tc = ws.cell(1, 1)
    tc.value = (f"{title}  --  Total AI Qual: {len(df_aiq)}  |  "
                f"{src_count_str(df_aiq)}")
    tc.font = Font(bold=True, color="FFFFFF"); tc.fill = AI_QUAL_TAB_FILL
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    header_row = 2
    if disp_counts:
        bk_str = "  |  ".join(
            f"{k}: {v}" for k, v in
            sorted(disp_counts.items(),
                   key=lambda x: OUTCOME_ORDER.get(x[0], 999)))
        ws.merge_cells(start_row=2, start_column=1, end_row=2,
                       end_column=max(len(cols_use), 1))
        bc = ws.cell(2, 1)
        bc.value = bk_str; bc.font = Font(italic=True, color="1A5276")
        bc.fill = AI_QUAL_FILL
        bc.alignment = Alignment(horizontal="left", vertical="center")
        bc.border = BOX; ws.row_dimensions[2].height = 18
        header_row = 3

    for ci, h in enumerate(cols_use):
        hfill = (AI_COL_FILL
                 if h in ("CallerAI Qualified",
                           "Business Disposition Confidence",
                           "Business Disposition")
                 else HEADER_FILL)
        _cell(ws, header_row, ci + 1, h, bold=True, fill=hfill)

    for ri, (_, row) in enumerate(df_aiq[cols_use].iterrows(),
                                  start=header_row + 1):
        row_fill = seg_fills.get(row.get("Source", ""), AI_QUAL_FILL)
        for ci, col in enumerate(cols_use):
            cell = ws.cell(ri, ci + 1)
            cell.value = row[col]; cell.font = Font()
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = BOX; cell.fill = row_fill

    for ci, col in enumerate(cols_use, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(ci)].width = w_map.get(col, 16)
    return ws

def write_ai_answered_tab(wb, df, sheet_name="AI Answered"):
    _mask = df["Disconnect Reason"].fillna("").str.strip().str.lower().isin(
        {r.lower() for r in CALLERAI_ANSWERED_REASONS})
    df_ans = df[_mask].copy()
    for col in ["Calling Outcome", "CallerAI Qualified", "Business Disposition",
                "Business Disposition Confidence",
                "Lead Owner", "Activity Added By", "Plan Name",
                "Disconnect Reason"]:
        if col in df_ans.columns: df_ans[col] = df_ans[col].fillna("")
    if "Revenue" in df_ans.columns:
        df_ans["Revenue"] = pd.to_numeric(df_ans["Revenue"], errors="coerce").fillna(0)

    df_ans["_ps"] = df_ans["Paid Check"].map({"Paid": 0, "Unpaid": 1}).fillna(1)
    df_ans.sort_values(["_ps", "Date"], inplace=True)
    df_ans.drop(columns=["_ps"], inplace=True)
    df_ans.reset_index(drop=True, inplace=True)

    paid_c   = int((df_ans["Paid Check"] == "Paid").sum())
    ai_q_c   = int((df_ans["CallerAI Qualified"] == "Yes").sum())
    cols_use = [c for c in DETAIL_COLS if c in df_ans.columns]
    w_map    = {c: w for c, w in zip(DETAIL_COLS, DETAIL_WIDTHS)}

    ws = wb.create_sheet(title=sheet_name[:31])
    ws.sheet_view.showGridLines = False

    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=max(len(cols_use), 1))
    tc = ws.cell(1, 1)
    tc.value = (f"AI Answered (user_request | CLIENT_INITIATED | "
                f"inactivity_timeout | transferred)"
                f"  --  Total: {len(df_ans)}  |  Paid: {paid_c}  |  "
                f"AI Qual: {ai_q_c}  |  {src_count_str(df_ans)}")
    tc.font = Font(bold=True, color="FFFFFF")
    tc.fill = PatternFill("solid", fgColor="1A7846")
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    for ci, h in enumerate(cols_use):
        hfill = (AI_COL_FILL
                 if h in ("CallerAI Qualified",
                           "Business Disposition Confidence",
                           "Business Disposition")
                 else HEADER_FILL)
        _cell(ws, 2, ci + 1, h, bold=True, fill=hfill)

    for ri, (_, row) in enumerate(df_ans[cols_use].iterrows(), start=3):
        is_paid  = (str(row.get("Paid Check", "")) == "Paid")
        is_ai_q  = (str(row.get("CallerAI Qualified", "")) == "Yes")
        row_fill = (PAID_FILL   if is_paid
                    else AI_QUAL_FILL if is_ai_q
                    else seg_fills.get(row.get("Source", ""), None))
        for ci, col in enumerate(cols_use):
            cell = ws.cell(ri, ci + 1)
            cell.value = row[col]; cell.font = Font(bold=is_paid)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = BOX
            if row_fill: cell.fill = row_fill

    for ci, col in enumerate(cols_use, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(ci)].width = w_map.get(col, 16)
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# STEP 7d — Per-campaign Purchased tabs + No LSQ Dispo tab
# ══════════════════════════════════════════════════════════════════════════════
def write_campaign_purchased_tab(wb, df_campaign, sheet_name, title):
    df_p = df_campaign[df_campaign["Paid Check"] == "Paid"].copy()
    for col in ["Calling Outcome", "Business Disposition",
                "Business Disposition Confidence",
                "Lead Owner", "Activity Added By", "Plan Name",
                "Disconnect Reason"]:
        if col in df_p.columns: df_p[col] = df_p[col].fillna("")
    if "Revenue" in df_p.columns:
        df_p["Revenue"] = pd.to_numeric(df_p["Revenue"], errors="coerce").fillna(0)
    if "Revenue_Attributed" in df_p.columns:
        df_p["Revenue_Attributed"] = (
            pd.to_numeric(df_p["Revenue_Attributed"], errors="coerce").fillna(0))
    df_p.sort_values("Date", inplace=True, kind="stable")
    df_p.reset_index(drop=True, inplace=True)

    total_rev = (float(df_p["Revenue_Attributed"].sum())
                 if "Revenue_Attributed" in df_p.columns
                 else float(df_p["Revenue"].sum())
                 if "Revenue" in df_p.columns else 0.0)

    cols_use  = [c for c in DETAIL_COLS if c in df_p.columns]
    w_map     = {c: w for c, w in zip(DETAIL_COLS, DETAIL_WIDTHS)}
    PURCH_HDR = PatternFill("solid", fgColor="1F7743")

    ws = wb.create_sheet(title=re.sub(r'[:\\/?*\[\]]', '', sheet_name)[:31])
    ws.sheet_view.showGridLines = False

    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=max(len(cols_use), 1))
    tc = ws.cell(1, 1)
    tc.value = (f"{title}  --  Total: {len(df_p)}  |  "
                f"Attributed Revenue: Rs {total_rev:,.0f}  |  {src_count_str(df_p)}")
    tc.font = Font(bold=True, color="FFFFFF"); tc.fill = PURCH_HDR
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    header_row = 2
    if "Plan Name" in df_p.columns:
        _plans  = df_p["Plan Name"].replace("", "(No Plan)").value_counts()
        bk_str  = "  |  ".join(f"{k}: {v}" for k, v in _plans.head(8).items())
        ws.merge_cells(start_row=2, start_column=1, end_row=2,
                       end_column=max(len(cols_use), 1))
        bc = ws.cell(2, 1)
        bc.value = bk_str; bc.font = Font(italic=True, color="1A5226")
        bc.fill = PAID_FILL
        bc.alignment = Alignment(horizontal="left", vertical="center")
        bc.border = BOX; ws.row_dimensions[2].height = 18
        header_row = 3

    for ci, h in enumerate(cols_use):
        _cell(ws, header_row, ci + 1, h, bold=True, fill=HEADER_FILL)

    for ri, (_, row) in enumerate(df_p[cols_use].iterrows(),
                                  start=header_row + 1):
        for ci, col in enumerate(cols_use):
            cell = ws.cell(ri, ci + 1)
            cell.value = row[col]; cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = BOX; cell.fill = PAID_FILL

    for ci, col in enumerate(cols_use, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(ci)].width = w_map.get(col, 16)
    return ws


def write_no_lsq_dispo_tab(wb, df, sheet_name="No LSQ Dispo Dump"):
    df_no = df[
        (df["Paid Check"] != "Paid") &
        (df["Calling Outcome"].isna() |
         (df["Calling Outcome"].astype(str).str.strip() == ""))
    ].copy()
    for col in ["Calling Outcome", "CallerAI Qualified", "Business Disposition",
                "Business Disposition Confidence", "Lead Owner",
                "Activity Added By", "Plan Name", "Disconnect Reason"]:
        if col in df_no.columns: df_no[col] = df_no[col].fillna("")
    if "Revenue" in df_no.columns:
        df_no["Revenue"] = pd.to_numeric(df_no["Revenue"], errors="coerce").fillna(0)

    df_no["_qs"]   = df_no["CallerAI Qualified"].map({"Yes": 0, "No": 1}).fillna(1)
    df_no["_srcs"] = df_no["Source"].map({
        WINBACK_SOURCE: 0, TTC_UNC_SOURCE: 1,
        TTC_SOURCE: 2,     RENEWAL_TTC_SOURCE: 3, BROKER_SOURCE: 4   # v22
    }).fillna(5)
    df_no.sort_values(["_qs", "_srcs", "Date"], inplace=True)
    df_no.drop(columns=["_qs", "_srcs"], inplace=True)
    df_no.reset_index(drop=True, inplace=True)

    total  = len(df_no)
    ai_q_c = (int((df_no["CallerAI Qualified"] == "Yes").sum())
               if "CallerAI Qualified" in df_no.columns else 0)
    cols_use    = [c for c in DETAIL_COLS if c in df_no.columns]
    w_map       = {c: w for c, w in zip(DETAIL_COLS, DETAIL_WIDTHS)}
    NO_DISP_HDR = PatternFill("solid", fgColor="7B3F00")

    ws = wb.create_sheet(title=sheet_name[:31])
    ws.sheet_view.showGridLines = False

    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=max(len(cols_use), 1))
    tc = ws.cell(1, 1)
    tc.value = (f"No LSQ Disposition Dump  --  Total: {total}  |  "
                f"AI Qualified (no follow-up): {ai_q_c}  |  {src_count_str(df_no)}")
    tc.font = Font(bold=True, color="FFFFFF"); tc.fill = NO_DISP_HDR
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    for ci, h in enumerate(cols_use):
        hfill = (AI_COL_FILL
                 if h in ("CallerAI Qualified",
                           "Business Disposition Confidence",
                           "Business Disposition")
                 else HEADER_FILL)
        _cell(ws, 2, ci + 1, h, bold=True, fill=hfill)

    for ri, (_, row) in enumerate(df_no[cols_use].iterrows(), start=3):
        is_ai_q  = (str(row.get("CallerAI Qualified", "")) == "Yes")
        row_fill = (AI_QUAL_FILL if is_ai_q
                    else seg_fills.get(row.get("Source", ""), None))
        for ci, col in enumerate(cols_use):
            cell = ws.cell(ri, ci + 1)
            cell.value = row[col]; cell.font = Font()
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = BOX
            if row_fill: cell.fill = row_fill

    for ci, col in enumerate(cols_use, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(ci)].width = w_map.get(col, 16)
    return ws


write_ai_qual_segment_tab(wb, "Winback AI Qual",     out_win,
                          "Winback -- AI Qualified Leads")
write_ai_qual_segment_tab(wb, "TTC+Unc AI Qual",     out_ttc_unc,
                          "TTC+Uncalled -- AI Qualified Leads")
write_ai_qual_segment_tab(wb, "TTC AI Qual",         out_ttc,
                          "TTC -- AI Qualified Leads")
write_ai_qual_segment_tab(wb, "Renewal TTC AI Qual", out_renewal_ttc,
                          "Renewal_TTC -- AI Qualified Leads")
write_ai_qual_segment_tab(wb, "Broker AI Qual",      out_broker,        # v22
                          "Broker -- AI Qualified Leads")
write_ai_qual_segment_tab(wb, "Overall AI Qual",     out_all,
                          "Overall -- AI Qualified Leads (All Campaigns)")
write_ai_answered_tab(wb, out_all, "AI Answered")

write_campaign_purchased_tab(wb, out_ttc_unc,     "TTC+Unc Purchased",
                             "TTC+Uncalled -- Purchased Leads")
write_campaign_purchased_tab(wb, out_ttc,         "TTC Purchased",
                             "TTC -- Purchased Leads")
write_campaign_purchased_tab(wb, out_win,         "Winback Purchased",
                             "Winback -- Purchased Leads")
write_campaign_purchased_tab(wb, out_renewal_ttc, "Renewal TTC Purchased",
                             "Renewal_TTC -- Purchased Leads")
write_campaign_purchased_tab(wb, out_broker,      "Broker Purchased",      # v22
                             "Broker -- Purchased Leads")

write_no_lsq_dispo_tab(wb, out_all, "No LSQ Dispo Dump")

tab_counts = detail_df["_tab"].value_counts()
print("\n-- Detail tab assignment ------------------------------------------------")
for t in PRIORITY_TAB_ORDER:
    if t in tab_counts.index:
        print(f"   {t:<30} {tab_counts[t]:>5}")
print(f"   {'TOTAL':<30} {tab_counts.sum():>5}")

print(f"\n-- AI Qualified tabs (incl. Paid) ---------------------------------------")
for seg_nm, seg_df in [
    (WINBACK_SOURCE,      out_win),
    (TTC_UNC_SOURCE,      out_ttc_unc),
    (TTC_SOURCE,          out_ttc),
    (RENEWAL_TTC_SOURCE,  out_renewal_ttc),
    (BROKER_SOURCE,       out_broker),       # v22
    ("Overall",           out_all),
]:
    n_total = int((seg_df["CallerAI Qualified"] == "Yes").sum())
    n_paid  = int(((seg_df["CallerAI Qualified"] == "Yes") &
                   (seg_df["Paid Check"] == "Paid")).sum())
    print(f"   {seg_nm:<30} {n_total:>5}  (incl. {n_paid} Paid)")
_ans_c = int(out_all["Disconnect Reason"].fillna("").str.strip().str.lower()
             .isin({r.lower() for r in CALLERAI_ANSWERED_REASONS}).sum())
print(f"   {'AI Answered':<30} {_ans_c:>5}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 8 — Save Leads_Summary_Report.xlsx  (local artifact for GHA)
# ══════════════════════════════════════════════════════════════════════════════
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
wb.save(SUMMARY_LOCAL)
print(f"\nOK Leads_Summary_Report.xlsx -> {SUMMARY_LOCAL}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 9 — Build Leads_Summary_Matrix.xlsx
# ══════════════════════════════════════════════════════════════════════════════
wb_m = openpyxl.Workbook()
ws_s = wb_m.active
ws_s.title = "Summary"
ws_s.sheet_view.showGridLines = False

nr = write_summary_matrix(ws_s, f"Today  {today_str}", summary_today,   1, 1)
nr += 2
nr = write_summary_matrix(ws_s, "Overall", summary_overall, nr, 1)
nr += 2

nr_1 = write_ai_qual_funnel_table(ws_s, "Overall AI Qualified Funnel",
                                  t_aiq_all,         nr, 1)
nr_2 = write_ai_qual_funnel_table(ws_s, "Winback AI Qualified Funnel",
                                  t_aiq_win,         nr, 5)
nr_3 = write_ai_qual_funnel_table(ws_s, "TTC+Uncalled AI Qualified Funnel",
                                  t_aiq_ttc_unc,     nr, 9)
nr_4 = write_ai_qual_funnel_table(ws_s, "TTC AI Qualified Funnel",
                                  t_aiq_ttc,         nr, 13)
nr_5 = write_ai_qual_funnel_table(ws_s, "Renewal_TTC AI Qualified Funnel",
                                  t_aiq_renewal_ttc, nr, 17)
nr_6 = write_ai_qual_funnel_table(ws_s, "Broker AI Qualified Funnel",   # v22
                                  t_aiq_broker,      nr, 21)
nr   = max(nr_1, nr_2, nr_3, nr_4, nr_5, nr_6) + 2

nr = write_planwise_block(
    ws_s, f"Planwise Today  {today_str}",
    build_planwise_rows(out_win_today, out_ttc_unc_today,
                        out_ttc_today, out_renewal_ttc_today,
                        out_broker_today), nr, 1)   # v22: +out_broker_today
nr += 2
nr = write_planwise_block(
    ws_s, "Planwise Overall",
    build_planwise_rows(out_win, out_ttc_unc, out_ttc, out_renewal_ttc,
                        out_broker), nr, 1)          # v22: +out_broker
nr += 2

nr = write_plan_revenue_table(
    ws_s, "Overall Plan-wise Revenue (All Campaigns)", t_plan_rev, nr, 1)
nr += 2
write_plan_revenue_table(
    ws_s, f"Today's Plan-wise Revenue ({today_str})", t_plan_rev_today, nr, 1)

for col, w in {"A": 22, "B": 14, "C": 16, "D": 14, "E": 14, "F": 14, "G": 14,
               "H": 14, "I": 14, "J": 14, "K": 14, "L": 14, "M": 14, "N": 14,
               "O": 14, "P": 14, "Q": 14, "R": 14, "S": 14, "T": 14,
               "U": 14, "V": 14, "W": 14}.items():  # v22
    ws_s.column_dimensions[col].width = w

# Matrix detail tabs
base_df = out.drop(columns=["_date_ts"], errors="ignore").copy()
_is_purch   = base_df["Paid Check"] == "Paid"
_is_aiq     = base_df["CallerAI Qualified"] == "Yes"
_has_nodisp = base_df["Calling Outcome"].isna()
_is_ttcsrc  = base_df["Source"] == TTC_SOURCE

df_purchased_m  = base_df[_is_purch].copy()
df_qualified_m  = base_df[~_is_purch & _is_aiq].copy()
df_ttc_m        = base_df[~_is_purch & ~_is_aiq &
                           (base_df["Calling Outcome"] == "Trying to Contact")].copy()
df_unanswered_m = base_df[~_is_purch & ~_is_aiq & _has_nodisp].copy()
df_uncalled_m   = base_df[~_is_purch & ~_is_aiq & _has_nodisp & _is_ttcsrc].copy()

def write_matrix_detail_sheet(wb, sheet_title, df_sub, header_title):
    ws = wb.create_sheet(title=sheet_title[:31])
    ws.sheet_view.showGridLines = False
    df_s = df_sub.copy()
    for col in ["Calling Outcome", "CallerAI Qualified", "Business Disposition",
                "Business Disposition Confidence", "Lead Owner",
                "Activity Added By", "Plan Name", "Disconnect Reason"]:
        if col in df_s.columns: df_s[col] = df_s[col].fillna("")
    if "Revenue" in df_s.columns:
        df_s["Revenue"] = pd.to_numeric(df_s["Revenue"], errors="coerce").fillna(0)

    sort_keys, sort_asc = [], []
    if "Paid Check" in df_s.columns:
        df_s["_ps"] = df_s["Paid Check"].map({"Paid": 0, "Unpaid": 1}).fillna(1)
        sort_keys.append("_ps"); sort_asc.append(True)
    if "CallerAI Qualified" in df_s.columns:
        df_s["_qs"] = df_s["CallerAI Qualified"].map({"Yes": 0, "No": 1}).fillna(1)
        sort_keys.append("_qs"); sort_asc.append(True)
    if "Date" in df_s.columns:
        sort_keys.append("Date"); sort_asc.append(False)
    if sort_keys:
        df_s.sort_values(sort_keys, ascending=sort_asc, inplace=True)
    df_s.drop(columns=[c for c in ["_ps", "_qs"] if c in df_s.columns],
              inplace=True)
    df_s.reset_index(drop=True, inplace=True)

    paid_c = (int((df_s["Paid Check"] == "Paid").sum())
               if "Paid Check" in df_s.columns else 0)
    ai_q_c = (int((df_s["CallerAI Qualified"] == "Yes").sum())
               if "CallerAI Qualified" in df_s.columns else 0)
    cols_use = [c for c in DETAIL_COLS if c in df_s.columns]
    w_map    = {c: w for c, w in zip(DETAIL_COLS, DETAIL_WIDTHS)}

    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=max(len(cols_use), 1))
    tc = ws.cell(1, 1)
    tc.value = (f"{header_title}  --  Total: {len(df_s)}  |  Paid: {paid_c}  |  "
                f"AI Qualified: {ai_q_c}  |  {src_count_str(df_s)}")
    tc.font = Font(bold=True, color="FFFFFF"); tc.fill = TITLE_FILL
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    for ci, h in enumerate(cols_use):
        hfill = (AI_COL_FILL
                 if h in ("CallerAI Qualified",
                           "Business Disposition Confidence",
                           "Business Disposition")
                 else HEADER_FILL)
        _cell(ws, 2, ci + 1, h, bold=True, fill=hfill)

    for ri, (_, row) in enumerate(df_s[cols_use].iterrows(), start=3):
        is_paid  = (str(row.get("Paid Check", "")) == "Paid")
        is_ai_q  = (str(row.get("CallerAI Qualified", "")) == "Yes")
        row_fill = (PAID_FILL   if is_paid
                    else AI_QUAL_FILL if is_ai_q
                    else seg_fills.get(row.get("Source", ""), None))
        for ci, col in enumerate(cols_use):
            cell = ws.cell(ri, ci + 1)
            cell.value = row[col]; cell.font = Font(bold=is_paid)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = BOX
            if row_fill: cell.fill = row_fill

    for ci, col in enumerate(cols_use, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(ci)].width = w_map.get(col, 16)
    return ws

write_matrix_detail_sheet(wb_m, "Purchased",
                          df_purchased_m,  "Purchased Leads")
write_matrix_detail_sheet(wb_m, "Qualified",
                          df_qualified_m,  "AI Qualified (excl. Purchased)")
write_matrix_detail_sheet(wb_m, "Trying to Contact",
                          df_ttc_m,        "Trying to Contact (excl. Purchased & AI Qualified)")
write_matrix_detail_sheet(wb_m, "Unanswered",
                          df_unanswered_m, "Unanswered -- No Disposition (excl. Purchased & AI Qualified)")
write_matrix_detail_sheet(wb_m, "Uncalled",
                          df_uncalled_m,   "Uncalled -- TTC Campaign, No LSQ Disposition")

wb_m.save(SUMMARY_MATRIX_LOCAL)
print(f"OK Leads_Summary_Matrix.xlsx -> {SUMMARY_MATRIX_LOCAL}")
print(f"   Purchased ({len(df_purchased_m)})  |  AI Qualified ({len(df_qualified_m)})  |  "
      f"TTC ({len(df_ttc_m)})  |  Unanswered ({len(df_unanswered_m)})  |  "
      f"Uncalled ({len(df_uncalled_m)})")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 10 — Write Lead Summary Report to Google Sheet  (v21 new step)
# ══════════════════════════════════════════════════════════════════════════════
write_summary_to_gsheet(
    gc, summary_today, summary_overall, today_str,
    t_aiq_all, t_aiq_win, t_aiq_ttc_unc, t_aiq_ttc, t_aiq_renewal_ttc,
    t_aiq_broker,   # v22
    t_plan_rev
)


# ══════════════════════════════════════════════════════════════════════════════
# Final console summary
# ══════════════════════════════════════════════════════════════════════════════
print(f"\n-- Final Summary --------------------------------------------------------")
print(f"   Today  {today_str}:")
for row in summary_today:
    rev = row.get("AI Qual & Paid Rev", 0) + row.get("AI Unqual & Paid Rev", 0)
    print(f"     {row['Segment']:<22}: {row['Attempted']:>6} attempted | "
          f"{row['AI Qualified']:>5} AI Qual | {row['Paid']:>5} paid | "
          f"Rs {rev:,} attributed")
# Grand total today
gt_t = build_grand_total_row(summary_today)
rev_t = gt_t.get("AI Qual & Paid Rev", 0) + gt_t.get("AI Unqual & Paid Rev", 0)
print(f"     {'GRAND TOTAL':<22}: {gt_t['Attempted']:>6} attempted | "
      f"{gt_t['AI Qualified']:>5} AI Qual | {gt_t['Paid']:>5} paid | "
      f"Rs {rev_t:,} attributed")

print(f"   Overall:")
for row in summary_overall:
    rev = row.get("AI Qual & Paid Rev", 0) + row.get("AI Unqual & Paid Rev", 0)
    print(f"     {row['Segment']:<22}: {row['Attempted']:>6} attempted | "
          f"{row['AI Qualified']:>5} AI Qual | {row['Paid']:>5} paid | "
          f"Rs {rev:,} attributed")
# Grand total overall
gt_o = build_grand_total_row(summary_overall)
rev_o = gt_o.get("AI Qual & Paid Rev", 0) + gt_o.get("AI Unqual & Paid Rev", 0)
print(f"     {'GRAND TOTAL':<22}: {gt_o['Attempted']:>6} attempted | "
      f"{gt_o['AI Qualified']:>5} AI Qual | {gt_o['Paid']:>5} paid | "
      f"Rs {rev_o:,} attributed")

print("\nDone!")
